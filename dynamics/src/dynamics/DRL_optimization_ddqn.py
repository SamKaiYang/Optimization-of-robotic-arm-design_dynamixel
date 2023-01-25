#!/usr/bin/env python3
# coding: utf-8
import importlib
import sys

import rospy

importlib.reload(sys)
import argparse
import csv
import math
import time
from collections import namedtuple
from math import pi
from os import path

import geometry_msgs.msg
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import plotly.graph_objs as go
import plotly.offline as py
import roboticstoolbox as rtb
import sympy as sp
# import dyna_space
from interface_control.msg import (cal_cmd, cal_process, cal_result, dyna_data,
                                dyna_space_data, optimal_design,
                                optimal_random, specified_parameter_design, tested_model_name)
from matplotlib import cm
from moveit_msgs.msg import DisplayTrajectory, RobotTrajectory
from mpl_toolkits.mplot3d import Axes3D
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, colors
from openpyxl.utils import get_column_letter
from plotly.offline import download_plotlyjs, iplot, plot
from scipy.interpolate import make_interp_spline  # draw smooth
from spatialmath import *
from trajectory_msgs.msg import JointTrajectory, JointTrajectoryPoint
# from dqn import DQN

np.set_printoptions(
    linewidth=100,
    formatter={"float": lambda x: f"{x:8.4g}" if abs(x) > 1e-10 else f"{0:8.4g}"},
)

from arm_workspace import arm_workspace_plane
# from robot_urdf import RandomRobot
from motor_module import motor_data
from random_robot import RandomRobot
from modular_robot_6dof import modular_robot_6dof
# DRL_optimization api
import sys
import os
import torch.nn as nn
import torch.nn.functional as F
curr_path = os.path.dirname(os.path.abspath(__file__))  # 当前文件所在绝对路径
parent_path = os.path.dirname(curr_path)  # 父路径
sys.path.append(parent_path)  # 添加路径到系统路径

import gym
import torch
import datetime
import numpy as np
from common.utils import save_results, make_dir
from common.utils import plot_rewards
from dqn import DQN

###### DDQN #############
# from ./pytorch-DQN/config import Config
import argparse
import random
# import gym
# import torch
from torch.optim import Adam
from pytorch_DQN.tester import Tester
from pytorch_DQN.buffer import ReplayBuffer
from pytorch_DQN.config import Config
from pytorch_DQN.core.util import get_class_attr_val
from pytorch_DQN.model import DQN
from pytorch_DQN.trainer import Trainer
from pytorch_DQN.config import Config
from pytorch_DQN.ddqn import DDQNAgent
from pytorch_DQN.core.util import get_output_folder

import matplotlib.pyplot as plt
from RobotOptEnv_dynamixel import RobotOptEnv
import tensorboardX
import yaml
file_path = curr_path + "/outputs/" 
curr_time = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")  # 获取当前时间

tb = tensorboardX.SummaryWriter()
# tensorboard_callback = tensorboardX.(log_dir=log_dir, histogram_freq=1)
algo_name = "DDQN"  # 算法名称
env_name = 'DDQN_RobotOptEnv'  # 环境名称



class drl_optimization:
    def __init__(self):
        # self.test = 0
        self.robot = modular_robot_6dof()
        self.env = RobotOptEnv()
        self.config = Config()
    def env_agent_config(self, cfg, seed=1):
        ''' 创建环境和智能体
        '''
        # env = gym.make(self.env)  # 创建环境
        env = self.env  # 创建环境
        env.seed(seed)  # 设置随机种子
        n_states = env.observation_space.shape[0]  # 状态维度
        rospy.loginfo("n_states: {}".format(n_states))
        n_actions = env.action_space.n  # 动作數量
        rospy.loginfo("n_actions: {}".format(n_actions))
        
        
        self.config.env = env_name
        self.config.gamma = 0.99 # 强化学习中的折扣因子
        self.config.epsilon = 1 # e-greedy策略中初始epsilon
        self.config.epsilon_min = 0.01 # e-greedy策略中的终止epsilon
        self.config.eps_decay = 500 # e-greedy策略中epsilon的衰减率
        self.config.frames = 160000 # 经验回放的容量
        self.config.use_cuda = True
        self.config.learning_rate = 1e-2  # 学习率 # TODO: fixed 0117 11:45 1e-1 -> 1e-2
        self.config.max_buff = 1000
        self.config.update_tar_interval = 100
        self.config.batch_size = 128 # mini-batch SGD中的批量大小
        self.config.print_interval = 50 # TODO: fixed 0117 11:45
        self.config.log_interval = 50 # TODO: fixed 0117 11:45
        self.config.win_reward = 100     # TODO: fixed 0116
        self.config.win_break = True

        self.config.action_dim = n_actions
        self.config.state_dim = n_states
        # model = MLP(n_states,n_actions)
        agent = DDQNAgent(self.config)  # 创建智能体
        return env, agent
    

class PlotConfig:
    ''' 绘图相关参数设置
    '''

    def __init__(self) -> None:
        self.algo_name = algo_name  # 算法名称
        self.env_name = env_name  # 环境名称
        self.device = torch.device(
            "cuda" if torch.cuda.is_available() else "cpu")  # 检测GPU
        self.result_path = curr_path + "/outputs/" + self.env_name + \
            '/' + curr_time + '/results/'  # 保存结果的路径
        self.model_path = curr_path + "/outputs/" + self.env_name + \
            '/' + curr_time + '/models/'  # 保存模型的路径
        self.save = True  # 是否保存图片

class RosTopic:
    def __init__(self):
        self.sub_taskcmd = rospy.Subscriber("/cal_command", cal_cmd, self.cmd_callback)
        self.sub_test_model_name = rospy.Subscriber("/tested_model_name", tested_model_name, self.test_model_name_callback)
        self.cmd_run = 0
    def cmd_callback(self, data):
        self.cmd = data.cmd
        rospy.loginfo("I heard command is %s", data.cmd)
        if data.cmd == 22:
            rospy.sleep(10)
            self.cmd_run = 1
        if data.cmd == 23:
            rospy.sleep(10)
            self.cmd_run = 2
    def test_model_name_callback(self, data):
        self.test_model_name  = data.tested_model_name
        print(self.test_model_name)


class Trainer:
    def __init__(self, agent, env, config: Config, model_path):
        self.agent = agent
        self.env = env
        self.config = config
        self.env_name = env_name  # 环境名称
        # non-Linear epsilon decay
        epsilon_final = self.config.epsilon_min
        epsilon_start = self.config.epsilon
        epsilon_decay = self.config.eps_decay
        self.epsilon_by_frame = lambda frame_idx: epsilon_final + (epsilon_start - epsilon_final) * math.exp(
            -1. * frame_idx / epsilon_decay)


        self.outputdir = model_path
        # # self.outputdir = get_output_folder(self.config.output, self.config.env)
        # self.outputdir = curr_path + "/outputs/" + self.env_name + \
        #     '/' + curr_time + '/models/'  # 保存模型的路径
        self.agent.save_config(self.outputdir)
        
        # self.board_logger = TensorBoardLogger(self.outputdir)

    def train(self, pre_fr=0, train_eps = 300):
        losses = []
        all_rewards = []
        episode_reward = 0
        ep_num = 0
        is_win = False

        state = self.env.reset()
        for fr in range(pre_fr + 1, self.config.frames + 1):
            epsilon = self.epsilon_by_frame(fr)
            action = self.agent.act(state, epsilon)

            next_state, reward, done, info = self.env.step(action)
            self.agent.buffer.add(state, action, reward, next_state, done)

            state = next_state
            episode_reward += reward
            # rospy.loginfo("aaaa")
            loss = 0
            if self.agent.buffer.size() > self.config.batch_size:
                loss = self.agent.learning(fr)
                losses.append(loss)
                # self.board_logger.scalar_summary('Loss per frame', fr, loss)
                tb.add_scalar("/trained-model/Loss_per_frame/", loss, fr)
                
            if fr % self.config.print_interval == 0:
                # print("frames: %5d, reward: %5f, loss: %4f episode: %4d" % (fr, np.mean(all_rewards[-10:]), loss, ep_num))
                rospy.loginfo('frames: {}, reward: {}, loss: {} episode: {}'.format(fr, np.mean(all_rewards[-10:]), loss, ep_num))
            if fr % self.config.log_interval == 0:
                # self.board_logger.scalar_summary('Reward per episode', ep_num, all_rewards[-1])
                tb.add_scalar("/trained-model/Reward_per_episode/", all_rewards[-1], ep_num)
            if self.config.checkpoint and fr % self.config.checkpoint_interval == 0:
                self.agent.save_checkpoint(fr, self.outputdir)

            if done:
                state = self.env.reset()
                all_rewards.append(episode_reward)
                episode_reward = 0
                ep_num += 1
                avg_reward = float(np.mean(all_rewards[-100:]))
                # self.board_logger.scalar_summary('Best 100-episodes average reward', ep_num, avg_reward)
                tb.add_scalar("/trained-model/Best_100_episodes_average_reward/", avg_reward, ep_num)
                if len(all_rewards) >= 100 and ep_num>=train_eps:
                    self.agent.save_model(self.outputdir, 'max_eps')
                    rospy.loginfo('Ran {} episodes max {}-episodes average reward is {}. Solved after {} trials ✔'.format(ep_num, train_eps, avg_reward, ep_num - 100))
                    if self.config.win_break:
                        break
                if len(all_rewards) >= 100 and avg_reward >= self.config.win_reward and all_rewards[-1] > self.config.win_reward:
                    is_win = True
                    self.agent.save_model(self.outputdir, 'best')
                    # print('Ran %d episodes best 100-episodes average reward is %3f. Solved after %d trials ✔' % (ep_num, avg_reward, ep_num - 100))
                    rospy.loginfo('Ran {} episodes best 100-episodes average reward is {}. Solved after {} trials ✔'.format(ep_num, avg_reward, ep_num - 100))
                    if self.config.win_break:
                        break

        if not is_win:
            print('Did not solve after %d episodes' % ep_num)
            self.agent.save_model(self.outputdir, 'last')
class Tester(object):

    def __init__(self, agent, env, model_path, num_episodes=50, max_ep_steps=400, test_ep_steps=100):
        self.num_episodes = num_episodes
        self.max_ep_steps = max_ep_steps
        self.test_ep_steps = test_ep_steps
        self.agent = agent
        self.env = env
        self.agent.is_training = False
        self.agent.load_weights(model_path)
        self.policy = lambda x: agent.act(x)
        self.xlsx_outpath = "./xlsx/"

    def test(self, debug=True, visualize=True): # debug = true
        excel_file = Workbook()
        sheet = excel_file.active
        i = 0
        avg_reward = 0
        for episode in range(self.num_episodes):
            s0 = self.env.tested_reset() # TODO: 改為陣列儲存
            episode_steps = 0
            episode_reward = 0.

            done = False
            while not done:
                if visualize:
                    self.env.render()

                action = self.policy(s0)
                s0, reward, done, info = self.env.step(action)
                
                episode_reward += reward
                episode_steps += 1


                

                if episode_steps + 1 > self.test_ep_steps:
                    done = True
                # TODO: 獲取分數高於...的reward結果,
            if debug:
                # TODO: 獲取分數高於...的reward結果,
                if episode_reward >= 100:
                    sheet.cell(row=i + 1, column=1).value = s0[0]
                    sheet.cell(row=i + 1, column=2).value = s0[1]
                    sheet.cell(row=i + 1, column=3).value = s0[2]
                    sheet.cell(row=i + 1, column=4).value = s0[3]
                    sheet.cell(row=i + 1, column=5).value = s0[4]
                    sheet.cell(row=i + 1, column=6).value = s0[5]
                    sheet.cell(row=i + 1, column=7).value = s0[6]
                    sheet.cell(row=i + 1, column=8).value = s0[7]
                    sheet.cell(row=i + 1, column=9).value = s0[8]
                    sheet.cell(row=i + 1, column=10).value = s0[9]
                    sheet.cell(row=i + 1, column=11).value = info[0] #axis 2
                    sheet.cell(row=i + 1, column=12).value = info[1] #axis 3
                    sheet.cell(row=i + 1, column=13).value = info[2] #motor 2
                    sheet.cell(row=i + 1, column=14).value = info[3] #motor 3
                    sheet.cell(row=i + 1, column=15).value = episode_reward
                    i = i + 1
                # print('[Test] episode: %3d, episode_reward: %5f' % (episode, episode_reward))
                rospy.loginfo('[Test] episode: {}, episode_reward: {}'.format(episode, episode_reward))
                tb.add_scalar("/tested-model/test_reward/", episode_reward, episode)
            avg_reward += episode_reward
        avg_reward /= self.num_episodes
        # print("avg reward: %5f" % (avg_reward))
        rospy.loginfo('avg reward: {}'.format(avg_reward))

        file_name = self.xlsx_outpath + "/tested_reward_state" +".xlsx"
        excel_file.save(file_name)

if __name__ == "__main__":
    rospy.init_node("optimization")
    a = 0
    # cfg = DQNConfig()
    cfg = 0
    drl = drl_optimization()
    plot_cfg = PlotConfig()
    ros_topic = RosTopic()
    ddqn_train_eps = 1000  # 训练的回合数
    ddqn_test_eps = 100  # 测试的回合数
    # train = Trainer()
    while not rospy.is_shutdown():
        # test all
        if ros_topic.cmd_run == 1:
            ros_topic.cmd_run = 0
            make_dir(plot_cfg.result_path, plot_cfg.model_path)  # 创建保存结果和模型路径的文件夹
            # 訓練
            drl.env.model_select = "train"
            drl.env.point_Workspace_cal_Monte_Carlo()
            train_env, train_agent = drl.env_agent_config(cfg, seed=1)
            train = Trainer(train_agent, train_env, drl.config, plot_cfg.model_path)
            train.train(train_eps = ddqn_train_eps)
            # 測試
            drl.env.model_select = "test"
            plot_cfg.model_path = plot_cfg.model_path +'model_last.pkl'
            test_env, test_agent = drl.env_agent_config(cfg, seed=10)
            test = Tester(test_agent, test_env, plot_cfg.model_path, num_episodes = ddqn_test_eps)
            test.test()
            break
        else:
            pass
        '''
        # test trained_model
        if ros_topic.cmd_run == 1:
            ros_topic.cmd_run = 0
            make_dir(plot_cfg.result_path, plot_cfg.model_path)  # 创建保存结果和模型路径的文件夹
            # 訓練
            train_env, train_agent = drl.env_agent_config(cfg, seed=1)
            train = Trainer(train_agent, train_env, drl.config, plot_cfg.model_path)
            train.train(train_eps = ddqn_train_eps)
            break
        else:
            pass

        '''
        # test tested_model
        if ros_topic.cmd_run == 2:
            ros_topic.cmd_run = 0
            # 測試
            drl.env.model_select = "test"
            plot_cfg.model_path = '/home/iclab/Documents/drl_robotics_arm_ws/src/Optimization-of-robotic-arm-design/dynamics/src/dynamics/outputs/DDQN_RobotOptEnv/'+ str(ros_topic.test_model_name) +'/models/model_best.pkl'# test 20230102
            test_env, test_agent = drl.env_agent_config(cfg, seed=10)
            test = Tester(test_agent, test_env, plot_cfg.model_path, test_ep_steps = ddqn_test_eps)
            test.test()
            break
        else:
            pass
