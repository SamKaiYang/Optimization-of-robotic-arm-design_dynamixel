#!/usr/bin/env python3
# coding: utf-8
import gym
from gym import spaces
import numpy as np
from math import pi

from sympy import false
# from torch import R
from interface_control.msg import optimal_design
from dynamics.arm_workspace import arm_workspace_plane
# from robot_urdf import RandomRobot
from dynamics.motor_module import motor_data
from dynamics.modular_robot_3dof import modular_robot_3dof
from dynamics.modular_robot_4dof import modular_robot_4dof
from dynamics.modular_robot_5dof import modular_robot_5dof
from dynamics.modular_robot_6dof import modular_robot_6dof
# from dynamics.stl_conv_6dof_urdf import stl_conv_urdf
from dynamics.stl_conv_6dof_urdf_dynamixel import stl_conv_urdf
import rospy
# one-hot encoder
from sklearn import preprocessing
import pandas as pd
import yaml
# add 可達性可操作性評估
import numpy as np
from roboticstoolbox import DHRobot, RevoluteDH
from spatialmath import SE3
import spatialmath as sm
from urdf_parser_py.urdf import URDF
import os
from openpyxl import load_workbook
from openpyxl import Workbook

# TODO: 初版 只考慮 6 dof 機器人的關節長度變化, 觀察各軸馬達極限之輸出最大torque值
class RobotOptEnv(gym.Env):
    metadata = {
        'render.modes': ['human', 'rgb_array'],
        'video.frames_per_second': 2
    }
    def __init__(self):
        self.robot = modular_robot_6dof()
        self.robot_urdf = stl_conv_urdf("single_arm_v12","test")
        
        # callback:Enter the parameters of the algorithm to be optimized on the interface
        self.sub_optimal_design = rospy.Subscriber(
            "/optimal_design", optimal_design, self.optimal_design_callback
        )
        # 使用者設定參數
        self.payload = 5.0
        self.payload_position = np.array([0, 0, 0.04])
        self.vel = np.array([2.356194, 2.356194, 2.356194, 2.356194, 2.356194, 2.356194])
        self.acc = np.array([2.356194, 2.356194, 2.356194, 2.356194, 2.356194, 2.356194])
        self.total_weight = 20 # Kg
        self.total_cost = 1800 # 元
        # 預設二,三軸軸長
        self.std_L2 = 35.0 # 預設標準值 第二軸 35 cm
        self.std_L3 = 35.0 # 預設標準值 第三軸 35 cm
        # 觀察參數 motor
        self.motor = motor_data()
        self.res = self.motor.dynamixel_member
        self.high_torque = float('inf') # 預設標準值 馬達極限 120.0 N max
        self.low_torque = float('-inf') # 預設標準值 馬達極限 -12.0 N max
        self.motor_cost_init = np.array([0,400,400,0,0,0], dtype=np.float64) # 預設最大馬達費用
        self.motor_weight_init = np.array([0,0.855,0.855,0,0,0], dtype=np.float64) # 預設最大馬達重量
        self.motor_cost = np.array([0,400,400,0,0,0], dtype=np.float64) # 馬達費用
        self.motor_weight = np.array([0,0.855,0.855,0,0,0], dtype=np.float64) # 馬達重量
        self.motor_rated = np.array([44.7,44.7,44.7,44.7,44.7,44.7], dtype=np.float64)
        # 使用者設定參數 & 觀察參數
        self.reach_distance = 0.6 # 使用者設定可達半徑最小值
        self.high_reach_eva = 1 # 預設觀測標準值
        self.low_reach_eva = 0 # 預設觀測標準值
        self.high_manipulability = 1  # 預設觀測標準值
        self.low_manipulability = 0  # 預設觀測標準值
        self.high_std_L2 = 70 # 預設觀測標準值
        self.low_std_L2 = -25 # 預設觀測標準值
        self.high_std_L3 = 70 # 預設觀測標準值
        self.low_std_L3 = -25 # 預設觀測標準值
        self.torque_done = np.array([false, false, false, false, false, false])
        self.torque_over = False
        self.prev_shaping = None
        # TODO: 增加馬達模組選型action
        self.action_space = spaces.Discrete(12) # TODO: fixed 12種action
        
        # TODO: observation space for torque, reach, motor cost, weight, manipulability
        self.observation_space = spaces.Box(np.array([self.low_torque,self.low_torque,self.low_torque,self.low_torque,self.low_torque,self.low_torque, self.low_reach_eva, self.low_manipulability, self.low_std_L2, self.low_std_L3 ]), 
                                            np.array([self.high_torque,self.high_torque,self.high_torque,self.high_torque,self.high_torque,self.high_torque, self.high_reach_eva, self.high_manipulability, self.high_std_L2, self.high_std_L3]), 
                                            dtype=np.float64)
        # TODO: reward 歸一化
        self.state = np.array([0,0,0,0,0,0,0,0,0,0], dtype=np.float64)
        self.pre_state = np.array([0,0,0,0,0,0,0,0,0,0], dtype=np.float64)

        #隨機抽樣點位初始化
        self.T_x = []
        self.T_y = []
        self.T_z = []
        self.T_roll = []
        self.T_pitch = []
        self.T_yaw = []
        self.xlsx_outpath = "./xlsx/"

        self.model_select = "train" # 選擇使用train or test model 

    def optimal_design_callback(self, data):
        # print(data.data)
        # TODO: 目標構型
        self.op_dof = data.dof
        self.op_payload = data.payload
        self.op_payload_position = data.payload_position
        self.op_vel = data.vel
        self.op_acc = data.acc
        self.op_radius = data.radius
        self.op_weight = data.arm_weight
        self.op_cost = data.cost
        
        print("op_dof:", self.op_dof)
        print("op_payload:", self.op_payload)
        print("op_payload_position:", self.op_payload_position)
        print("op_vel:", self.op_vel)
        print("op_acc:", self.op_acc)
        print("op_radius:", self.op_radius)
        print("op_weight:", self.op_weight)
        print("op_cost:", self.op_cost)
        
        self.payload = self.op_payload
        self.payload_position = np.array(self.op_payload_position)
        self.vel = np.array(self.op_vel[0:6])
        self.acc = np.array(self.op_acc[0:6])
        self.total_weight = self.op_weight # Kg
        self.total_cost = self.op_cost # 元
        self.reach_distance = self.op_radius # 使用者設定可達半徑最小值
        
    # TODO: fixed
    def step(self, action):
        assert self.action_space.contains(action), "%r (%s) invalid"%(action, type(action))
        self.pre_state[0:6] = self.state[0:6]
        # TODO: 向量編碼動作
        if action == 0: # 軸2  # 短 # 型號1
            self.std_L2 -= 1.0
            # 配置軸2 motor 型號1
            motor_type = 0 # 型號
            axis = 2 # 軸數
            
        elif action == 1: # 軸2  # 長 # 型號1
            self.std_L2 += 1.0
            motor_type = 0 # 型號
            axis = 2 # 軸數
            
        elif action == 2: # 軸2  # 短 # 型號2
            self.std_L2 -= 1.0
            motor_type = 1 # 型號
            axis = 2 # 軸數
            
        elif action == 3: # 軸2  # 長 # 型號2
            self.std_L2 += 1.0
            motor_type = 1 # 型號
            axis = 2 # 軸數
            
        elif action == 4: # 軸2  # 短 # 型號3
            self.std_L2 -= 1.0
            motor_type = 2 # 型號
            axis = 2 # 軸數
            
        elif action == 5: # 軸2  # 長 # 型號3
            self.std_L2 += 1.0
            motor_type = 2 # 型號
            axis = 2 # 軸數
            
        elif action == 6: # 軸3  # 短 # 型號1
            self.std_L3 -= 1.0
            motor_type = 0 # 型號
            axis = 3 # 軸數
            
        elif action == 7: # 軸3  # 長 # 型號1
            self.std_L3 += 1.0
            motor_type = 0 # 型號
            axis = 3 # 軸數
            
        elif action == 8: # 軸3  # 短 # 型號2
            self.std_L3 -= 1.0
            motor_type = 1 # 型號
            axis = 3 # 軸數
            
        elif action == 9: # 軸3  # 長 # 型號2
            self.std_L3 += 1.0
            motor_type = 1 # 型號
            axis = 3 # 軸數
            
        elif action == 10: # 軸3  # 短 # 型號3
            self.std_L3 -= 1.0
            motor_type = 2 # 型號
            axis = 3 # 軸數
            
        elif action == 11: # 軸3  # 長 # 型號3
            self.std_L3 += 1.0
            motor_type = 2 # 型號
            axis = 3 # 軸數

        # 輸入action後 二,三軸軸長
        self.robot_urdf.specified_generate_write_urdf(self.std_L2, self.std_L3)
        self.robot.__init__() # 重製機器人
        torque = self.dynamics_torque_limit()
        self.state[0:6] = torque
        # # 可達性
        # self.state[6] = self.reach_evaluate()
        # 計算成本與重量    
        self.motor_cost[axis-1] = self.res.cost[motor_type]
        self.motor_weight[axis-1] = self.res.weight[motor_type]
        # print(self.motor_cost)
        # print(self.motor_weight)
        cost = sum(self.motor_cost) 
        weight = sum(self.motor_weight)
        # self.state[7] = cost
        # self.state[8] = weight
        
        # self.state[9] = self.manipulability_evaluate()
        # 可達性 # 可操作性
        self.state[6], self.state[7] = self.reach_manipulability_evaluate(self.model_select)
        # rospy.loginfo("reach_score: %s", self.state[6])
        # rospy.loginfo("manipulability_score: %s", self.state[9])
        self.state[8] = self.std_L2
        self.state[9] = self.std_L3
        self.motor_rated[axis-1] = self.res.rated_torque[motor_type]
        # rospy.loginfo("configuration cost & weight: %s, %s", cost, weight)
        self.counts += 1
        # rospy.loginfo("self.state: %s", self.state)
        reward = 0

        
        # TODO: fixed
        shaping = (
            -1 * np.sqrt(self.state[0] * self.state[0] + self.state[1] * self.state[1] + self.state[2] * self.state[2] + self.state[3] * self.state[3] + self.state[4] * self.state[4] + self.state[5] * self.state[5])
            + 100 * self.state[6] # 可達性
            # - 0.01 * self.state[7] # cost
            # - 10 * self.state[8] # weight
            + 1000 * self.state[7] # 可操作性
        ) 

        if self.prev_shaping is not None:
            reward = shaping - self.prev_shaping
        self.prev_shaping = shaping

        # 判斷超出最大扭矩
        for i in range(6):
            # TODO:consider cost & weight 
            # if np.abs(self.state[i]) > self.motor_rated[i]:
            if np.abs(self.state[i]) > 44.7: # TODO: fixed 
                self.torque_over = True # 超出最大扭矩
                break # TODO

        terminated = False
        
        # 避免軸長小於0
        if self.state[8] <= 0 or self.state[9] <=  0:
            # terminated = True
            reward = -1000
        if self.torque_over == True and self.state[6] < 0.6: 
            reward += -20
        elif self.torque_over == True and 0.6 <= self.state[6] < 0.8: 
            reward += -10
        elif self.torque_over == True and 0.8 <= self.state[6] <= 1.0:   
            reward += -5
        elif self.torque_over == False and self.state[6] < 0.6: 
            reward += +5
        elif self.torque_over == False and 0.6 <= self.state[6] < 0.8:
            reward += +10
        elif self.torque_over == False and 0.8 <= self.state[6] < 0.9:
            reward += +20
        elif self.torque_over == False and 0.9 <= self.state[6] < 1.0:
            terminated = True
            reward += +50
        elif self.torque_over == False and self.state[6] == 1.0:
            terminated = True
            reward += +100
        
        if self.counts == 30:
            if self.state[6] < 0.6: # TODO: fixed
                reward += -50
            else:
                reward += -10
            if self.torque_over == True: # TODO: fixed 0112 00:22 改為超過最大的馬達型號torque
                reward += -100
            else:
                reward += -10
            terminated = True
            self.counts = 0
        

        self.torque_over = False #reset
        rospy.loginfo("counts: %s", self.counts)
        rospy.loginfo("step_reward: %s", reward)
        rospy.loginfo("================================")
        # print("================================")
        current_design = [self.std_L2, self.std_L3, self.motor_rated[1], self.motor_rated[2]]
        return self.state, reward, terminated, current_design

    # reset环境状态 
    def reset(self):
        # random state (手臂長度隨機)
        self.std_L2, self.std_L3 = self.robot_urdf.opt_random_generate_write_urdf() # 啟用隨機的L2,L3長度urdf
        self.robot.__init__() # 重製機器人
        torque = self.dynamics_torque_limit()
        self.state[0:6] = torque
        # 生成隨機 payload (kg)
        rand_payload = np.random.uniform(low=1, high=4)
        self.payload = rand_payload
        # rospy.loginfo("payload: %s", self.payload)
        self.point_Workspace_cal_Monte_Carlo() # 在當前reset出來的機械手臂構型下, 生成點位
        self.random_select_point() # 先隨機抽樣30個點位
        self.prev_shaping = None
        # reach_score = self.reach_evaluate()
        # manipulability_score = self.manipulability_evaluate()
        reach_score, manipulability_score = self.reach_manipulability_evaluate(self.model_select)
        self.state[6] = reach_score
        # self.state[7] = sum(self.motor_cost_init)
        # self.state[8] = sum(self.motor_weight_init)
        self.state[7] = manipulability_score
        self.state[8] = self.std_L2
        self.state[9] = self.std_L3
        self.counts = 0
        return self.state    
    
    # reset环境状态 
    def tested_reset(self):
        # random state (手臂長度隨機)
        self.std_L2, self.std_L3 = self.robot_urdf.opt_random_generate_write_urdf() # 啟用隨機的L2,L3長度urdf
        self.robot.__init__() # 重製機器人
        self.payload = self.op_payload
        self.payload_position = np.array(self.op_payload_position)
        self.vel = np.array(self.op_vel[0:6])
        self.acc = np.array(self.op_acc[0:6])
        self.total_weight = self.op_weight # Kg
        self.total_cost = self.op_cost # 元
        self.reach_distance = self.op_radius # 使用者設定可達半徑最小值

        torque = self.dynamics_torque_limit()
        self.state[0:6] = torque
        self.prev_shaping = None
        # reach_score = self.reach_evaluate()
        # manipulability_score = self.manipulability_evaluate()
        reach_score, manipulability_score = self.reach_manipulability_evaluate(self.model_select)
        self.state[6] = reach_score
        # self.state[7] = sum(self.motor_cost_init)
        # self.state[8] = sum(self.motor_weight_init)
        self.state[7] = manipulability_score
        self.state[8] = self.std_L2
        self.state[9] = self.std_L3
        self.counts = 0
        return self.state    

    # 視覺化呈現，它只會回應出呼叫那一刻的畫面給你，要它持續出現，需要寫個迴圈
    def render(self, mode='human'):
        return None
        
    def close(self):
        return None
        
    def dynamics_torque_limit(self):
        """
        Calculate the maximum torque required by

        each axis when the arm of each axis is the longest and the acceleration is the highest
        """
        torque = np.array([np.zeros(shape=6)])
        # axis_angle = np.array([np.zeros(shape=6)])
        axis_angle = []
        append_torque_limit_list = []
        temp_torque_max = []
        temp_torque_min = []
        Torque_Max = []
        # 角度轉換
        du = pi / 180
        # 度
        # radian = 180 / pi
        # 弧度
        self.robot.payload(self.payload, self.payload_position)  # set payload
        torque = np.array([np.zeros(shape=6)])
        q_list = [0, 90, -90, 180, -180]
        T_cell = (
            len(q_list)
            * len(q_list)
            * len(q_list)
            * len(q_list)
            * len(q_list)
            * len(q_list)
        )

        for i in range(len(q_list)):
            q1 = q_list[i]
            percent = i / T_cell * 100
            for j in range(len(q_list)):
                q2 = q_list[j]
                for k in range(len(q_list)):
                    q3 = q_list[k]
                    for l in range(len(q_list)):
                        q4 = q_list[l]
                        for m in range(len(q_list)):
                            q5 = q_list[m]
                            for n in range(len(q_list)):
                                q6 = q_list[n]
                                axis_angle.append([q1, q2, q3, q4, q5, q6])
                                load = np.array(
                                    [
                                        self.robot.rne(
                                            [
                                                q1 * du,
                                                q2 * du,
                                                q3 * du,
                                                q4 * du,
                                                q5 * du,
                                                q6 * du,
                                            ],
                                            self.vel,
                                            self.acc
                                        )
                                    ]
                                )
                                torque = np.append(torque, load, axis=0)

        for i in range(6):
            axis = i
            toque_max_index = np.argmax(torque[:, axis])
            toque_min_index = np.argmin(torque[:, axis])

            temp_torque_max = torque[toque_max_index].tolist()
            temp_torque_max.extend(axis_angle[toque_max_index])
            temp_torque_min = torque[toque_min_index].tolist()
            temp_torque_min.extend(axis_angle[toque_min_index])
            append_torque_limit_list.append(temp_torque_max)
            append_torque_limit_list.append(temp_torque_min)
            Torque_Max.append(abs(torque[toque_max_index][i]))
            self.torque_dynamics_limit = Torque_Max
        return self.torque_dynamics_limit

    def reach_manipulability_evaluate(self,model_select):
        if model_select == "train":
            # import xlsx
            df = load_workbook("./xlsx/task_point.xlsx")
            sheets = df.worksheets
            sheet1 = sheets[0]
            rows = sheet1.rows
            cols = sheet1.columns
            T_tmp = []
            score = []
            manipulability_index = []
            i = 0
            false_done = False
            count = 0
            for row in rows:
                row_val = [col.value for col in row]
                T_tmp.append(SE3(row_val[0], row_val[1], row_val[2]) * SE3.RPY([np.deg2rad(row_val[3]), np.deg2rad(row_val[4]), np.deg2rad(row_val[5])]))
                ik_q = self.robot.ikine_LMS(T=T_tmp[i])
                if ik_q.success == True:
                    count += 1
                    manipulability_index.append(self.robot.manipulability(q=ik_q.q))
                i = i + 1

            if count == 0:
                return(0,0)
            else:
                final_score = count / i
                if count == 1:
                    return(final_score, manipulability_index[0]) # 回傳 manipulability[0]
                else:
                    return(final_score, np.mean(manipulability_index)) # 回傳 manipulability 取平均
        elif model_select == "test":
            # import xlsx
            df = load_workbook("./xlsx/task_point_6dof_tested.xlsx")
            sheets = df.worksheets
            sheet1 = sheets[0]
            rows = sheet1.rows
            cols = sheet1.columns
            T_tmp = []
            score = []
            manipulability_index = []
            i = 0
            false_done = False
            count = 0
            for row in rows:
                row_val = [col.value for col in row]
                T_tmp.append(SE3(row_val[0], row_val[1], row_val[2]) * SE3.RPY([np.deg2rad(row_val[3]), np.deg2rad(row_val[4]), np.deg2rad(row_val[5])]))
                ik_q = self.robot.ikine_LMS(T=T_tmp[i])
                if ik_q.success == True:
                    count += 1
                    manipulability_index.append(self.robot.manipulability(q=ik_q.q))
                i = i + 1

            if count == 0:
                return(0,0)
            else:
                final_score = count / i
                if count == 1:
                    return(final_score, manipulability_index[0]) # 回傳 manipulability[0]
                else:
                    return(final_score, np.mean(manipulability_index)) # 回傳 manipulability 取平均
    
    def point_Workspace_cal_Monte_Carlo(self):
        """
        Through the "Work space" page in the interface to calculate of the robot
        """
        i = 0
        # 角度轉換
        du = pi / 180
        # 度
        radian = 180 / pi
        # 弧度

        self.q1_s = -160
        self.q1_end = 160
        self.q2_s = -160
        self.q2_end = 160
        self.q3_s = -160
        self.q3_end = 160
        self.q4_s = -160
        self.q4_end = 160
        self.q5_s = -160
        self.q5_end = 160
        self.q6_s = -160
        self.q6_end = 160
        N = 1000
        theta1 = self.q1_end + (self.q1_end - self.q1_s) * np.random.rand(N, 1)
        theta2 = self.q2_end + (self.q2_end - self.q2_s) * np.random.rand(N, 1)
        theta3 = self.q3_end + (self.q3_end - self.q3_s) * np.random.rand(N, 1)
        theta4 = self.q4_end + (self.q4_end - self.q4_s) * np.random.rand(N, 1)
        theta5 = self.q5_end + (self.q5_end - self.q5_s) * np.random.rand(N, 1)
        theta6 = self.q6_end + (self.q6_end - self.q6_s) * np.random.rand(N, 1)

        for i in range(N):
            q1 = theta1[i, 0]
            q2 = theta2[i, 0]
            q3 = theta3[i, 0]
            q4 = theta4[i, 0]
            q5 = theta5[i, 0]
            q6 = theta6[i, 0]
            self.T = self.robot.fkine(
                [q1 * du, q2 * du, q3 * du, q4 * du, q5 * du, q6 * du]
            )
            t = np.round(self.T.t, 3)
            r = np.round(self.T.rpy('deg'), 3)
            self.T_x.append(t[0])
            self.T_y.append(t[1])
            self.T_z.append(t[2])
            self.T_roll.append(int(r[0]))
            self.T_pitch.append(int(r[1]))
            self.T_yaw.append(int(r[2]))
            i = i + 1
        
    def random_select_point(self):
        excel_file = Workbook()
        sheet = excel_file.active

        for i in range(10):
            x = np.random.randint(1,999)
            sheet.cell(row=i + 1, column=1).value = self.T_x[x]
            sheet.cell(row=i + 1, column=2).value = self.T_y[x]
            sheet.cell(row=i + 1, column=3).value = self.T_z[x]
            sheet.cell(row=i + 1, column=4).value = self.T_roll[x]
            sheet.cell(row=i + 1, column=5).value = self.T_pitch[x]
            sheet.cell(row=i + 1, column=6).value = self.T_yaw[x]

        file_name = self.xlsx_outpath + "/task_point" +".xlsx"
        excel_file.save(file_name)

    # def select_point_tested(self):
    #     task_point_6dof_tested

class RobotOptEnv_3dof(gym.Env):
    metadata = {
        'render.modes': ['human', 'rgb_array'],
        'video.frames_per_second': 2
    }
    def __init__(self):
        self.robot = modular_robot_3dof() # TODO: fixed 3dof
        self.robot_urdf = stl_conv_urdf("single_arm_v12","test")
        
        # callback:Enter the parameters of the algorithm to be optimized on the interface
        self.sub_optimal_design = rospy.Subscriber(
            "/optimal_design", optimal_design, self.optimal_design_callback
        )
        # 使用者設定參數
        self.payload = 5.0
        self.payload_position = np.array([0, 0, 0.04])
        self.vel = np.array([2.356194, 2.356194, 2.356194])# TODO: fixed 3dof
        self.acc = np.array([2.356194, 2.356194, 2.356194])# TODO: fixed 3dof
        self.total_weight = 20 # Kg
        self.total_cost = 1800 # 元
        # 預設二,三軸軸長
        self.std_L2 = 35.0 # 預設標準值 第二軸 35 cm
        self.std_L3 = 35.0 # 預設標準值 第三軸 35 cm
        # 觀察參數 motor
        self.motor = motor_data()
        self.res = self.motor.dynamixel_member
        self.high_torque = 120 # 預設標準值 馬達極限 120.0 N max
        self.low_torque = -120 # 預設標準值 馬達極限 -12.0 N max
        self.motor_cost_init = np.array([0,400,400,0,0,0], dtype=np.float64) # 預設最大馬達費用
        self.motor_weight_init = np.array([0,0.855,0.855,0,0,0], dtype=np.float64) # 預設最大馬達重量
        self.motor_cost = np.array([0,400,400,0,0,0], dtype=np.float64) # 馬達費用 # TODO: fixed 3dof
        self.motor_weight = np.array([0,0.855,0.855], dtype=np.float64) # 馬達重量# TODO: fixed 3dof
        self.motor_rated = np.array([44.7,44.7,44.7], dtype=np.float64)# TODO: fixed 3dof
        # 使用者設定參數 & 觀察參數
        self.reach_distance = 0.6 # 使用者設定可達半徑最小值
        self.high_reach_eva = 1 # 預設觀測標準值
        self.low_reach_eva = 0 # 預設觀測標準值
        self.high_manipulability = 1  # 預設觀測標準值
        self.low_manipulability = 0  # 預設觀測標準值
        self.high_std_L2 = 50 # 預設觀測標準值
        self.low_std_L2 = 0 # 預設觀測標準值
        self.high_std_L3 = 50 # 預設觀測標準值
        self.low_std_L3 = 0 # 預設觀測標準值
        self.torque_done = np.array([false, false, false])# TODO: fixed 3dof
        self.torque_over = False
        self.prev_shaping = None
        # TODO: 增加馬達模組選型action
        self.action_space = spaces.Discrete(12) # TODO: fixed 12種action
        
        # TODO: observation space for torque, reach, motor cost, weight, manipulability
        self.observation_space = spaces.Box(np.array([self.low_torque,self.low_torque,self.low_torque,self.low_torque,self.low_torque,self.low_torque, self.low_reach_eva, -float('inf'), -float('inf'), self.low_manipulability, self.low_std_L2, self.low_std_L3 ]), 
                                            np.array([self.high_torque,self.high_torque,self.high_torque,self.high_torque,self.high_torque,self.high_torque, self.high_reach_eva, self.total_cost, self.total_weight, self.high_manipulability, self.high_std_L2, self.high_std_L3]), 
                                            dtype=np.float64)
        # TODO: reward 歸一化
        self.state = np.array([0,0,0,0,0,0,0,0,0,0,0,0], dtype=np.float64)
        # self.pre_state = np.array([0,0,0,0,0,0,0,0,0,0,0,0], dtype=np.float64)# TODO: fixed 3dof

        #隨機抽樣點位初始化
        self.T_x = []
        self.T_y = []
        self.T_z = []
        self.T_roll = []
        self.T_pitch = []
        self.T_yaw = []
        self.xlsx_outpath = "./xlsx/"

        self.model_select = "train" # 選擇使用train or test model 

    def optimal_design_callback(self, data):
        # print(data.data)
        # TODO: 目標構型
        self.op_dof = data.dof
        self.op_payload = data.payload
        self.op_payload_position = data.payload_position
        self.op_vel = data.vel
        self.op_acc = data.acc
        self.op_radius = data.radius
        self.op_weight = data.arm_weight
        self.op_cost = data.cost
        
        print("op_dof:", self.op_dof)
        print("op_payload:", self.op_payload)
        print("op_payload_position:", self.op_payload_position)
        print("op_vel:", self.op_vel)
        print("op_acc:", self.op_acc)
        print("op_radius:", self.op_radius)
        print("op_weight:", self.op_weight)
        print("op_cost:", self.op_cost)
        
        self.payload = self.op_payload
        self.payload_position = np.array(self.op_payload_position)
        self.vel = np.array(self.op_vel[0:3])# TODO: fixed 3dof
        self.acc = np.array(self.op_acc[0:3])# TODO: fixed 3dof
        self.total_weight = self.op_weight # Kg
        self.total_cost = self.op_cost # 元
        self.reach_distance = self.op_radius # 使用者設定可達半徑最小值
        
    # TODO: fixed
    def step(self, action):
        assert self.action_space.contains(action), "%r (%s) invalid"%(action, type(action))
        # self.pre_state[0:3] = self.state[0:3]# TODO: fixed 3dof
        # TODO: 向量編碼動作
        if action == 0: # 軸2  # 短 # 型號1
            self.std_L2 -= 1.0
            # 配置軸2 motor 型號1
            motor_type = 0 # 型號
            axis = 2 # 軸數
            
        elif action == 1: # 軸2  # 長 # 型號1
            self.std_L2 += 1.0
            motor_type = 0 # 型號
            axis = 2 # 軸數
            
        elif action == 2: # 軸2  # 短 # 型號2
            self.std_L2 -= 1.0
            motor_type = 1 # 型號
            axis = 2 # 軸數
            
        elif action == 3: # 軸2  # 長 # 型號2
            self.std_L2 += 1.0
            motor_type = 1 # 型號
            axis = 2 # 軸數
            
        elif action == 4: # 軸2  # 短 # 型號3
            self.std_L2 -= 1.0
            motor_type = 2 # 型號
            axis = 2 # 軸數
            
        elif action == 5: # 軸2  # 長 # 型號3
            self.std_L2 += 1.0
            motor_type = 2 # 型號
            axis = 2 # 軸數
            
        elif action == 6: # 軸3  # 短 # 型號1
            self.std_L3 -= 1.0
            motor_type = 0 # 型號
            axis = 3 # 軸數
            
        elif action == 7: # 軸3  # 長 # 型號1
            self.std_L3 += 1.0
            motor_type = 0 # 型號
            axis = 3 # 軸數
            
        elif action == 8: # 軸3  # 短 # 型號2
            self.std_L3 -= 1.0
            motor_type = 1 # 型號
            axis = 3 # 軸數
            
        elif action == 9: # 軸3  # 長 # 型號2
            self.std_L3 += 1.0
            motor_type = 1 # 型號
            axis = 3 # 軸數
            
        elif action == 10: # 軸3  # 短 # 型號3
            self.std_L3 -= 1.0
            motor_type = 2 # 型號
            axis = 3 # 軸數
            
        elif action == 11: # 軸3  # 長 # 型號3
            self.std_L3 += 1.0
            motor_type = 2 # 型號
            axis = 3 # 軸數
        
        # 輸入action後 二,三軸軸長
        self.robot_urdf.specified_generate_write_urdf(self.std_L2, self.std_L3)
        self.robot.__init__() # 重製機器人
        torque = self.dynamics_torque_limit()
        self.state[0:3] = torque# TODO: fixed 3dof
        # # 可達性
        # self.state[6] = self.reach_evaluate()
        # 計算成本與重量    
        self.motor_cost[axis-1] = self.res.cost[motor_type]
        self.motor_weight[axis-1] = self.res.weight[motor_type]
        # print(self.motor_cost)
        # print(self.motor_weight)
        cost = sum(self.motor_cost) 
        weight = sum(self.motor_weight)
        # self.state[7] = cost
        # self.state[8] = weight
        
        # self.state[9] = self.manipulability_evaluate()
        # 可達性 # 可操作性
        self.state[6], self.state[7] = self.reach_manipulability_evaluate(self.model_select)
        # rospy.loginfo("reach_score: %s", self.state[6])
        # rospy.loginfo("manipulability_score: %s", self.state[9])
        self.state[8] = self.std_L2
        self.state[9] = self.std_L3
        self.motor_rated[axis-1] = self.res.rated_torque[motor_type]
        # rospy.loginfo("configuration cost & weight: %s, %s", cost, weight)
        self.counts += 1
        # rospy.loginfo("self.state: %s", self.state)
        reward = 0
        # TODO: fixed
        shaping = (
            # TODO: fixed 3dof
            -1 * np.sqrt(self.state[0] * self.state[0] + self.state[1] * self.state[1] + self.state[2] * self.state[2])
            + 100 * self.state[6] # 可達性
            # - 0.01 * self.state[7] # cost
            # - 10 * self.state[8] # weight
            + 1000 * self.state[9] # 可操作性
        ) 

        if self.prev_shaping is not None:
            reward = shaping - self.prev_shaping
        self.prev_shaping = shaping

        # 判斷超出最大扭矩
        for i in range(3): # TODO: fixed 3dof
            # TODO:consider cost & weight 
            # if np.abs(self.state[i]) > self.motor_rated[i]:
            if np.abs(self.state[i]) > 44.7: # TODO: fixed 
                self.torque_over = True # 超出最大扭矩
                break # TODO

        terminated = False
        # if down 完成任务 
        
        # if self.counts == 10:
        #     if self.state[6] <= 0.6: # TODO: fixed
        #         terminated = True
        #         reward += -50
        #     if self.torque_over == True: # TODO: fixed 0112 00:22 改為超過最大的馬達型號torque
        #         terminated = True
        #         reward += -100

        # if self.torque_over == False and self.state[6] == 1 and self.state[8] < self.op_weight and self.state[7] < self.op_cost:  # TODO: fixed 增加 cost & weight & ... 
        #     terminated = True
        #     reward += +50
        if self.state[10] > 0 and self.state[11] > 0:
            if self.torque_over == True and self.state[6] < 0.6: 
                reward += -20
            elif self.torque_over == True and 0.6 <= self.state[6] < 0.8: 
                reward += -10
            elif self.torque_over == True and 0.8 <= self.state[6] <= 1.0:   
                reward += -5
            elif self.torque_over == False and self.state[6] < 0.6: 
                reward += +5
            elif self.torque_over == False and 0.6 <= self.state[6] < 0.8:
                reward += +10
            elif self.torque_over == False and 0.8 <= self.state[6] < 0.9:
                reward += +20
            elif self.torque_over == False and 0.9 <= self.state[6] < 1.0:
                terminated = True
                reward += +50
            elif self.torque_over == False and self.state[6] == 1.0:
                terminated = True
                reward += +100
        # else:
        #     pass
            
            if self.counts == 30:
                if self.state[6] < 0.6: # TODO: fixed
                    reward += -50
                else:
                    reward += 10
                if self.torque_over == True: # TODO: fixed 0112 00:22 改為超過最大的馬達型號torque
                    reward += -100
                else:
                    reward += +10
                terminated = True
                self.counts = 0
                # reward += +30
        else:
            terminated = True
            reward += -100

        self.torque_over = False #reset
        rospy.loginfo("counts: %s", self.counts)
        rospy.loginfo("step_reward: %s", reward)
        rospy.loginfo("================================")
        # print("================================")
        current_design = [self.std_L2, self.std_L3, self.motor_rated[1], self.motor_rated[2]]
        return self.state, reward, terminated, current_design

    # reset环境状态 
    def reset(self):
        # random state (手臂長度隨機)
        self.std_L2, self.std_L3 = self.robot_urdf.opt_random_generate_write_urdf() # 啟用隨機的L2,L3長度urdf
        self.robot.__init__() # 重製機器人
        torque = self.dynamics_torque_limit()
        self.state[0:3] = torque# TODO: fixed 3dof
        # 生成隨機 payload (kg)
        rand_payload = np.random.uniform(low=1, high=4)
        self.payload = rand_payload
        # rospy.loginfo("payload: %s", self.payload)
        self.point_Workspace_cal_Monte_Carlo() # 在當前reset出來的機械手臂構型下, 生成點位
        self.random_select_point() # 先隨機抽樣30個點位
        self.prev_shaping = None
        # reach_score = self.reach_evaluate()
        # manipulability_score = self.manipulability_evaluate()
        reach_score, manipulability_score = self.reach_manipulability_evaluate(self.model_select)
        self.state[6] = reach_score
        # self.state[7] = sum(self.motor_cost_init)
        # self.state[8] = sum(self.motor_weight_init)
        self.state[7] = manipulability_score
        self.state[8] = self.std_L2
        self.state[9] = self.std_L3
        self.counts = 0
        return self.state    
    
    # reset环境状态 
    def tested_reset(self):
        # random state (手臂長度隨機)
        self.std_L2, self.std_L3 = self.robot_urdf.opt_random_generate_write_urdf() # 啟用隨機的L2,L3長度urdf
        self.robot.__init__() # 重製機器人
        self.payload = self.op_payload
        self.payload_position = np.array(self.op_payload_position)
        self.vel = np.array(self.op_vel[0:3])# TODO: fixed 3dof
        self.acc = np.array(self.op_acc[0:3])# TODO: fixed 3dof
        self.total_weight = self.op_weight # Kg
        self.total_cost = self.op_cost # 元
        self.reach_distance = self.op_radius # 使用者設定可達半徑最小值

        torque = self.dynamics_torque_limit()
        self.state[0:3] = torque# TODO: fixed 3dof
        self.prev_shaping = None
        # reach_score = self.reach_evaluate()
        # manipulability_score = self.manipulability_evaluate()
        reach_score, manipulability_score = self.reach_manipulability_evaluate(self.model_select)
        self.state[6] = reach_score
        # self.state[7] = sum(self.motor_cost_init)
        # self.state[8] = sum(self.motor_weight_init)
        self.state[7] = manipulability_score
        self.state[8] = self.std_L2
        self.state[9] = self.std_L3
        self.counts = 0
        return self.state    

    # 視覺化呈現，它只會回應出呼叫那一刻的畫面給你，要它持續出現，需要寫個迴圈
    def render(self, mode='human'):
        return None
        
    def close(self):
        return None
        
    def dynamics_torque_limit(self):
        """
        Calculate the maximum torque required by

        each axis when the arm of each axis is the longest and the acceleration is the highest
        """
        torque = np.array([np.zeros(shape=3)])# TODO: fixed 3dof
        # axis_angle = np.array([np.zeros(shape=6)])
        axis_angle = []
        append_torque_limit_list = []
        temp_torque_max = []
        temp_torque_min = []
        Torque_Max = []
        # 角度轉換
        du = pi / 180
        # 度
        # radian = 180 / pi
        # 弧度
        self.robot.payload(self.payload, self.payload_position)  # set payload
        torque = np.array([np.zeros(shape=3)])# TODO: fixed 3dof
        q_list = [0, 90, -90, 180, -180]
        # TODO: fixed 3dof
        T_cell = (
            len(q_list)
            * len(q_list)
            * len(q_list)
        )
        # TODO: fixed 3dof
        for i in range(len(q_list)):
            q1 = q_list[i]
            percent = i / T_cell * 100
            for j in range(len(q_list)):
                q2 = q_list[j]
                for k in range(len(q_list)):
                    q3 = q_list[k]
                    axis_angle.append([q1, q2, q3])
                    load = np.array(
                        [
                            self.robot.rne(
                                [
                                    q1 * du,
                                    q2 * du,
                                    q3 * du,
                                    # q4 * du,
                                    # q5 * du,
                                    # q6 * du,
                                ],
                                self.vel,
                                self.acc
                            )
                        ]
                    )
                    torque = np.append(torque, load, axis=0)
        # TODO: fixed 3dof
        for i in range(3):
            axis = i
            toque_max_index = np.argmax(torque[:, axis])
            toque_min_index = np.argmin(torque[:, axis])

            temp_torque_max = torque[toque_max_index].tolist()
            temp_torque_max.extend(axis_angle[toque_max_index])
            temp_torque_min = torque[toque_min_index].tolist()
            temp_torque_min.extend(axis_angle[toque_min_index])
            append_torque_limit_list.append(temp_torque_max)
            append_torque_limit_list.append(temp_torque_min)
            Torque_Max.append(abs(torque[toque_max_index][i]))
            self.torque_dynamics_limit = Torque_Max
        return self.torque_dynamics_limit

    def reach_manipulability_evaluate(self,model_select):
        if model_select == "train":
            # import xlsx
            df = load_workbook("./xlsx/task_point_3dof.xlsx")# TODO: fixed 3dof
            sheets = df.worksheets
            sheet1 = sheets[0]
            rows = sheet1.rows
            cols = sheet1.columns
            T_tmp = []
            score = []
            manipulability_index = []
            i = 0
            false_done = False
            count = 0
            for row in rows:
                row_val = [col.value for col in row]
                # TODO: fixed
                T_tmp.append(SE3(row_val[0], row_val[1], row_val[2]))# TODO: fixed 3dof
                ik_q = self.robot.ikine_LMS(T=T_tmp[i], mask = [1,1,1,0,0,0])# TODO: fixed 3dof
                if ik_q.success == True:
                    count += 1
                    manipulability_index.append(self.robot.manipulability(q=ik_q.q))
                i = i + 1

            if count == 0:
                return(0,0)
            else:
                final_score = count / i
                if count == 1:
                    return(final_score, manipulability_index[0]) # 回傳 manipulability[0]
                else:
                    return(final_score, np.mean(manipulability_index)) # 回傳 manipulability 取平均
        elif model_select == "test":
            # import xlsx
            df = load_workbook("./xlsx/task_point_3dof_tested.xlsx")# TODO: fixed 3dof
            sheets = df.worksheets
            sheet1 = sheets[0]
            rows = sheet1.rows
            cols = sheet1.columns
            T_tmp = []
            score = []
            manipulability_index = []
            i = 0
            false_done = False
            count = 0
            for row in rows:
                row_val = [col.value for col in row]
                # TODO: fixed
                T_tmp.append(SE3(row_val[0], row_val[1], row_val[2]))# TODO: fixed 3dof
                ik_q = self.robot.ikine_LMS(T=T_tmp[i], mask = [1,1,1,0,0,0])# TODO: fixed 3dof
                if ik_q.success == True:
                    count += 1
                    manipulability_index.append(self.robot.manipulability(q=ik_q.q))
                i = i + 1

            if count == 0:
                return(0,0)
            else:
                final_score = count / i
                if count == 1:
                    return(final_score, manipulability_index[0]) # 回傳 manipulability[0]
                else:
                    return(final_score, np.mean(manipulability_index)) # 回傳 manipulability 取平均
    
    def point_Workspace_cal_Monte_Carlo(self):
        """
        Through the "Work space" page in the interface to calculate of the robot
        """
        i = 0
        # 角度轉換
        du = pi / 180
        # 度
        radian = 180 / pi
        # 弧度

        self.q1_s = -160
        self.q1_end = 160
        self.q2_s = -160
        self.q2_end = 160
        self.q3_s = -160
        self.q3_end = 160
        # TODO: fixed 3dof
        # self.q4_s = -160
        # self.q4_end = 160
        # self.q5_s = -160
        # self.q5_end = 160
        # self.q6_s = -160
        # self.q6_end = 160
        N = 1000
        theta1 = self.q1_end + (self.q1_end - self.q1_s) * np.random.rand(N, 1)
        theta2 = self.q2_end + (self.q2_end - self.q2_s) * np.random.rand(N, 1)
        theta3 = self.q3_end + (self.q3_end - self.q3_s) * np.random.rand(N, 1)
        # TODO: fixed 3dof
        # theta4 = self.q4_end + (self.q4_end - self.q4_s) * np.random.rand(N, 1)
        # theta5 = self.q5_end + (self.q5_end - self.q5_s) * np.random.rand(N, 1)
        # theta6 = self.q6_end + (self.q6_end - self.q6_s) * np.random.rand(N, 1)

        for i in range(N):
            q1 = theta1[i, 0]
            q2 = theta2[i, 0]
            q3 = theta3[i, 0]
            # TODO: fixed 3dof
            # q4 = theta4[i, 0]
            # q5 = theta5[i, 0]
            # q6 = theta6[i, 0]
            self.T = self.robot.fkine(
                [q1 * du, q2 * du, q3 * du]# TODO: fixed 3dof
            )
            t = np.round(self.T.t, 3)
            # r = np.round(self.T.rpy('deg'), 3)# TODO: fixed 3dof
            self.T_x.append(t[0])
            self.T_y.append(t[1])
            self.T_z.append(t[2])
            # TODO: fixed 3dof
            # self.T_roll.append(int(r[0]))
            # self.T_pitch.append(int(r[1]))
            # self.T_yaw.append(int(r[2]))
            i = i + 1
        
    def random_select_point(self):
        excel_file = Workbook()
        sheet = excel_file.active

        for i in range(10):
            x = np.random.randint(1,999)
            sheet.cell(row=i + 1, column=1).value = self.T_x[x]
            sheet.cell(row=i + 1, column=2).value = self.T_y[x]
            sheet.cell(row=i + 1, column=3).value = self.T_z[x]
            # TODO: fixed 3dof
            # sheet.cell(row=i + 1, column=4).value = self.T_roll[x]
            # sheet.cell(row=i + 1, column=5).value = self.T_pitch[x]
            # sheet.cell(row=i + 1, column=6).value = self.T_yaw[x]

        file_name = self.xlsx_outpath + "/task_point_3dof" +".xlsx"# TODO: fixed 3dof
        excel_file.save(file_name)

    # def select_point_tested(self):
    #     task_point_6dof_tested
if __name__ == '__main__':
    env = RobotOptEnv()
    
    action = env.action_space.sample()
    print(action)
    print(action[0])
    print(type(action[0]))
    print(action[1])
    print(action[2])
    
