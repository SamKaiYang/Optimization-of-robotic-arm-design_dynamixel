#!/usr/bin/env python3
# coding: utf-8
import gym
from gym import spaces
import numpy as np
from math import pi

from sympy import false
# from torch import R
from interface_control.msg import optimal_design
from dynamics.arm_workspace import arm_workspace_plane
# from robot_urdf import RandomRobot
from dynamics.motor_module import motor_data
from dynamics.modular_robot_3dof import modular_robot_3dof
from dynamics.modular_robot_4dof import modular_robot_4dof
from dynamics.modular_robot_5dof import modular_robot_5dof
from dynamics.modular_robot_6dof import modular_robot_6dof
# from dynamics.stl_conv_6dof_urdf import stl_conv_urdf
from dynamics.stl_conv_6dof_urdf_dynamixel import stl_conv_urdf
import rospy
# one-hot encoder
from sklearn import preprocessing
import pandas as pd
import yaml
# add 可達性可操作性評估
import numpy as np
from roboticstoolbox import DHRobot, RevoluteDH
from spatialmath import SE3
import spatialmath as sm
from urdf_parser_py.urdf import URDF
import os
from openpyxl import load_workbook
from openpyxl import Workbook

# TODO: 初版 只考慮 6 dof 機器人的關節長度變化, 觀察各軸馬達極限之輸出最大torque值
class RobotOptEnv(gym.Env):
    metadata = {
        'render.modes': ['human', 'rgb_array'],
        'video.frames_per_second': 2
    }
    def __init__(self):
        self.robot = modular_robot_6dof()
        self.robot_urdf = stl_conv_urdf("single_arm_v12","test")
        # self.robot_urdf.init_dynamixel_diff_inertia()
        # callback:Enter the parameters of the algorithm to be optimized on the interface
        self.sub_optimal_design = rospy.Subscriber(
            "/optimal_design", optimal_design, self.optimal_design_callback
        )
        # 使用者設定參數
        self.payload = 5.0
        self.payload_position = np.array([0, 0, 0.04])
        self.vel = np.array([3.04, 3.04, 3.04, 3.04, 3.04, 3.04])
        self.acc = np.array([2.356194, 2.356194, 2.356194, 2.356194, 2.356194, 2.356194])
        self.total_weight = 20 # Kg
        self.total_cost = 1800 # 元
        self.reachable_tmp = 0
        # 預設二,三軸軸長
        self.std_L2 = 35.0 # 預設標準值 第二軸 35 cm
        self.std_L3 = 35.0 # 預設標準值 第三軸 35 cm
        # 觀察參數 motor
        self.motor = motor_data()
        self.res = self.motor.dynamixel_member
        self.high_torque = float('inf') # 預設標準值 馬達極限 120.0 N max
        self.low_torque = float('-inf') # 預設標準值 馬達極限 -12.0 N max
        self.motor_cost_init = np.array([0,400,400,0,0,0], dtype=np.float64) # 預設最大馬達費用
        self.motor_weight_init = np.array([0,0.855,0.855,0,0,0], dtype=np.float64) # 預設最大馬達重量
        self.motor_cost = np.array([0,400,400,0,0,0], dtype=np.float64) # 馬達費用
        self.motor_weight = np.array([0,0.855,0.855,0,0,0], dtype=np.float64) # 馬達重量
        self.motor_rated = np.array([44.7,44.7,44.7,44.7,44.7,44.7], dtype=np.float64)
        # 使用者設定參數 & 觀察參數
        self.reach_distance = 0.6 # 使用者設定可達半徑最小值
        self.high_reach_eva = 1 # 預設觀測標準值
        self.low_reach_eva = 0 # 預設觀測標準值
        self.high_manipulability = 3  # 預設觀測標準值
        self.low_manipulability = 0  # 預設觀測標準值
        self.high_std_L2 = 70 # 預設觀測標準值
        self.low_std_L2 = -25 # 預設觀測標準值
        self.high_std_L3 = 70 # 預設觀測標準值
        self.low_std_L3 = -25 # 預設觀測標準值
        self.high_ratio_over = 10 # 預設觀測標準值
        self.low_ratio_over = 0 # 預設觀測標準值
        self.high_torque_over = 10 # 預設觀測標準值
        self.low_torque_over = 0 # 預設觀測標準值
        self.high_power_consumption = float('inf')
        self.low_power_consumption = float('-inf')
        self.torque_done = np.array([false, false, false, false, false, false])
        self.torque_over = False
        self.ratio_over = False
        self.prev_shaping = None
        self.motor_type_axis_2 = 5.1
        self.motor_type_axis_3 = 5.1
        self.mission_time = 0
        self.low_torque_cost = -200
        self.high_torque_cost = 200
        # FIXME:
        self.action_select = 'fixed' 
        self.point_test_excel = './xlsx/task_point_6dof_tested_ori_random.xlsx'
        self.MAX_LENGTH = 40
        self.MIN_LENGTH = 5
        # TODO: 增加馬達模組選型action
        self.action_space = spaces.Discrete(12) # TODO: fixed 12種action
        
        # TODO: observation space for torque, reach, motor cost, weight, manipulability
        self.observation_space = spaces.Box(np.array([self.low_ratio_over,self.low_torque_over,self.low_power_consumption, self.low_reach_eva, self.low_manipulability, self.low_std_L2, self.low_std_L3, self.low_torque_cost ]), 
                                            np.array([self.high_ratio_over,self.high_torque_over,self.high_power_consumption, self.high_reach_eva, self.high_manipulability, self.high_std_L2, self.high_std_L3, self.high_torque_cost]), 
                                            dtype=np.float64)
        # TODO: reward 歸一化
        self.state = np.array([0,0,0,0,0,0,0,0], dtype=np.float64)
        self.pre_state = np.array([0,0,0,0,0,0,0,0], dtype=np.float64)

        #隨機抽樣點位初始化
        self.T_x = []
        self.T_y = []
        self.T_z = []
        self.T_roll = []
        self.T_pitch = []
        self.T_yaw = []
        self.xlsx_outpath = "./xlsx/"

        self.model_select = "train" # 選擇使用train or test model 

        self.op_dof = None
        self.op_payload = None
        self.op_payload_position = None
        self.op_vel = None
        self.op_acc = None
        self.op_radius = None
        self.op_weight = None
        self.op_cost = None

    def optimal_design_callback(self, data):
        # print(data.data)
        # TODO: 目標構型
        self.op_dof = data.dof
        self.op_payload = data.payload
        self.op_payload_position = data.payload_position
        self.op_vel = data.vel
        self.op_acc = data.acc
        self.op_radius = data.radius
        self.op_weight = data.arm_weight
        self.op_cost = data.cost
        
        print("op_dof:", self.op_dof)
        print("op_payload:", self.op_payload)
        print("op_payload_position:", self.op_payload_position)
        print("op_vel:", self.op_vel)
        print("op_acc:", self.op_acc)
        print("op_radius:", self.op_radius)
        print("op_weight:", self.op_weight)
        print("op_cost:", self.op_cost)
        
        self.payload = self.op_payload
        self.payload_position = np.array(self.op_payload_position)
        self.vel = np.array(self.op_vel[0:6])
        self.acc = np.array(self.op_acc[0:6])
        self.total_weight = self.op_weight # Kg
        self.total_cost = self.op_cost # 元
        self.reach_distance = self.op_radius # 使用者設定可達半徑最小值
    # def one_hot_encode(action, num_actions):
    #     one_hot = np.zeros(num_actions)
    #     one_hot[action] = 1
    #     return one_hot
    # TODO: fixed
    def step(self, action):
        assert self.action_space.contains(action), "%r (%s) invalid"%(action, type(action))
        if self.action_select == 'variable':
            if action == 0: # 軸2  # 短 # 型號1
                self.std_L2 -= 1.0
                # 配置軸2 motor 型號1
                #[5.1, 25.3, 44.7]
                self.motor_type_axis_2 = 5.1 # 型號
            
            elif action == 1: # 軸2  # 長 # 型號1
                self.std_L2 += 1.0
                self.motor_type_axis_2 = 5.1 # 型號
                
            elif action == 2: # 軸2  # 短 # 型號2
                self.std_L2 -= 1.0
                self.motor_type_axis_2 = 25.3 # 型號

            elif action == 3: # 軸2  # 長 # 型號2
                self.std_L2 += 1.0
                self.motor_type_axis_2 = 25.3 # 型號
                
            elif action == 4: # 軸2  # 短 # 型號3
                self.std_L2 -= 1.0
                self.motor_type_axis_2 = 44.7 # 型號

            elif action == 5: # 軸2  # 長 # 型號3
                self.std_L2 += 1.0
                self.motor_type_axis_2 = 44.7 # 型號
                
            elif action == 6: # 軸3  # 短 # 型號1
                self.std_L3 -= 1.0
                self.motor_type_axis_3 = 5.1 # 型號
                
            elif action == 7: # 軸3  # 長 # 型號1
                self.std_L3 += 1.0
                self.motor_type_axis_3 = 5.1 # 型號
                
            elif action == 8: # 軸3  # 短 # 型號2
                self.std_L3 -= 1.0
                self.motor_type_axis_3 = 25.3 # 型號
                
            elif action == 9: # 軸3  # 長 # 型號2
                self.std_L3 += 1.0
                self.motor_type_axis_3 = 25.3 # 型號

            elif action == 10: # 軸3  # 短 # 型號3
                self.std_L3 -= 1.0
                self.motor_type_axis_3 = 44.7 # 型號

            elif action == 11: # 軸3  # 長 # 型號3
                self.std_L3 += 1.0
                self.motor_type_axis_3 = 44.7 # 型號
        elif self.action_select == 'fixed':
            if action == 0: # 軸2  # 短 # 型號1
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                # 配置軸2 motor 型號1
                #[5.1, 25.3, 44.7]
                self.motor_type_axis_2 = 5.1 # 型號
                
            elif action == 1: # 軸2  # 長 # 型號1
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_2 = 5.1 # 型號
                
            elif action == 2: # 軸2  # 短 # 型號2
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                self.motor_type_axis_2 = 25.3 # 型號

            elif action == 3: # 軸2  # 長 # 型號2
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_2 = 25.3 # 型號
                
            elif action == 4: # 軸2  # 短 # 型號3
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                self.motor_type_axis_2 = 44.7 # 型號

            elif action == 5: # 軸2  # 長 # 型號3
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_2 = 44.7 # 型號
                
            elif action == 6: # 軸3  # 短 # 型號1
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                self.motor_type_axis_3 = 5.1 # 型號
                
            elif action == 7: # 軸3  # 長 # 型號1
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_3 = 5.1 # 型號
                
            elif action == 8: # 軸3  # 短 # 型號2
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                self.motor_type_axis_3 = 25.3 # 型號
                
            elif action == 9: # 軸3  # 長 # 型號2
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_3 = 25.3 # 型號

            elif action == 10: # 軸3  # 短 # 型號3
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                self.motor_type_axis_3 = 44.7 # 型號

            elif action == 11: # 軸3  # 長 # 型號3
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_3 = 44.7 # 型號

        # 輸入action後 二,三軸軸長
        self.robot_urdf.specified_generate_write_urdf(self.std_L2, self.std_L3)
        self.robot.__init__() # 重製機器人
        self.robot.payload(self.payload, self.payload_position)  # set payload
        reach_score, manipulability_score = self.only_reachable_manipulability(self.model_select)
        if reach_score == 1:
            ratio_over, torque_over, consumption, reach_score, manipulability_score = self.performance_evaluate(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
        else:
            ratio_over = 10
            torque_over = 10
            consumption = float('inf')
        # ratio_over, torque_over, consumption, reach_score, manipulability_score = self.performance_evaluate(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
        # ratio_over, torque_over, consumption = self.power_consumption(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
        # rospy.loginfo("reach_score: %s", reach_score)
        self.state[0] = ratio_over
        self.state[1] = torque_over
        self.state[2] = np.abs(consumption)
        # 可達性 # 可操作性
        # self.state[3], self.state[4] = self.reach_manipulability_evaluate(self.model_select)
        self.state[3] = reach_score
        self.state[4] = manipulability_score
        self.state[5] = self.std_L2
        self.state[6] = self.std_L3
        self.state[7] = self.motor_type_axis_2 + self.motor_type_axis_3
        self.counts += 1
        reward = 0
        
        
        # TODO: fixed
        # shaping = (
        #     # - self.state[0] # 超過轉速量
        #     - self.state[7] # 馬達扭矩成本
        #     - self.state[2] # 功耗
        #     # + 100 * self.state[3] # 可達性
        #     + 1000 * self.state[4] # 可操作性
        # ) 
        if self.state[3] == 1 and self.reachable_tmp == 1:
            shaping = (
                        - self.state[7] # 馬達扭矩成本
                        - self.state[2] # 功耗
                        + 1000 * self.state[4] # 可操作性
                    ) 
            if self.prev_shaping != None:
                reward = shaping - self.prev_shaping
            self.prev_shaping = shaping
        elif self.state[3] == 1 and self.reachable_tmp != 1:
            tmp_shaping = (
                        - self.state[7] # 馬達扭矩成本
                        - self.state[2] # 功耗
                        + 1000 * self.state[4] # 可操作性
                    ) 
            self.prev_shaping = tmp_shaping
            reward = 0
            # self.prev_shaping = tmp_shaping
        elif self.state[3] != 1 and self.reachable_tmp == 1:
            tmp_shaping = (
                        + 1000 * self.state[4] # 可操作性
                    ) 
            self.prev_shaping = tmp_shaping
            reward = 0
            # self.prev_shaping = tmp_shaping
        else:
            shaping = (
                        + 1000 * self.state[4] # 可操作性
                    ) 
            if self.prev_shaping != None:
                reward = shaping - self.prev_shaping
            self.prev_shaping = shaping
        self.reachable_tmp = self.state[3]
        # if self.prev_shaping is not None:
        #     reward = shaping - self.prev_shaping
        # self.prev_shaping = shaping
        rospy.loginfo("-------------------------")
        rospy.loginfo("shaping reward: %s", reward)

        if self.state[0] > 0:
            self.ratio_over = True

        if self.state[1] > 0:
            self.torque_over = True

        terminated = False
        # TODO: case 1
        # if self.state[5] <= self.MIN_LENGTH or self.state[6] <= self.MIN_LENGTH  or self.state[5] >= self.MAX_LENGTH or self.state[6] >= self.MAX_LENGTH:
        #     terminated = True
        #     reward += -200
        # if self.ratio_over == True or self.torque_over == True and self.state[3] < 0.6: 
        #     reward += -20
        # elif self.ratio_over == True or self.torque_over == True and 0.6 <= self.state[3] < 0.8: 
        #     reward += -10
        # elif self.ratio_over == True or self.torque_over == True and 0.8 <= self.state[3] <= 1.0:   
        #     reward += -5
        # elif self.ratio_over == False and self.torque_over == False and self.state[3] < 0.6: 
        #     reward += +5
        # elif self.ratio_over == False and self.torque_over == False and 0.6 <= self.state[3] < 0.8:
        #     reward += +10
        # elif self.ratio_over == False and self.torque_over == False and 0.8 <= self.state[3] < 0.9:
        #     reward += +20
        # elif self.ratio_over == False and self.torque_over == False and 0.9 <= self.state[3] < 1.0:
        #     # terminated = True
        #     reward += +50
        # # success
        # elif self.ratio_over == False and self.torque_over == False and self.state[3] == 1.0:
        #     terminated = True
        #     reward += +100
        #     self.counts = 0
        # TODO: case2
        # if self.ratio_over or self.torque_over:
        #     percent = 100 - self.state[3] * 100    
        #     reward += -percent
        # else:
        #     percent = self.state[3] * 100
        #     reward += percent
        #     # success
        #     if percent == 100:
        #         terminated = True
        #         self.counts = 0
        # TODO: case 3
        if self.state[5] <= self.MIN_LENGTH or self.state[6] <= self.MIN_LENGTH  or self.state[5] >= self.MAX_LENGTH or self.state[6] >= self.MAX_LENGTH:
            terminated = True
            reward += -200
        percent = 100 - self.state[3] * 100
        reward += -percent
        # reachable == 1
        if percent == 0:
            ratio_score = self.ratio_over
            torque_score = self.torque_over
            reward -= ratio_score * 3
            reward -= torque_score * 3
            if self.ratio_over == 0 and self.torque_over == 0:
                reward += 200
                terminated = True
                self.counts = 0
        if self.counts == 50: # max_steps
            terminated = True
            self.counts = 0
        
        self.ratio_over = False
        self.torque_over = False #reset
       
        current_design = [self.std_L2, self.std_L3, self.motor_rated[1], self.motor_rated[2]]
        return self.state, reward, terminated, current_design
    # reset环境状态 
    def reset(self):
        if self.model_select == "train":
            rospy.loginfo("model_select:%s", self.model_select)
            if self.action_select == 'variable':
                self.std_L2, self.std_L3 = self.robot_urdf.opt_random_generate_write_urdf() # 啟用隨機的L2,L3長度urdf
            elif self.action_select == 'fixed':
                random_total_arm_length = np.random.uniform(low=10, high=60) # FIX: 臂長長度
                self.std_L2, self.std_L3 = self.robot_urdf.opt_specify_random_generate_write_urdf(random_total_arm_length) # 啟用隨機的L2,L3長度urdf, 並指定總臂長
            
            # self.std_L2, self.std_L3 = self.robot_urdf.opt_specify_random_generate_write_urdf(random_total_arm_length) # 啟用隨機的L2,L3長度urdf, 並指定總臂長
            self.robot.__init__() # 重製機器人
            self.motor_type_axis_2 = 44.7
            self.motor_type_axis_3 = 44.7
            self.mission_time = np.random.uniform(low = 20, high = 50)
            
            # rospy.loginfo("random_total_arm_length: %s", random_total_arm_length)
            rospy.loginfo("std_L2: %s", self.std_L2)
            rospy.loginfo("std_L3: %s", self.std_L3)
            rospy.loginfo("mission_time: %s", self.mission_time)
            
            # 生成隨機 payload (kg)
            rand_payload = np.random.uniform(low=1, high=6)
            # FIXME: 修改未成功匯入payload給予機器人的問題
            self.payload = rand_payload
            self.payload_position =  [0, 0, 0.04]
            self.robot.payload(self.payload, self.payload_position)  # set payload
            self.point_Workspace_cal_Monte_Carlo() # 在當前reset出來的機械手臂構型下, 生成點位
            self.random_select_point() # 先隨機抽樣30個點位
            reach_score, manipulability_score = self.only_reachable_manipulability(self.model_select)
            if reach_score == 1:
                ratio_over, torque_over, consumption, reach_score, manipulability_score = self.performance_evaluate(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
            else:
                ratio_over = 10
                torque_over = 10
                consumption = float('inf')
            self.prev_shaping = None
            self.reachable_tmp = 0
            self.state[0] = ratio_over
            self.state[1] = torque_over
            self.state[2] = np.abs(consumption)
            self.state[3] = reach_score
            self.state[4] = manipulability_score
            self.state[5] = self.std_L2
            self.state[6] = self.std_L3
            self.state[7] = self.motor_type_axis_2 + self.motor_type_axis_3
            self.counts = 0
            return self.state
        elif self.model_select == "test":
            # random state (手臂長度隨機)
            rospy.loginfo("model_select:%s", self.model_select)
            total_arm_length = self.op_radius
            # if self.action_select == 'variable':
            #     # 固定random 大小臂長 上下5cm 內
            #     self.std_L2, self.std_L3 = self.robot_urdf.opt_random_generate_write_urdf() # 啟用隨機的L2,L3長度urdf
            # elif self.action_select == 'fixed':
            #     total_arm_length = self.op_radius
            #     # 固定random 大小臂長 上下5cm 內
            #     self.std_L2, self.std_L3 = self.robot_urdf.opt_specify_random_generate_write_urdf(total_arm_length) # 啟用隨機的L2,L3長度urdf, 並指定總臂長
            
            
            # TODO: add 固定random 大小臂長 上下5cm 內
            self.std_L2, self.std_L3 = self.robot_urdf.opt_specify_test_generate_write_urdf(total_arm_length)
            self.robot.__init__() # 重製機器人
            # FIXME: 修改未成功匯入payload給予機器人的問題
            self.payload = self.op_payload
            self.payload_position = self.op_payload_position
            self.robot.payload(self.payload, self.payload_position)  # set payload
            rospy.loginfo("mission_time: %s", self.mission_time)
            self.total_weight = self.op_weight # Kg
            self.total_cost = self.op_cost # 元
            self.reach_distance = self.op_radius # 使用者設定可達半徑最小值
            self.motor_type_axis_2 = 44.7
            self.motor_type_axis_3 = 44.7
            # ratio_over, torque_over, consumption, reach_score, manipulability_score = self.performance_evaluate(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
            reach_score, manipulability_score = self.only_reachable_manipulability(self.model_select)
            if reach_score == 1:
                ratio_over, torque_over, consumption, reach_score, manipulability_score = self.performance_evaluate(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
            else:
                ratio_over = 10
                torque_over = 10
                consumption = float('inf')
            self.state[0] = ratio_over
            self.state[1] = torque_over
            self.state[2] = np.abs(consumption)
            self.prev_shaping = None
            self.reachable_tmp = 0
            self.state[3] = reach_score
            self.state[4] = manipulability_score
            self.state[5] = self.std_L2
            self.state[6] = self.std_L3
            self.state[7] = self.motor_type_axis_2 + self.motor_type_axis_3
            self.counts = 0
            return self.state        
    def original_design(self,std_L2,std_L3, motor_1, motor_2, payload, mission_time):
        #  指定手臂長度
        print("aaaaa")
        self.robot_urdf.specified_generate_write_urdf(std_L2, std_L3)
        self.robot.__init__() # 重製機器人
        motor_type_axis_2 = motor_1
        motor_type_axis_3 = motor_2
        self.payload = payload
        op_payload_position = [0, 0, 0.04]
        # self.payload_position = [np.array(op_payload_position)]
        self.robot.payload(self.payload, op_payload_position)  # set payload
        # mission_time = 30
        model_select = "test"
        # rospy.loginfo("mission_time: %s", mission_time)
        self.point_Workspace_cal_Monte_Carlo() # 在當前reset出來的機械手臂構型下, 生成點位
        self.random_select_point() # 先隨機抽樣30個點位
        self.mission_time = mission_time
        ratio_over, torque_over, consumption, reach_score, manipulability_score = self.performance_evaluate(model_select, motor_type_axis_2, motor_type_axis_3)
        # ratio_over, torque_over, consumption = self.power_consumption(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
        # torque = self.dynamics_torque_limit()
        rospy.loginfo("ratio_over: %s", ratio_over)
        rospy.loginfo("torque_over: %s", torque_over)
        rospy.loginfo("consumption: %s", np.abs(consumption))
        # 生成隨機 payload (kg)
        rospy.loginfo("reach_score: %s", reach_score)
        rospy.loginfo("manipulability_score: %s", manipulability_score)
        rospy.loginfo("std_L2: %s", std_L2)
        rospy.loginfo("std_L3: %s", std_L3)
        rospy.loginfo("motor_torque_cost: %s", motor_type_axis_2 + motor_type_axis_3)
        
    # 視覺化呈現，它只會回應出呼叫那一刻的畫面給你，要它持續出現，需要寫個迴圈
    def render(self, mode='human'):
        return None
        
    def close(self):
        return None
        
    def dynamics_torque_limit(self):
        """
        Calculate the maximum torque required by

        each axis when the arm of each axis is the longest and the acceleration is the highest
        """
        torque = np.array([np.zeros(shape=6)])
        # axis_angle = np.array([np.zeros(shape=6)])
        axis_angle = []
        append_torque_limit_list = []
        temp_torque_max = []
        temp_torque_min = []
        Torque_Max = []
        # 角度轉換
        du = pi / 180
        # 度
        # radian = 180 / pi
        # 弧度
        self.robot.payload(self.payload, self.payload_position)  # set payload
        torque = np.array([np.zeros(shape=6)])
        q_list = [0, 90, -90, 180, -180]
        T_cell = (
            len(q_list)
            * len(q_list)
            * len(q_list)
            * len(q_list)
            * len(q_list)
            * len(q_list)
        )

        for i in range(len(q_list)):
            q1 = q_list[i]
            percent = i / T_cell * 100
            for j in range(len(q_list)):
                q2 = q_list[j]
                for k in range(len(q_list)):
                    q3 = q_list[k]
                    for l in range(len(q_list)):
                        q4 = q_list[l]
                        for m in range(len(q_list)):
                            q5 = q_list[m]
                            for n in range(len(q_list)):
                                q6 = q_list[n]
                                axis_angle.append([q1, q2, q3, q4, q5, q6])
                                load = np.array(
                                    [
                                        self.robot.rne(
                                            [
                                                q1 * du,
                                                q2 * du,
                                                q3 * du,
                                                q4 * du,
                                                q5 * du,
                                                q6 * du,
                                            ],
                                            self.vel,
                                            self.acc
                                        )
                                    ]
                                )
                                torque = np.append(torque, load, axis=0)

        for i in range(6):
            axis = i
            toque_max_index = np.argmax(torque[:, axis])
            toque_min_index = np.argmin(torque[:, axis])

            temp_torque_max = torque[toque_max_index].tolist()
            temp_torque_max.extend(axis_angle[toque_max_index])
            temp_torque_min = torque[toque_min_index].tolist()
            temp_torque_min.extend(axis_angle[toque_min_index])
            append_torque_limit_list.append(temp_torque_max)
            append_torque_limit_list.append(temp_torque_min)
            Torque_Max.append(abs(torque[toque_max_index][i]))
            self.torque_dynamics_limit = Torque_Max
        return self.torque_dynamics_limit
    
    def point_Workspace_cal_Monte_Carlo(self):
        """
        Through the "Work space" page in the interface to calculate of the robot
        """
        i = 0
        # 角度轉換
        du = pi / 180
        # 度
        radian = 180 / pi
        # 弧度

        self.q1_s = -180
        self.q1_end = 180
        self.q2_s = -50
        self.q2_end = 230
        self.q3_s = -150
        self.q3_end = 150
        self.q4_s = -180
        self.q4_end = 180
        self.q5_s = -180
        self.q5_end = 180
        self.q6_s = -180
        self.q6_end = 180
        N = 10 # 改為直接random 10個點
        theta1 = np.around(self.q1_s + (self.q1_end - self.q1_s) * np.random.rand(N, 1), 0)
        theta2 = np.around(self.q2_s + (self.q2_end - self.q2_s) * np.random.rand(N, 1), 0)
        theta3 = np.around(self.q3_s + (self.q3_end - self.q3_s) * np.random.rand(N, 1), 0)
        theta4 = np.around(self.q4_s + (self.q4_end - self.q4_s) * np.random.rand(N, 1), 0)
        theta5 = np.around(self.q5_s + (self.q5_end - self.q5_s) * np.random.rand(N, 1), 0)
        theta6 = np.around(self.q6_s + (self.q6_end - self.q6_s) * np.random.rand(N, 1), 0)

        for i in range(N):
            q1 = theta1[i, 0]
            q2 = theta2[i, 0]
            q3 = theta3[i, 0]
            q4 = theta4[i, 0]
            q5 = theta5[i, 0]
            q6 = theta6[i, 0]
            self.T = self.robot.fkine(
                [q1 * du, q2 * du, q3 * du, q4 * du, q5 * du, q6 * du]
            )
            t = np.around(self.T.t, 2)
            r = np.around(self.T.rpy('deg'), 0)
            self.T_x.append(t[0])
            self.T_y.append(t[1])
            self.T_z.append(t[2])
            self.T_roll.append(int(r[0]))
            self.T_pitch.append(int(r[1]))
            self.T_yaw.append(int(r[2]))
            i = i + 1
        
    def random_select_point(self):
        excel_file = Workbook()
        sheet = excel_file.active

        for i in range(10):
            # x = np.random.randint(1,999)
            sheet.cell(row=i + 1, column=1).value = self.T_x[i]
            sheet.cell(row=i + 1, column=2).value = self.T_y[i]
            sheet.cell(row=i + 1, column=3).value = self.T_z[i]
            sheet.cell(row=i + 1, column=4).value = self.T_roll[i]
            sheet.cell(row=i + 1, column=5).value = self.T_pitch[i]
            sheet.cell(row=i + 1, column=6).value = self.T_yaw[i]

        file_name = self.xlsx_outpath + "/task_point" +".xlsx"
        excel_file.save(file_name)

    def only_reachable_manipulability(self,model_select):
        if model_select == "train":
            # import xlsx
            df = load_workbook("./xlsx/task_point.xlsx")
        elif model_select == "test":
            df = load_workbook(self.point_test_excel)
        sheets = df.worksheets
        sheet1 = sheets[0]
        rows = sheet1.rows
        T_tmp = []
        i = 0
        count = 0
        manipulability_index = []
        for row in rows:
            row_val = [col.value for col in row]
            T_tmp.append(SE3(row_val[0], row_val[1], row_val[2]) * SE3.RPY([np.deg2rad(row_val[3]), np.deg2rad(row_val[4]), np.deg2rad(row_val[5])]))
            ik_q = self.robot.ikine_LMS(T=T_tmp[i])

            if ik_q.success == True:
                count += 1
                manipulability_index.append(self.robot.manipulability(q=ik_q.q))
            i = i + 1
        final_score = count / i
        if count == 0:
            return(0, 0)
        if count == 1:
            return(final_score, manipulability_index[0]) # 回傳 manipulability[0]
        else:
            return(final_score, np.mean(manipulability_index)) # 回傳 manipulability 取平均
    def performance_evaluate(self,model_select,axis2_motor_type, axis3_motor_type):
        if model_select == "train":
            # import xlsx
            df = load_workbook("./xlsx/task_point.xlsx")
        elif model_select == "test":
            # import xlsx
            # df = load_workbook("./xlsx/task_point_6dof_tested.xlsx")
            # df = load_workbook("./xlsx/task_point_6dof_tested_ori_random.xlsx")
            df = load_workbook(self.point_test_excel)
        sheets = df.worksheets
        # sheet1 = sheets[0]
        # rows = sheet1.rows
        # cols = sheet1.columns
        # T_tmp = []
        # score = []
        # ratio_over = 0
        # torque_over = 0
        # num_torque = np.array([np.zeros(shape=6)])
        # total_time = self.mission_time
        # # 采样间隔
        # sample_interval = 0.2
        # manipulability_index = []
        # i = 0
        # false_done = False
        # count = 0
        # for row in rows:
        #     row_val = [col.value for col in row]
        #     T_tmp.append(SE3(row_val[0], row_val[1], row_val[2]) * SE3.RPY([np.deg2rad(row_val[3]), np.deg2rad(row_val[4]), np.deg2rad(row_val[5])]))
        #     ik_q = self.robot.ikine_LMS(T=T_tmp[i])

        #     if ik_q.success == True:
        #         count += 1
        #         manipulability_index.append(self.robot.manipulability(q=ik_q.q))
        #     if i >= 1:
        #         # 生成时间向量
        #         traj_time = total_time/9 # 除以點位數量
        #         time_vector = np.linspace(0, traj_time, int(traj_time/sample_interval) + 1)
        #         traj = self.robot.jtraj(T_tmp[i-1],T_tmp[i],time_vector)

        #         if np.amax(traj.sd) > 3.04:
        #             ratio_over = ratio_over + 1
        #         torque = self.robot.rne(traj.s,traj.sd,traj.sdd)

        #         row = abs(torque[:,1]) # 取出第2行
        #         result_2 = row[row > axis2_motor_type] # 取出大于阈值的数字
        #         row = abs(torque[:,2]) # 取出第3行
        #         result_3 = row[row > axis3_motor_type] # 取出大于阈值的数字
        #         if len(result_2)>0 or len(result_3) >0:
        #             torque_over = torque_over + 1
        #         num_torque = np.append(num_torque, torque)
        #     i = i + 1

        # num_samples = int(total_time / sample_interval)
        # total_energy = 0
        # for j in range(num_samples):
        #     energy = abs(num_torque[j]) * sample_interval
        #     total_energy += energy
            
        # if count == 0:
        #     return(ratio_over, torque_over, total_energy,0,0)
        # else:
        #     final_score = count / i
        #     if count == 1:
        #         return(ratio_over, torque_over, total_energy, final_score, manipulability_index[0]) # 回傳 manipulability[0]
        #     else:
        #         return(ratio_over, torque_over, total_energy, final_score, np.mean(manipulability_index)) # 回傳 manipulability 取平均
        sheet1 = sheets[0]
        rows = sheet1.rows
        T_tmp = []
        T_traj = []
        ik_q_traj = []
        ratio_over = 0
        torque_over = 0
        num_torque = np.array([np.zeros(shape=6)])
        total_time = self.mission_time
        # 采样间隔
        sample_interval = 0.2
        manipulability_index = []
        i = 0
        count = 0
        max_diff = []
        max_diff_tol = 0
        traj_time = []
        diff = np.array([np.zeros(shape=6)])
        for row in rows:
            row_val = [col.value for col in row]
            T_tmp.append(SE3(row_val[0], row_val[1], row_val[2]) * SE3.RPY([np.deg2rad(row_val[3]), np.deg2rad(row_val[4]), np.deg2rad(row_val[5])]))
            ik_q = self.robot.ikine_LMS(T=T_tmp[i])

            if ik_q.success == True:
                count += 1
                T_traj.append(T_tmp[i])
                ik_q_traj.append(ik_q.q)
                manipulability_index.append(self.robot.manipulability(q=ik_q.q))
                # print("ik_q.q",ik_q.q)
                if count >= 2: # 兩個點位以上開始計算
                    diff = np.abs(np.subtract(ik_q_traj[-2], ik_q_traj[-1]))
                    max_diff_tol = max_diff_tol + np.max(diff)
                    max_diff.append(np.max(diff))
                    # print(max_diff)
            i = i + 1
        for k in range(len(max_diff)):
            traj_time.append(max_diff[k] / max_diff_tol * total_time)
        print("traj_time",traj_time)
        for m in range(len(max_diff)):
            time_vector = np.linspace(0, traj_time[m], int(traj_time[m]/sample_interval) + 1)
            traj = self.robot.jtraj(T_traj[m],T_traj[m+1],time_vector)
            # print(traj.s)
            if np.amax(traj.sd) > 3.04:
                ratio_over = ratio_over + 1
            torque = self.robot.rne(traj.s,traj.sd,traj.sdd)
            row = abs(torque[:,1]) # 取出第2行
            result_2 = row[row > axis2_motor_type] # 取出大于阈值的数字
            row = abs(torque[:,2]) # 取出第3行
            result_3 = row[row > axis3_motor_type] # 取出大于阈值的数字
            if len(result_2)>0 or len(result_3) >0:
                torque_over = torque_over + 1
            num_torque = np.append(num_torque, torque)
        total_energy = 0
        for j in range(len(num_torque)):
            energy = abs(num_torque[j]) * sample_interval
            total_energy += energy
                
        if count == 0:
            return(0, 0, 0,0,0)
        else:
            final_score = count / i
            if count == 1:
                return(0, 0, 0, final_score, manipulability_index[0]) # 回傳 manipulability[0]
            else:
                return(ratio_over, torque_over, total_energy, final_score, np.mean(manipulability_index)) # 回傳 manipulability 取平均


class RobotOptEnv_3dof(gym.Env):
    metadata = {
        'render.modes': ['human', 'rgb_array'],
        'video.frames_per_second': 2
    }
    def __init__(self):
        self.robot = modular_robot_3dof() # TODO: fixed 3dof
        self.robot_urdf = stl_conv_urdf("single_arm_v12","test")
        
        # callback:Enter the parameters of the algorithm to be optimized on the interface
        self.sub_optimal_design = rospy.Subscriber(
            "/optimal_design", optimal_design, self.optimal_design_callback
        )
        # 使用者設定參數
        self.payload = 5.0
        self.payload_position = np.array([0, 0, 0.04])
        self.vel = np.array([2.356194, 2.356194, 2.356194])# TODO: fixed 3dof
        self.acc = np.array([2.356194, 2.356194, 2.356194])# TODO: fixed 3dof
        self.total_weight = 20 # Kg
        self.total_cost = 1800 # 元
        # 預設二,三軸軸長
        self.std_L2 = 35.0 # 預設標準值 第二軸 35 cm
        self.std_L3 = 35.0 # 預設標準值 第三軸 35 cm
        # 觀察參數 motor
        self.motor = motor_data()
        self.res = self.motor.dynamixel_member
        self.high_torque = float('inf') # 預設標準值 馬達極限 120.0 N max
        self.low_torque = float('-inf') # 預設標準值 馬達極限 -12.0 N max
        self.motor_cost_init = np.array([0,400,400,0,0,0], dtype=np.float64) # 預設最大馬達費用
        self.motor_weight_init = np.array([0,0.855,0.855,0,0,0], dtype=np.float64) # 預設最大馬達重量
        self.motor_cost = np.array([0,400,400,0,0,0], dtype=np.float64) # 馬達費用 # TODO: fixed 3dof
        self.motor_weight = np.array([0,0.855,0.855], dtype=np.float64) # 馬達重量# TODO: fixed 3dof
        self.motor_rated = np.array([44.7,44.7,44.7], dtype=np.float64)# TODO: fixed 3dof
        # 使用者設定參數 & 觀察參數
        self.reach_distance = 0.6 # 使用者設定可達半徑最小值
        self.high_reach_eva = 1 # 預設觀測標準值
        self.low_reach_eva = 0 # 預設觀測標準值
        self.high_manipulability = 1  # 預設觀測標準值
        self.low_manipulability = 0  # 預設觀測標準值
        self.high_std_L2 = 70 # 預設觀測標準值
        self.low_std_L2 = -25 # 預設觀測標準值
        self.high_std_L3 = 70 # 預設觀測標準值
        self.low_std_L3 = -25 # 預設觀測標準值
        self.high_ratio_over = 10 # 預設觀測標準值
        self.low_ratio_over = 0 # 預設觀測標準值
        self.high_torque_over = 10 # 預設觀測標準值
        self.low_torque_over = 0 # 預設觀測標準值
        self.high_power_consumption = float('inf')
        self.low_power_consumption = float('-inf')
        self.torque_done = np.array([false, false, false])# TODO: fixed 3dof
        self.torque_over = False
        self.ratio_over = False
        self.prev_shaping = None
        self.motor_type_axis_2 = 5.1
        self.motor_type_axis_3 = 5.1
        # FIXME:
        self.action_select = 'fixed' 
        self.point_test_excel = './xlsx/task_point_6dof_tested_ori_random'
        # TODO: 增加馬達模組選型action
        self.action_space = spaces.Discrete(12) # TODO: fixed 12種action
        
        # TODO: observation space for torque, reach, motor cost, weight, manipulability
        self.observation_space = spaces.Box(np.array([self.low_ratio_over,self.low_torque_over,self.low_power_consumption, self.low_reach_eva, self.low_manipulability, self.low_std_L2, self.low_std_L3 ]), 
                                            np.array([self.high_ratio_over,self.high_torque_over,self.high_power_consumption, self.high_reach_eva, self.high_manipulability, self.high_std_L2, self.high_std_L3]), 
                                            dtype=np.float64)
        # TODO: reward 歸一化
        self.state = np.array([0,0,0,0,0,0,0], dtype=np.float64)
        self.pre_state = np.array([0,0,0,0,0,0,0], dtype=np.float64)
        #隨機抽樣點位初始化
        self.T_x = []
        self.T_y = []
        self.T_z = []
        self.T_roll = []
        self.T_pitch = []
        self.T_yaw = []
        self.xlsx_outpath = "./xlsx/"

        self.model_select = "train" # 選擇使用train or test model 

    def optimal_design_callback(self, data):
        # print(data.data)
        # TODO: 目標構型
        self.op_dof = data.dof
        self.op_payload = data.payload
        self.op_payload_position = data.payload_position
        self.op_vel = data.vel
        self.op_acc = data.acc
        self.op_radius = data.radius
        self.op_weight = data.arm_weight
        self.op_cost = data.cost
        
        print("op_dof:", self.op_dof)
        print("op_payload:", self.op_payload)
        print("op_payload_position:", self.op_payload_position)
        print("op_vel:", self.op_vel)
        print("op_acc:", self.op_acc)
        print("op_radius:", self.op_radius)
        print("op_weight:", self.op_weight)
        print("op_cost:", self.op_cost)
        
        self.payload = self.op_payload
        self.payload_position = np.array(self.op_payload_position)
        self.vel = np.array(self.op_vel[0:3])# TODO: fixed 3dof
        self.acc = np.array(self.op_acc[0:3])# TODO: fixed 3dof
        self.total_weight = self.op_weight # Kg
        self.total_cost = self.op_cost # 元
        self.reach_distance = self.op_radius # 使用者設定可達半徑最小值
        
    # TODO: fixed
    def step(self, action):
        assert self.action_space.contains(action), "%r (%s) invalid"%(action, type(action))
        # self.pre_state[0:3] = self.state[0:3]# TODO: fixed 3dof
        # TODO: 向量編碼動作
        # TODO: 向量編碼動作
        if action == 0: # 軸2  # 短 # 型號1
            self.std_L2 -= 1.0
            self.std_L3 += 1.0
            # 配置軸2 motor 型號1
            #[5.1, 25.3, 44.7]
            self.motor_type_axis_2 = 5.1 # 型號
            
        elif action == 1: # 軸2  # 長 # 型號1
            self.std_L2 += 1.0
            self.std_L3 -= 1.0
            self.motor_type_axis_2 = 5.1 # 型號
            
        elif action == 2: # 軸2  # 短 # 型號2
            self.std_L2 -= 1.0
            self.std_L3 += 1.0
            self.motor_type_axis_2 = 25.3 # 型號

        elif action == 3: # 軸2  # 長 # 型號2
            self.std_L2 += 1.0
            self.std_L3 -= 1.0
            self.motor_type_axis_2 = 25.3 # 型號
            
        elif action == 4: # 軸2  # 短 # 型號3
            self.std_L2 -= 1.0
            self.std_L3 += 1.0
            self.motor_type_axis_2 = 44.7 # 型號

        elif action == 5: # 軸2  # 長 # 型號3
            self.std_L2 += 1.0
            self.std_L3 -= 1.0
            self.motor_type_axis_2 = 44.7 # 型號
            
        elif action == 6: # 軸3  # 短 # 型號1
            self.std_L3 -= 1.0
            self.std_L3 += 1.0
            self.motor_type_axis_3 = 5.1 # 型號
            
        elif action == 7: # 軸3  # 長 # 型號1
            self.std_L3 += 1.0
            self.std_L3 -= 1.0
            self.motor_type_axis_3 = 5.1 # 型號
            
        elif action == 8: # 軸3  # 短 # 型號2
            self.std_L3 -= 1.0
            self.std_L3 += 1.0
            self.motor_type_axis_3 = 25.3 # 型號
            
        elif action == 9: # 軸3  # 長 # 型號2
            self.std_L3 += 1.0
            self.std_L3 -= 1.0
            self.motor_type_axis_3 = 25.3 # 型號

        elif action == 10: # 軸3  # 短 # 型號3
            self.std_L3 -= 1.0
            self.std_L3 += 1.0
            self.motor_type_axis_3 = 44.7 # 型號

        elif action == 11: # 軸3  # 長 # 型號3
            self.std_L3 += 1.0
            self.std_L3 -= 1.0
            self.motor_type_axis_3 = 44.7 # 型號
        
        # 輸入action後 二,三軸軸長
        self.robot_urdf.specified_generate_write_urdf(self.std_L2, self.std_L3)
        self.robot.__init__() # 重製機器人
        # FIXME: add set payload
        self.robot.payload(self.payload, self.payload_position)  # set payload
        ratio_over, torque_over, consumption, reach_score, manipulability_score = self.performance_evaluate(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
        self.state[0] = ratio_over
        self.state[1] = torque_over
        self.state[2] = np.abs(consumption)
        self.state[3] = reach_score
        self.state[4] = manipulability_score
        self.state[5] = self.std_L2
        self.state[6] = self.std_L3
        self.counts += 1
        reward = 0
        # TODO: fixed
        shaping = (
            # - self.state[0] # 超過轉速量
            - self.state[2] # 功耗
            # + 100 * self.state[3] # 可達性
            + 1000 * self.state[4] # 可操作性
        ) 
        rospy.loginfo("shaping: %s", shaping)
        if self.prev_shaping is not None:
            reward = shaping - self.prev_shaping
        self.prev_shaping = shaping
        

        if self.state[0] > 0:
            self.ratio_over = True

        if self.state[1] > 0:
            self.torque_over = True

        terminated = False
    
        # 避免軸長小於0
        if self.state[5] <= 0 or self.state[6] <=  0:
            # terminated = True
            reward += -200
        if self.ratio_over == True or self.torque_over == True and self.state[3] < 0.6: 
            reward += -20
        elif self.ratio_over == True or self.torque_over == True and 0.6 <= self.state[3] < 0.8: 
            reward += -10
        elif self.ratio_over == True or self.torque_over == True and 0.8 <= self.state[3] <= 1.0:   
            reward += -5
        elif self.ratio_over == False and self.torque_over == False and self.state[3] < 0.6: 
            reward += +5
        elif self.ratio_over == False and self.torque_over == False and 0.6 <= self.state[3] < 0.8:
            reward += +10
        elif self.ratio_over == False and self.torque_over == False and 0.8 <= self.state[3] < 0.9:
            reward += +20
        elif self.ratio_over == False and self.torque_over == False and 0.9 <= self.state[3] < 1.0:
            # terminated = True
            reward += +50
        # success
        elif self.ratio_over == False and self.torque_over == False and self.state[3] == 1.0:
            terminated = True
            reward += +100
            self.counts = 0
        
        if self.counts == 50: # max_steps
            terminated = True
            self.counts = 0
        
        self.ratio_over = False
        self.torque_over = False #reset
        rospy.loginfo("counts: %s", self.counts)
        rospy.loginfo("step_reward: %s", reward)
        rospy.loginfo("step_reward: %s", self.state)
        rospy.loginfo("================================")
        # print("================================")
        current_design = [self.std_L2, self.std_L3, self.motor_rated[1], self.motor_rated[2]]
        return self.state, reward, terminated, current_design

    # reset环境状态 
    def reset(self):
        if self.model_select == "train":
            rospy.loginfo("model_select:%s", self.model_select)
            # random state (手臂長度隨機)
            random_total_arm_length = np.random.uniform(low=20, high=80)
            self.std_L2, self.std_L3 = self.robot_urdf.opt_specify_random_generate_write_urdf(random_total_arm_length) # 啟用隨機的L2,L3長度urdf, 並指定總臂長
            # self.std_L2, self.std_L3 = self.robot_urdf.opt_random_generate_write_urdf() # 啟用隨機的L2,L3長度urdf
            self.robot.__init__() # 重製機器人
            self.motor_type_axis_2 = 5.1
            self.motor_type_axis_3 = 5.1
            self.mission_time = np.random.uniform(low = 20, high = 50)
            rospy.loginfo("random_total_arm_length: %s", random_total_arm_length)
            rospy.loginfo("std_L2: %s", self.std_L2)
            rospy.loginfo("std_L3: %s", self.std_L3)
            rospy.loginfo("mission_time: %s", self.mission_time)
            # 生成隨機 payload (kg)
            rand_payload = np.random.uniform(low=1, high=4)
            # FIXME: 修改未成功匯入payload給予機器人的問題
            self.payload = rand_payload
            self.payload_position =  [0, 0, 0.04]
            self.robot.payload(self.payload, self.payload_position)  # set payload
            self.point_Workspace_cal_Monte_Carlo() # 在當前reset出來的機械手臂構型下, 生成點位
            self.random_select_point() # 先隨機抽樣30個點位
            ratio_over, torque_over, consumption, reach_score, manipulability_score = self.performance_evaluate(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
            self.prev_shaping = None
            self.state[0] = ratio_over
            self.state[1] = torque_over
            self.state[2] = np.abs(consumption)
            self.state[3] = reach_score
            self.state[4] = manipulability_score
            self.state[5] = self.std_L2
            self.state[6] = self.std_L3
            self.counts = 0
            return self.state
        elif self.model_select == "test":
            # random state (手臂長度隨機)
            rospy.loginfo("model_select:%s", self.model_select)
            total_arm_length = self.op_radius
            self.std_L2, self.std_L3 = self.robot_urdf.opt_specify_random_generate_write_urdf(total_arm_length) # 啟用隨機的L2,L3長度urdf, 並指定總臂長
            # self.std_L2, self.std_L3 = self.robot_urdf.opt_random_generate_write_urdf() # 啟用隨機的L2,L3長度urdf
            self.robot.__init__() # 重製機器人
            # FIXME: 修改未成功匯入payload給予機器人的問題
            self.payload = self.op_payload
            self.payload_position = self.op_payload_position
            self.robot.payload(self.payload, self.payload_position)
            rospy.loginfo("mission_time: %s", self.mission_time)
            self.total_weight = self.op_weight # Kg
            self.total_cost = self.op_cost # 元
            self.reach_distance = self.op_radius # 使用者設定可達半徑最小值
            self.motor_type_axis_2 = 5.1
            self.motor_type_axis_3 = 5.1
            ratio_over, torque_over, consumption, reach_score, manipulability_score = self.performance_evaluate(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
            self.prev_shaping = None
            self.state[0] = ratio_over
            self.state[1] = torque_over
            self.state[2] = np.abs(consumption)
            self.state[3] = reach_score
            self.state[4] = manipulability_score
            self.state[5] = self.std_L2
            self.state[6] = self.std_L3
            self.counts = 0
            return self.state        
    def original_design(self,std_L2,std_L3, motor_1, motor_2, payload, mission_time):
        #  指定手臂長度
        self.robot_urdf.specified_generate_write_urdf(std_L2, std_L3)
        self.robot.__init__() # 重製機器人
        motor_type_axis_2 = motor_1
        motor_type_axis_3 = motor_2
        self.payload = payload
        op_payload_position = [0, 0, 0.04]
        self.payload_position = [np.array(op_payload_position)]
        self.robot.payload(self.payload, op_payload_position)  # set payload
        # mission_time = 30
        model_select = "test"
        # rospy.loginfo("mission_time: %s", mission_time)
        self.point_Workspace_cal_Monte_Carlo() # 在當前reset出來的機械手臂構型下, 生成點位
        self.random_select_point() # 先隨機抽樣30個點位
        self.mission_time = mission_time
        ratio_over, torque_over, consumption, reach_score, manipulability_score = self.performance_evaluate(model_select, motor_type_axis_2, motor_type_axis_3)
        # ratio_over, torque_over, consumption = self.power_consumption(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
        # torque = self.dynamics_torque_limit()
        rospy.loginfo("ratio_over: %s", ratio_over)
        rospy.loginfo("torque_over: %s", torque_over)
        rospy.loginfo("consumption: %s", np.abs(consumption))
        # 生成隨機 payload (kg)
        rospy.loginfo("reach_score: %s", reach_score)
        rospy.loginfo("manipulability_score: %s", manipulability_score)
        rospy.loginfo("std_L2: %s", std_L2)
        rospy.loginfo("std_L3: %s", std_L3)

    # 視覺化呈現，它只會回應出呼叫那一刻的畫面給你，要它持續出現，需要寫個迴圈
    def render(self, mode='human'):
        return None
        
    def close(self):
        return None
        
    def dynamics_torque_limit(self):
        """
        Calculate the maximum torque required by

        each axis when the arm of each axis is the longest and the acceleration is the highest
        """
        torque = np.array([np.zeros(shape=3)])# TODO: fixed 3dof
        # axis_angle = np.array([np.zeros(shape=6)])
        axis_angle = []
        append_torque_limit_list = []
        temp_torque_max = []
        temp_torque_min = []
        Torque_Max = []
        # 角度轉換
        du = pi / 180
        # 度
        # radian = 180 / pi
        # 弧度
        self.robot.payload(self.payload, self.payload_position)  # set payload
        torque = np.array([np.zeros(shape=3)])# TODO: fixed 3dof
        q_list = [0, 90, -90, 180, -180]
        # TODO: fixed 3dof
        T_cell = (
            len(q_list)
            * len(q_list)
            * len(q_list)
        )
        # TODO: fixed 3dof
        for i in range(len(q_list)):
            q1 = q_list[i]
            percent = i / T_cell * 100
            for j in range(len(q_list)):
                q2 = q_list[j]
                for k in range(len(q_list)):
                    q3 = q_list[k]
                    axis_angle.append([q1, q2, q3])
                    load = np.array(
                        [
                            self.robot.rne(
                                [
                                    q1 * du,
                                    q2 * du,
                                    q3 * du,
                                    # q4 * du,
                                    # q5 * du,
                                    # q6 * du,
                                ],
                                self.vel,
                                self.acc
                            )
                        ]
                    )
                    torque = np.append(torque, load, axis=0)
        # TODO: fixed 3dof
        for i in range(3):
            axis = i
            toque_max_index = np.argmax(torque[:, axis])
            toque_min_index = np.argmin(torque[:, axis])

            temp_torque_max = torque[toque_max_index].tolist()
            temp_torque_max.extend(axis_angle[toque_max_index])
            temp_torque_min = torque[toque_min_index].tolist()
            temp_torque_min.extend(axis_angle[toque_min_index])
            append_torque_limit_list.append(temp_torque_max)
            append_torque_limit_list.append(temp_torque_min)
            Torque_Max.append(abs(torque[toque_max_index][i]))
            self.torque_dynamics_limit = Torque_Max
        return self.torque_dynamics_limit
    
    def point_Workspace_cal_Monte_Carlo(self):
        """
        Through the "Work space" page in the interface to calculate of the robot
        """
        i = 0
        # 角度轉換
        du = pi / 180
        # 度
        radian = 180 / pi
        # 弧度

        self.q1_s = -180
        self.q1_end = 180
        self.q2_s = -50
        self.q2_end = 230
        self.q3_s = -150
        self.q3_end = 150
        # TODO: fixed 3dof
        # self.q4_s = -160
        # self.q4_end = 160
        # self.q5_s = -160
        # self.q5_end = 160
        # self.q6_s = -160
        # self.q6_end = 160
        N = 10
        theta1 = self.q1_s + (self.q1_end - self.q1_s) * np.random.rand(N, 1)
        theta2 = self.q2_s + (self.q2_end - self.q2_s) * np.random.rand(N, 1)
        theta3 = self.q3_s + (self.q3_end - self.q3_s) * np.random.rand(N, 1)
        # TODO: fixed 3dof
        # theta4 = self.q4_end + (self.q4_end - self.q4_s) * np.random.rand(N, 1)
        # theta5 = self.q5_end + (self.q5_end - self.q5_s) * np.random.rand(N, 1)
        # theta6 = self.q6_end + (self.q6_end - self.q6_s) * np.random.rand(N, 1)
        
        for i in range(N):
            q1 = theta1[i, 0]
            q2 = theta2[i, 0]
            q3 = theta3[i, 0]
            # TODO: fixed 3dof
            # q4 = theta4[i, 0]
            # q5 = theta5[i, 0]
            # q6 = theta6[i, 0]
            self.T = self.robot.fkine(
                [q1 * du, q2 * du, q3 * du]# TODO: fixed 3dof
            )
            t = np.round(self.T.t, 3)
            # r = np.round(self.T.rpy('deg'), 3)# TODO: fixed 3dof
            self.T_x.append(t[0])
            self.T_y.append(t[1])
            self.T_z.append(t[2])
            # TODO: fixed 3dof
            # self.T_roll.append(int(r[0]))
            # self.T_pitch.append(int(r[1]))
            # self.T_yaw.append(int(r[2]))
            i = i + 1
        
    def random_select_point(self):
        excel_file = Workbook()
        sheet = excel_file.active

        for i in range(10):
            # x = np.random.randint(1,999)
            sheet.cell(row=i + 1, column=1).value = self.T_x[i]
            sheet.cell(row=i + 1, column=2).value = self.T_y[i]
            sheet.cell(row=i + 1, column=3).value = self.T_z[i]
            # TODO: fixed 3dof
            # sheet.cell(row=i + 1, column=4).value = self.T_roll[x]
            # sheet.cell(row=i + 1, column=5).value = self.T_pitch[x]
            # sheet.cell(row=i + 1, column=6).value = self.T_yaw[x]

        file_name = self.xlsx_outpath + "/task_point_3dof" +".xlsx"# TODO: fixed 3dof
        excel_file.save(file_name)

    def performance_evaluate(self,model_select,axis2_motor_type, axis3_motor_type):
        if model_select == "train":
            # import xlsx
            df = load_workbook("./xlsx/task_point_3dof.xlsx")
        elif model_select == "test":
            # import xlsx
            df = load_workbook("./xlsx/task_point_3dof_tested.xlsx")
        sheets = df.worksheets
        sheet1 = sheets[0]
        rows = sheet1.rows
        cols = sheet1.columns
        T_tmp = []
        score = []
        ratio_over = 0
        torque_over = 0
        num_torque = np.array([np.zeros(shape=3)]) # TODO: fixed 3dof
        total_time = self.mission_time
        # 采样间隔
        sample_interval = 0.2
        manipulability_index = []
        i = 0
        false_done = False
        count = 0
        for row in rows:
            row_val = [col.value for col in row]
            # T_tmp.append(SE3(row_val[0], row_val[1], row_val[2]) * SE3.RPY([np.deg2rad(row_val[3]), np.deg2rad(row_val[4]), np.deg2rad(row_val[5])]))
            # ik_q = self.robot.ikine_LMS(T=T_tmp[i])
            T_tmp.append(SE3(row_val[0], row_val[1], row_val[2]))# TODO: fixed 3dof
            ik_q = self.robot.ikine_LMS(T=T_tmp[i], mask = [1,1,1,0,0,0])# TODO: fixed 3dof
            if ik_q.success == True:
                count += 1
                manipulability_index.append(self.robot.manipulability(q=ik_q.q))
            if i >= 1:
                # 生成时间向量
                traj_time = total_time/9 # 除以點位數量
                time_vector = np.linspace(0, traj_time, int(traj_time/sample_interval) + 1)
                traj = self.robot.jtraj(T_tmp[i-1],T_tmp[i],time_vector)
            
                if np.amax(traj.sd) > 3.04:
                    ratio_over = ratio_over + 1
                torque = self.robot.rne(traj.s,traj.sd,traj.sdd)

                row = abs(torque[:,1]) # 取出第2行
                result_2 = row[row > axis2_motor_type] # 取出大于阈值的数字
                row = abs(torque[:,2]) # 取出第3行
                result_3 = row[row > axis3_motor_type] # 取出大于阈值的数字
                if len(result_2)>0 or len(result_3) >0:
                    torque_over = torque_over + 1
                num_torque = np.append(num_torque, torque)
            i = i + 1

        num_samples = int(total_time / sample_interval)
        total_energy = 0
        for j in range(num_samples):
            energy = abs(num_torque[j]) * sample_interval
            total_energy += energy
            
        if count == 0:
            return(ratio_over, torque_over, total_energy,0,0)
        else:
            final_score = count / i
            if count == 1:
                return(ratio_over, torque_over, total_energy, final_score, manipulability_index[0]) # 回傳 manipulability[0]
            else:
                return(ratio_over, torque_over, total_energy, final_score, np.mean(manipulability_index)) # 回傳 manipulability 取平均
       
class RobotOptEnv_5dof(gym.Env):
    metadata = {
        'render.modes': ['human', 'rgb_array'],
        'video.frames_per_second': 2
    }
    def __init__(self):
        self.robot = modular_robot_5dof() # TODO: fixed 5dof
        self.robot_urdf = stl_conv_urdf("single_arm_v12","test")
        # self.robot_urdf.init_dynamixel_diff_inertia()
        # callback:Enter the parameters of the algorithm to be optimized on the interface
        self.sub_optimal_design = rospy.Subscriber(
            "/optimal_design", optimal_design, self.optimal_design_callback
        )
        # 使用者設定參數
        self.payload = 5.0
        self.payload_position = np.array([0, 0, 0.04])
        self.vel = np.array([2.356194, 2.356194, 2.356194, 2.356194, 2.356194]) # TODO: fixed 5dof
        self.acc = np.array([2.356194, 2.356194, 2.356194, 2.356194, 2.356194]) # TODO: fixed 5dof
        self.total_weight = 20 # Kg
        self.total_cost = 1800 # 元
        # 預設二,三軸軸長
        self.std_L2 = 35.0 # 預設標準值 第二軸 35 cm
        self.std_L3 = 35.0 # 預設標準值 第三軸 35 cm
        # 觀察參數 motor
        self.motor = motor_data()
        self.res = self.motor.dynamixel_member
        self.high_torque = float('inf') # 預設標準值 馬達極限 120.0 N max
        self.low_torque = float('-inf') # 預設標準值 馬達極限 -12.0 N max
        self.motor_cost_init = np.array([0,400,400,0,0,0], dtype=np.float64) # 預設最大馬達費用
        self.motor_weight_init = np.array([0,0.855,0.855,0,0,0], dtype=np.float64) # 預設最大馬達重量
        self.motor_cost = np.array([0,400,400,0,0], dtype=np.float64) # 馬達費用 # TODO: fixed 5dof
        self.motor_weight = np.array([0,0.855,0.855,0,0], dtype=np.float64) # 馬達重量 # TODO: fixed 5dof
        self.motor_rated = np.array([44.7,44.7,44.7,44.7,44.7], dtype=np.float64) # TODO: fixed 5dof
        # 使用者設定參數 & 觀察參數
        self.reach_distance = 0.6 # 使用者設定可達半徑最小值
        self.high_reach_eva = 1 # 預設觀測標準值
        self.low_reach_eva = 0 # 預設觀測標準值
        self.high_manipulability = 1  # 預設觀測標準值
        self.low_manipulability = 0  # 預設觀測標準值
        self.high_std_L2 = 70 # 預設觀測標準值
        self.low_std_L2 = -25 # 預設觀測標準值
        self.high_std_L3 = 70 # 預設觀測標準值
        self.low_std_L3 = -25 # 預設觀測標準值
        self.high_ratio_over = 10 # 預設觀測標準值
        self.low_ratio_over = 0 # 預設觀測標準值
        self.high_torque_over = 10 # 預設觀測標準值
        self.low_torque_over = 0 # 預設觀測標準值
        self.high_power_consumption = float('inf')
        self.low_power_consumption = float('-inf')
        self.torque_done = np.array([false, false, false])# TODO: fixed 5dof
        self.torque_over = False
        self.ratio_over = False
        self.prev_shaping = None
        self.motor_type_axis_2 = 5.1
        self.motor_type_axis_3 = 5.1
        # FIXME:
        self.action_select = 'fixed' 
        self.point_test_excel = './xlsx/task_point_6dof_tested_ori_random'
        # TODO: 增加馬達模組選型action
        self.action_space = spaces.Discrete(12) # TODO: fixed 12種action
        
        # TODO: observation space for torque, reach, motor cost, weight, manipulability
        self.observation_space = spaces.Box(np.array([self.low_ratio_over,self.low_torque_over,self.low_power_consumption, self.low_reach_eva, self.low_manipulability, self.low_std_L2, self.low_std_L3 ]), 
                                            np.array([self.high_ratio_over,self.high_torque_over,self.high_power_consumption, self.high_reach_eva, self.high_manipulability, self.high_std_L2, self.high_std_L3]), 
                                            dtype=np.float64)
        # TODO: reward 歸一化
        self.state = np.array([0,0,0,0,0,0,0], dtype=np.float64)
        self.pre_state = np.array([0,0,0,0,0,0,0], dtype=np.float64)

        #隨機抽樣點位初始化
        self.T_x = []
        self.T_y = []
        self.T_z = []
        self.T_roll = []
        self.T_pitch = []
        self.T_yaw = []
        self.xlsx_outpath = "./xlsx/"

        self.model_select = "train" # 選擇使用train or test model 

        # self.op_dof = None
        # self.op_payload = None
        # self.op_payload_position = None
        # self.op_vel = None
        # self.op_acc = None
        # self.op_radius = None
        # self.op_weight = None
        # self.op_cost = None

    def optimal_design_callback(self, data):
        # print(data.data)
        # TODO: 目標構型
        self.op_dof = data.dof
        self.op_payload = data.payload
        self.op_payload_position = data.payload_position
        self.op_vel = data.vel
        self.op_acc = data.acc
        self.op_radius = data.radius
        self.op_weight = data.arm_weight
        self.op_cost = data.cost
        
        print("op_dof:", self.op_dof)
        print("op_payload:", self.op_payload)
        print("op_payload_position:", self.op_payload_position)
        print("op_vel:", self.op_vel)
        print("op_acc:", self.op_acc)
        print("op_radius:", self.op_radius)
        print("op_weight:", self.op_weight)
        print("op_cost:", self.op_cost)
        
        self.payload = self.op_payload
        self.payload_position = np.array(self.op_payload_position)
        self.vel = np.array(self.op_vel[0:5]) # TODO: fixed 5dof
        self.acc = np.array(self.op_acc[0:5]) # TODO: fixed 5dof
        self.total_weight = self.op_weight # Kg
        self.total_cost = self.op_cost # 元
        self.reach_distance = self.op_radius # 使用者設定可達半徑最小值
        
    # TODO: fixed
    def step(self, action):
        assert self.action_space.contains(action), "%r (%s) invalid"%(action, type(action))
        if action == 0: # 軸2  # 短 # 型號1
            self.std_L2 -= 1.0
            self.std_L3 += 1.0
            # 配置軸2 motor 型號1
            #[5.1, 25.3, 44.7]
            self.motor_type_axis_2 = 5.1 # 型號
            
        elif action == 1: # 軸2  # 長 # 型號1
            self.std_L2 += 1.0
            self.std_L3 -= 1.0
            self.motor_type_axis_2 = 5.1 # 型號
            
        elif action == 2: # 軸2  # 短 # 型號2
            self.std_L2 -= 1.0
            self.std_L3 += 1.0
            self.motor_type_axis_2 = 25.3 # 型號

        elif action == 3: # 軸2  # 長 # 型號2
            self.std_L2 += 1.0
            self.std_L3 -= 1.0
            self.motor_type_axis_2 = 25.3 # 型號
            
        elif action == 4: # 軸2  # 短 # 型號3
            self.std_L2 -= 1.0
            self.std_L3 += 1.0
            self.motor_type_axis_2 = 44.7 # 型號

        elif action == 5: # 軸2  # 長 # 型號3
            self.std_L2 += 1.0
            self.std_L3 -= 1.0
            self.motor_type_axis_2 = 44.7 # 型號
            
        elif action == 6: # 軸3  # 短 # 型號1
            self.std_L3 -= 1.0
            self.std_L3 += 1.0
            self.motor_type_axis_3 = 5.1 # 型號
            
        elif action == 7: # 軸3  # 長 # 型號1
            self.std_L3 += 1.0
            self.std_L3 -= 1.0
            self.motor_type_axis_3 = 5.1 # 型號
            
        elif action == 8: # 軸3  # 短 # 型號2
            self.std_L3 -= 1.0
            self.std_L3 += 1.0
            self.motor_type_axis_3 = 25.3 # 型號
            
        elif action == 9: # 軸3  # 長 # 型號2
            self.std_L3 += 1.0
            self.std_L3 -= 1.0
            self.motor_type_axis_3 = 25.3 # 型號

        elif action == 10: # 軸3  # 短 # 型號3
            self.std_L3 -= 1.0
            self.std_L3 += 1.0
            self.motor_type_axis_3 = 44.7 # 型號

        elif action == 11: # 軸3  # 長 # 型號3
            self.std_L3 += 1.0
            self.std_L3 -= 1.0
            self.motor_type_axis_3 = 44.7 # 型號
        
        # 輸入action後 二,三軸軸長
        self.robot_urdf.specified_generate_write_urdf(self.std_L2, self.std_L3)
        self.robot.__init__() # 重製機器人
        # FIXME: 修改未成功匯入payload給予機器人的問題
        self.robot.payload(self.payload, self.payload_position)  # set payload
        ratio_over, torque_over, consumption, reach_score, manipulability_score = self.performance_evaluate(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
        self.state[0] = ratio_over
        self.state[1] = torque_over
        self.state[2] = np.abs(consumption)
        self.state[3] = reach_score
        self.state[4] = manipulability_score
        self.state[5] = self.std_L2
        self.state[6] = self.std_L3
        self.counts += 1
        reward = 0

        
        # TODO: fixed
        shaping = (
            # - self.state[0] # 超過轉速量
            - self.state[2] # 功耗
            # + 100 * self.state[3] # 可達性
            + 1000 * self.state[4] # 可操作性
        ) 
        rospy.loginfo("shaping: %s", shaping)
        if self.prev_shaping is not None:
            reward = shaping - self.prev_shaping
        self.prev_shaping = shaping
        

        if self.state[0] > 0:
            self.ratio_over = True

        if self.state[1] > 0:
            self.torque_over = True

        terminated = False
    
        # 避免軸長小於0
        if self.state[5] <= 0 or self.state[6] <=  0:
            # terminated = True
            reward += -200
        if self.ratio_over == True or self.torque_over == True and self.state[3] < 0.6: 
            reward += -20
        elif self.ratio_over == True or self.torque_over == True and 0.6 <= self.state[3] < 0.8: 
            reward += -10
        elif self.ratio_over == True or self.torque_over == True and 0.8 <= self.state[3] <= 1.0:   
            reward += -5
        elif self.ratio_over == False and self.torque_over == False and self.state[3] < 0.6: 
            reward += +5
        elif self.ratio_over == False and self.torque_over == False and 0.6 <= self.state[3] < 0.8:
            reward += +10
        elif self.ratio_over == False and self.torque_over == False and 0.8 <= self.state[3] < 0.9:
            reward += +20
        elif self.ratio_over == False and self.torque_over == False and 0.9 <= self.state[3] < 1.0:
            # terminated = True
            reward += +50
        # success
        elif self.ratio_over == False and self.torque_over == False and self.state[3] == 1.0:
            terminated = True
            reward += +100
            self.counts = 0
        
        if self.counts == 50: # max_steps
            terminated = True
            self.counts = 0
        
        self.ratio_over = False
        self.torque_over = False #reset
        rospy.loginfo("counts: %s", self.counts)
        rospy.loginfo("step_reward: %s", reward)
        rospy.loginfo("step_reward: %s", self.state)
        rospy.loginfo("================================")
        # print("================================")
        current_design = [self.std_L2, self.std_L3, self.motor_rated[1], self.motor_rated[2]]
        return self.state, reward, terminated, current_design
    # reset环境状态 
    # TODO: 修改為目前最新的計算方式 reset
    # FIXME: 修改未成功匯入payload給予機器人的問題
    def reset(self):
        if self.model_select == "train":
            rospy.loginfo("model_select:%s", self.model_select)
            # random state (手臂長度隨機)
            random_total_arm_length = np.random.uniform(low=20, high=80)
            self.std_L2, self.std_L3 = self.robot_urdf.opt_specify_random_generate_write_urdf(random_total_arm_length) # 啟用隨機的L2,L3長度urdf, 並指定總臂長
            # self.std_L2, self.std_L3 = self.robot_urdf.opt_random_generate_write_urdf() # 啟用隨機的L2,L3長度urdf
            self.robot.__init__() # 重製機器人
            self.motor_type_axis_2 = 5.1
            self.motor_type_axis_3 = 5.1
            self.mission_time = np.random.uniform(low = 20, high = 50)
            rospy.loginfo("random_total_arm_length: %s", random_total_arm_length)
            rospy.loginfo("std_L2: %s", self.std_L2)
            rospy.loginfo("std_L3: %s", self.std_L3)
            rospy.loginfo("mission_time: %s", self.mission_time)
            rand_payload = np.random.uniform(low=1, high=4)
            self.payload = rand_payload
            self.payload_position =  [0, 0, 0.04]
            self.robot.payload(self.payload, self.payload_position)
            # rospy.loginfo("payload: %s", self.payload)
            self.point_Workspace_cal_Monte_Carlo() # 在當前reset出來的機械手臂構型下, 生成點位
            self.random_select_point() # 先隨機抽樣30個點位
            ratio_over, torque_over, consumption, reach_score, manipulability_score = self.performance_evaluate(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
            self.prev_shaping = None
            self.state[0] = ratio_over
            self.state[1] = torque_over
            self.state[2] = np.abs(consumption)
            self.state[3] = reach_score
            self.state[4] = manipulability_score
            self.state[5] = self.std_L2
            self.state[6] = self.std_L3
            self.counts = 0
            return self.state
        elif self.model_select == "test":
            # random state (手臂長度隨機)
            rospy.loginfo("model_select:%s", self.model_select)
            total_arm_length = self.op_radius
            self.std_L2, self.std_L3 = self.robot_urdf.opt_specify_random_generate_write_urdf(total_arm_length) # 啟用隨機的L2,L3長度urdf, 並指定總臂長
            # self.std_L2, self.std_L3 = self.robot_urdf.opt_random_generate_write_urdf() # 啟用隨機的L2,L3長度urdf
            self.robot.__init__() # 重製機器人
            # FIXME: 修改未成功匯入payload給予機器人的問題
            self.payload = self.op_payload
            self.payload_position = self.op_payload_position
            self.robot.payload(self.payload, self.payload_position)
            rospy.loginfo("mission_time: %s", self.mission_time)
            self.total_weight = self.op_weight # Kg
            self.total_cost = self.op_cost # 元
            self.reach_distance = self.op_radius # 使用者設定可達半徑最小值
            self.motor_type_axis_2 = 5.1
            self.motor_type_axis_3 = 5.1
            ratio_over, torque_over, consumption, reach_score, manipulability_score = self.performance_evaluate(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
            self.prev_shaping = None
            self.state[0] = ratio_over
            self.state[1] = torque_over
            self.state[2] = np.abs(consumption)
            self.state[3] = reach_score
            self.state[4] = manipulability_score
            self.state[5] = self.std_L2
            self.state[6] = self.std_L3
            self.counts = 0
            return self.state        
    
    def original_design(self,std_L2,std_L3, motor_1, motor_2, payload, mission_time):
        #  指定手臂長度
        self.robot_urdf.specified_generate_write_urdf(std_L2, std_L3)
        self.robot.__init__() # 重製機器人
        motor_type_axis_2 = motor_1
        motor_type_axis_3 = motor_2
        self.payload = payload
        self.payload = payload
        op_payload_position = [0, 0, 0.04]
        # self.payload_position = [np.array(op_payload_position)]
        self.robot.payload(self.payload, op_payload_position)  # set payload
        # mission_time = 30
        model_select = "test"
        # rospy.loginfo("mission_time: %s", mission_time)
        self.point_Workspace_cal_Monte_Carlo() # 在當前reset出來的機械手臂構型下, 生成點位
        self.random_select_point() # 先隨機抽樣30個點位
        self.mission_time = mission_time
        ratio_over, torque_over, consumption, reach_score, manipulability_score = self.performance_evaluate(model_select, motor_type_axis_2, motor_type_axis_3)
        # ratio_over, torque_over, consumption = self.power_consumption(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
        # torque = self.dynamics_torque_limit()
        rospy.loginfo("ratio_over: %s", ratio_over)
        rospy.loginfo("torque_over: %s", torque_over)
        rospy.loginfo("consumption: %s", np.abs(consumption))
        # 生成隨機 payload (kg)
        rospy.loginfo("reach_score: %s", reach_score)
        rospy.loginfo("manipulability_score: %s", manipulability_score)
        rospy.loginfo("std_L2: %s", std_L2)
        rospy.loginfo("std_L3: %s", std_L3)
    # 視覺化呈現，它只會回應出呼叫那一刻的畫面給你，要它持續出現，需要寫個迴圈
    def render(self, mode='human'):
        return None
        
    def close(self):
        return None

    def dynamics_torque_limit(self):
        """
        Calculate the maximum torque required by

        each axis when the arm of each axis is the longest and the acceleration is the highest
        """
        torque = np.array([np.zeros(shape=5)]) # TODO: fixed 5dof
        # axis_angle = np.array([np.zeros(shape=6)])
        axis_angle = []
        append_torque_limit_list = []
        temp_torque_max = []
        temp_torque_min = []
        Torque_Max = []
        # 角度轉換
        du = pi / 180
        # 度
        # radian = 180 / pi
        # 弧度
        self.robot.payload(self.payload, self.payload_position)  # set payload
        torque = np.array([np.zeros(shape=5)]) # TODO: fixed 5dof
        q_list = [0, 90, -90, 180, -180]
        # TODO: fixed 5dof
        T_cell = (
            len(q_list)
            * len(q_list)
            * len(q_list)
            * len(q_list)
            * len(q_list)
        )
        # TODO: fixed 5dof
        for i in range(len(q_list)):
            q1 = q_list[i]
            percent = i / T_cell * 100
            for j in range(len(q_list)):
                q2 = q_list[j]
                for k in range(len(q_list)):
                    q3 = q_list[k]
                    for l in range(len(q_list)):
                        q4 = q_list[l]
                        for m in range(len(q_list)):
                            q5 = q_list[m]
                            axis_angle.append([q1, q2, q3, q4, q5])
                            load = np.array(
                                [
                                    self.robot.rne(
                                        [
                                            q1 * du,
                                            q2 * du,
                                            q3 * du,
                                            q4 * du,
                                            q5 * du,
                                        ],
                                        self.vel,
                                        self.acc
                                    )
                                ]
                            )
                            torque = np.append(torque, load, axis=0)
        # TODO: fixed 5dof
        for i in range(5):
            axis = i
            toque_max_index = np.argmax(torque[:, axis])
            toque_min_index = np.argmin(torque[:, axis])

            temp_torque_max = torque[toque_max_index].tolist()
            temp_torque_max.extend(axis_angle[toque_max_index])
            temp_torque_min = torque[toque_min_index].tolist()
            temp_torque_min.extend(axis_angle[toque_min_index])
            append_torque_limit_list.append(temp_torque_max)
            append_torque_limit_list.append(temp_torque_min)
            Torque_Max.append(abs(torque[toque_max_index][i]))
            self.torque_dynamics_limit = Torque_Max
        return self.torque_dynamics_limit

  
    def point_Workspace_cal_Monte_Carlo(self):
        """
        Through the "Work space" page in the interface to calculate of the robot
        """
        i = 0
        # 角度轉換
        du = pi / 180
        # 度
        radian = 180 / pi
        # 弧度

        self.q1_s = -180
        self.q1_end = 180
        self.q2_s = -50
        self.q2_end = 230
        self.q3_s = -150
        self.q3_end = 150
        self.q4_s = -180
        self.q4_end = 180
        self.q5_s = -180
        self.q5_end = 180
        
        N = 10 # 改為直接random 10個點
        theta1 = self.q1_s + (self.q1_end - self.q1_s) * np.random.rand(N, 1)
        theta2 = self.q2_s + (self.q2_end - self.q2_s) * np.random.rand(N, 1)
        theta3 = self.q3_s + (self.q3_end - self.q3_s) * np.random.rand(N, 1)
        theta4 = self.q4_s + (self.q4_end - self.q4_s) * np.random.rand(N, 1)
        theta5 = self.q5_s + (self.q5_end - self.q5_s) * np.random.rand(N, 1)

        for i in range(N):
            q1 = theta1[i, 0]
            q2 = theta2[i, 0]
            q3 = theta3[i, 0]
            q4 = theta4[i, 0]
            q5 = theta5[i, 0]
            # TODO: fixed 5dof
            # q6 = theta6[i, 0]
            self.T = self.robot.fkine(
                [q1 * du, q2 * du, q3 * du, q4 * du, q5 * du, 0]# TODO: fixed 5dof
            )
            t = np.round(self.T.t, 3)
            r = np.round(self.T.rpy('deg'), 3)
            self.T_x.append(t[0])
            self.T_y.append(t[1])
            self.T_z.append(t[2])
            self.T_roll.append(int(r[0]))
            self.T_pitch.append(int(r[1]))
            self.T_yaw.append(int(r[2]))
            i = i + 1
        
    def random_select_point(self):
        excel_file = Workbook()
        sheet = excel_file.active

        for i in range(10):
            # x = np.random.randint(1,999)
            sheet.cell(row=i + 1, column=1).value = self.T_x[i]
            sheet.cell(row=i + 1, column=2).value = self.T_y[i]
            sheet.cell(row=i + 1, column=3).value = self.T_z[i]
            sheet.cell(row=i + 1, column=4).value = self.T_roll[i]
            sheet.cell(row=i + 1, column=5).value = self.T_pitch[i]
            sheet.cell(row=i + 1, column=6).value = self.T_yaw[i]

        file_name = self.xlsx_outpath + "/task_point_5dof" +".xlsx" # TODO: fixed 5dof
        excel_file.save(file_name)

    def performance_evaluate(self,model_select,axis2_motor_type, axis3_motor_type):
        if model_select == "train":
            # import xlsx
            df = load_workbook("./xlsx/task_point_5dof.xlsx")
        elif model_select == "test":
            # import xlsx
            df = load_workbook("./xlsx/task_point_5dof_tested.xlsx")
        sheets = df.worksheets
        sheet1 = sheets[0]
        rows = sheet1.rows
        cols = sheet1.columns
        T_tmp = []
        score = []
        ratio_over = 0
        torque_over = 0
        num_torque = np.array([np.zeros(shape=6)])
        total_time = self.mission_time
        # 采样间隔
        sample_interval = 0.2
        manipulability_index = []
        i = 0
        false_done = False
        count = 0
        for row in rows:
            row_val = [col.value for col in row]
            T_tmp.append(SE3(row_val[0], row_val[1], row_val[2]) * SE3.RPY([np.deg2rad(row_val[3]), np.deg2rad(row_val[4]), np.deg2rad(row_val[5])]))
            ik_q = self.robot.ikine_LMS(T=T_tmp[i])
            if ik_q.success == True:
                count += 1
                manipulability_index.append(self.robot.manipulability(q=ik_q.q))
            if i >= 1:
                # 生成时间向量
                traj_time = total_time/9 # 除以點位數量
                time_vector = np.linspace(0, traj_time, int(traj_time/sample_interval) + 1)
                traj = self.robot.jtraj(T_tmp[i-1],T_tmp[i],time_vector)
            
                if np.amax(traj.sd) > 3.04:
                    ratio_over = ratio_over + 1
                torque = self.robot.rne(traj.s,traj.sd,traj.sdd)

                row = abs(torque[:,1])  # 取出第2行
                result_2 = row[row > axis2_motor_type] # 取出大于阈值的数字
                row = abs(torque[:,2]) # 取出第3行
                result_3 = row[row > axis3_motor_type] # 取出大于阈值的数字
                if len(result_2)>0 or len(result_3) >0:
                    torque_over = torque_over + 1
                num_torque = np.append(num_torque, torque)
            i = i + 1

        num_samples = int(total_time / sample_interval)
        total_energy = 0
        for j in range(num_samples):
            energy = abs(num_torque[j]) * sample_interval
            total_energy += energy
            
        if count == 0:
            return(ratio_over, torque_over, total_energy,0,0)
        else:
            final_score = count / i
            if count == 1:
                return(ratio_over, torque_over, total_energy, final_score, manipulability_index[0]) # 回傳 manipulability[0]
            else:
                return(ratio_over, torque_over, total_energy, final_score, np.mean(manipulability_index)) # 回傳 manipulability 取平均
        
if __name__ == '__main__':
    env = RobotOptEnv()
    
    action = env.action_space.sample()
    print(action)
    print(action[0])
    print(type(action[0]))
    print(action[1])
    print(action[2])
    
