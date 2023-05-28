#!/usr/bin/env python3
# coding: utf-8
import gym
from gym import spaces
import numpy as np
from math import pi

from sympy import false
# from torch import R
from interface_control.msg import optimal_design
from dynamics.arm_workspace import arm_workspace_plane
# from robot_urdf import RandomRobot
from dynamics.motor_module import motor_data
from dynamics.modular_robot_3dof import modular_robot_3dof
from dynamics.modular_robot_4dof import modular_robot_4dof
from dynamics.modular_robot_5dof import modular_robot_5dof
from dynamics.modular_robot_6dof import modular_robot_6dof
# from dynamics.stl_conv_6dof_urdf import stl_conv_urdf
from dynamics.stl_conv_6dof_urdf_dynamixel import stl_conv_urdf
import rospy
# one-hot encoder
from sklearn import preprocessing
import pandas as pd
import yaml
# add 可達性可操作性評估
import numpy as np
from roboticstoolbox import DHRobot, RevoluteDH
from spatialmath import SE3
import spatialmath as sm
from urdf_parser_py.urdf import URDF
import os
from openpyxl import load_workbook
from openpyxl import Workbook
import time
import os
from termcolor import cprint
import numpy as np
from pybullet_test import motion_model
HERE = os.path.dirname(__file__)
SINGLE_ARM = os.path.join(HERE,'urdf', 'single_arm_v12.urdf')


# TODO: 初版 只考慮 6 dof 機器人的關節長度變化, 觀察各軸馬達極限之輸出最大torque值
class RobotOptEnv(gym.Env):
    metadata = {
        'render.modes': ['human', 'rgb_array'],
        'video.frames_per_second': 2
    }
    def __init__(self):
        self.robot = modular_robot_6dof()
        self.robot_urdf = stl_conv_urdf("single_arm_v12","test")
        self.motion_plan = motion_model()
        # self.robot_urdf.init_dynamixel_diff_inertia()
        # callback:Enter the parameters of the algorithm to be optimized on the interface
        self.sub_optimal_design = rospy.Subscriber(
            "/optimal_design", optimal_design, self.optimal_design_callback
        )
        # 使用者設定參數
        self.payload = 5.0
        self.payload_position = np.array([0, 0, 0.04])
        self.vel = np.array([3.04, 3.04, 3.04, 3.04, 3.04, 3.04])
        self.acc = np.array([2.356194, 2.356194, 2.356194, 2.356194, 2.356194, 2.356194])
        self.total_weight = 20 # Kg
        self.total_cost = 1800 # 元
        # 預設二,三軸軸長
        self.std_L2 = 35.0 # 預設標準值 第二軸 35 cm
        self.std_L3 = 35.0 # 預設標準值 第三軸 35 cm
        # 觀察參數 motor
        self.motor = motor_data()
        self.res = self.motor.dynamixel_member
        self.high_torque = float('inf') # 預設標準值 馬達極限 120.0 N max
        self.low_torque = float('-inf') # 預設標準值 馬達極限 -12.0 N max
        self.motor_cost_init = np.array([0,400,400,0,0,0], dtype=np.float64) # 預設最大馬達費用
        self.motor_weight_init = np.array([0,0.855,0.855,0,0,0], dtype=np.float64) # 預設最大馬達重量
        self.motor_cost = np.array([0,400,400,0,0,0], dtype=np.float64) # 馬達費用
        self.motor_weight = np.array([0,0.855,0.855,0,0,0], dtype=np.float64) # 馬達重量
        self.motor_rated = np.array([44.7,44.7,44.7,44.7,44.7,44.7], dtype=np.float64)
        # 使用者設定參數 & 觀察參數
        self.reach_distance = 0.6 # 使用者設定可達半徑最小值
        self.high_reach_eva = 1 # 預設觀測標準值
        self.low_reach_eva = 0 # 預設觀測標準值
        self.high_motion_eva = 1 # 預設觀測標準值
        self.low_motion_eva = 0 # 預設觀測標準值
        # self.high_manipulability = 3  # 預設觀測標準值
        # self.low_manipulability = 0  # 預設觀測標準值
        self.high_std_L2 = 70 # 預設觀測標準值
        self.low_std_L2 = -25 # 預設觀測標準值
        self.high_std_L3 = 70 # 預設觀測標準值
        self.low_std_L3 = -25 # 預設觀測標準值
        # self.high_ratio_over = 10 # 預設觀測標準值
        # self.low_ratio_over = 0 # 預設觀測標準值
        self.high_torque_over = 10 # 預設觀測標準值
        self.low_torque_over = 0 # 預設觀測標準值
        # self.high_power_consumption = float('inf')
        # self.low_power_consumption = float('-inf')
        self.torque_done = np.array([false, false, false, false, false, false])
        self.torque_over = False
        self.ratio_over = False
        self.prev_shaping = None
        self.motor_type_axis_2 = 5.1
        self.motor_type_axis_3 = 5.1
        self.mission_time = 0
        self.low_torque_cost = -200
        self.high_torque_cost = 200
        # FIXME:
        self.action_select = 'fixed' 
        self.point_test_excel = './xlsx/task_point_6dof_tested_ori_random.xlsx'
        self.MAX_LENGTH = 40
        self.MIN_LENGTH = 5
        # TODO: 增加馬達模組選型action
        self.action_space = spaces.Discrete(10) # TODO: fixed 12種action

        # TODO: observation space for torque over 6DoF, reach, motion, axis 2, axis 3
        self.observation_space = spaces.Box(np.array([self.low_torque_over, self.low_reach_eva, self.low_motion_eva, self.low_std_L2, self.low_std_L3 ]), 
                                            np.array([self.high_torque_over, self.high_reach_eva, self.high_motion_eva,self.high_std_L2, self.high_std_L3]), 
                                            dtype=np.float64)
        # TODO: reward 歸一化
        self.state = np.array([0,0,0,0,0], dtype=np.float64)
        self.pre_state = np.array([0,0,0,0,0], dtype=np.float64)
        #隨機抽樣點位初始化
        self.T_x = []
        self.T_y = []
        self.T_z = []
        self.T_roll = []
        self.T_pitch = []
        self.T_yaw = []
        #隨機抽樣點位角度初始化
        self.joint_1 = []
        self.joint_2 = []
        self.joint_3 = []
        self.joint_4 = []
        self.joint_5 = []
        self.joint_6 = []
        self.xlsx_outpath = "./xlsx/"

        self.model_select = "train" # 選擇使用train or test model 

        self.op_dof = None
        self.op_payload = None
        self.op_payload_position = None
        self.op_vel = None
        self.op_acc = None
        self.op_radius = None
        self.op_weight = None
        self.op_cost = None

    def optimal_design_callback(self, data):
        # print(data.data)
        # TODO: 目標構型
        self.op_dof = data.dof
        self.op_payload = data.payload
        self.op_payload_position = data.payload_position
        self.op_vel = data.vel
        self.op_acc = data.acc
        self.op_radius = data.radius
        self.op_weight = data.arm_weight
        self.op_cost = data.cost
        
        print("op_dof:", self.op_dof)
        print("op_payload:", self.op_payload)
        print("op_payload_position:", self.op_payload_position)
        print("op_vel:", self.op_vel)
        print("op_acc:", self.op_acc)
        print("op_radius:", self.op_radius)
        print("op_weight:", self.op_weight)
        print("op_cost:", self.op_cost)
        
        self.payload = self.op_payload
        self.payload_position = np.array(self.op_payload_position)
        self.vel = np.array(self.op_vel[0:6])
        self.acc = np.array(self.op_acc[0:6])
        self.total_weight = self.op_weight # Kg
        self.total_cost = self.op_cost # 元
        self.reach_distance = self.op_radius # 使用者設定可達半徑最小值
    # def one_hot_encode(action, num_actions):
    #     one_hot = np.zeros(num_actions)
    #     one_hot[action] = 1
    #     return one_hot
    # TODO: fixed
    def step(self, action):
        assert self.action_space.contains(action), "%r (%s) invalid"%(action, type(action))
        if self.action_select == 'variable':
            if action == 0: 
                self.std_L2 -= 1.0
            elif action == 1:
                self.std_L2 += 1.0
            elif action == 2:
                self.std_L3 -= 1.0
            elif action == 3:
                self.std_L3 += 1.0
            elif action == 4:
                self.motor_type_axis_2 = 5.1
            elif action == 5:
                self.motor_type_axis_2 = 25.3
            elif action == 6:
                self.motor_type_axis_2 = 44.7
            elif action == 7:
                self.motor_type_axis_3 = 5.1 
            elif action == 8:
                self.motor_type_axis_3 = 25.3
            elif action == 9:
                self.motor_type_axis_3 = 44.7
          
        elif self.action_select == 'fixed':
            if action == 0: # 軸2  # 短 # 型號1
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                # 配置軸2 motor 型號1
                #[5.1, 25.3, 44.7]
                self.motor_type_axis_2 = 5.1 # 型號
                
            elif action == 1: # 軸2  # 長 # 型號1
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_2 = 5.1 # 型號
                
            elif action == 2: # 軸2  # 短 # 型號2
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                self.motor_type_axis_2 = 25.3 # 型號

            elif action == 3: # 軸2  # 長 # 型號2
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_2 = 25.3 # 型號
                
            elif action == 4: # 軸2  # 短 # 型號3
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                self.motor_type_axis_2 = 44.7 # 型號

            elif action == 5: # 軸2  # 長 # 型號3
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_2 = 44.7 # 型號
                
            elif action == 6: # 軸3  # 短 # 型號1
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                self.motor_type_axis_3 = 5.1 # 型號
                
            elif action == 7: # 軸3  # 長 # 型號1
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_3 = 5.1 # 型號
                
            elif action == 8: # 軸3  # 短 # 型號2
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                self.motor_type_axis_3 = 25.3 # 型號
                
            elif action == 9: # 軸3  # 長 # 型號2
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_3 = 25.3 # 型號

            elif action == 10: # 軸3  # 短 # 型號3
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                self.motor_type_axis_3 = 44.7 # 型號

            elif action == 11: # 軸3  # 長 # 型號3
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_3 = 44.7 # 型號

        # 輸入action後 二,三軸軸長
        self.robot_urdf.specified_generate_write_urdf(self.std_L2, self.std_L3)
        self.robot.__init__() # 重製機器人
        # FIXME: 修改未成功匯入payload給予機器人的問題
        self.robot.payload(self.payload, self.payload_position)  # set payload
        # TODO: 撰寫motion planning 
        torque = self.dynamics_torque_limit()
        rospy.loginfo("torque: %s", torque)
        rospy.loginfo("motor_type_axis_2: %s", self.motor_type_axis_2)
        rospy.loginfo("motor_type_axis_3: %s", self.motor_type_axis_3)
        torque_over = self.torque_score_result(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3, torque)
        self.state[0] = torque_over
        reach_score = self.reachability_performance_evaluate(self.model_select)
        if reach_score == 1:
            reach_score, motion_score = self.motion_planning_performance_evaluate(self.std_L2, self.std_L3, self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
        else:
            motion_score = 0
        # reach_score, motion_score = self.motion_planning_performance_evaluate(self.std_L2, self.std_L3, self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
        self.state[1] = reach_score
        self.state[2] = motion_score
        self.prev_shaping = None
        self.state[3] = self.std_L2
        self.state[4] = self.std_L3
        self.counts += 1
        reward = 0

        terminated = False
        # 避免軸長小於0
        if self.state[3] <= self.MIN_LENGTH or self.state[4] <= self.MIN_LENGTH  or self.state[3] >= self.MAX_LENGTH or self.state[4] >= self.MAX_LENGTH:
            terminated = True
            reward += -200
            
        percent = 100 - self.state[1] * 100
        reward += -percent
        
        # motion_percent = 100 - self.state[2] * 100
        # reward += -motion_percent

        torque_score = self.state[0]
        reward -= torque_score * 10 # TODO: fixed reward
        
        if percent == 0:
            motion_percent = self.state[2] * 100
            reward += motion_percent
            torque_score = self.state[0]
            if torque_score == 0:
                reward += 100
            if motion_percent == 100:
                torque_score = self.state[0]
                if torque_score == 0:
                    reward += 100
                    terminated = True
                    self.counts = 0

        if self.counts == 50: # max_steps
            terminated = True
            self.counts = 0

        current_design = [self.std_L2, self.std_L3, self.torque_over, motion_score]
        self.torque_over = False #reset
        rospy.loginfo("counts: %s", self.counts)
        rospy.loginfo("step_reward: %s", reward)
        rospy.loginfo("step_state: %s", self.state)
        rospy.loginfo("================================")
        # print("================================")
        
        return self.state, reward, terminated, current_design
    # reset环境状态 
    def reset(self):
        if self.model_select == "train":
            rospy.loginfo("model_select:%s", self.model_select)
            if self.action_select == 'variable':
                self.std_L2, self.std_L3 = self.robot_urdf.opt_random_generate_write_urdf() # 啟用隨機的L2,L3長度urdf
            elif self.action_select == 'fixed':
                random_total_arm_length = np.random.uniform(low=10, high=60) # FIX: 臂長長度
                self.std_L2, self.std_L3 = self.robot_urdf.opt_specify_random_generate_write_urdf(random_total_arm_length) # 啟用隨機的L2,L3長度urdf, 並指定總臂長
            self.robot.__init__() # 重製機器人
            self.motor_type_axis_2 = 44.7
            self.motor_type_axis_3 = 44.7
            rand_payload = np.random.uniform(low=1, high=4)
            self.payload = rand_payload
            self.robot.payload(self.payload, self.payload_position)  # set payload
            torque = self.dynamics_torque_limit()
            torque_over = self.torque_score_result(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3, torque)
            self.state[0] = torque_over
            self.point_Workspace_cal_Monte_Carlo() # 在當前reset出來的機械手臂構型下, 生成點位
            self.random_select_point() # 先隨機抽樣30個點位
            self.prev_shaping = None
            reach_score, motion_score = self.motion_planning_performance_evaluate(self.std_L2, self.std_L3, self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
            self.state[1] = reach_score
            self.state[2] = motion_score
            self.state[3] = self.std_L2
            self.state[4] = self.std_L3
            self.counts = 0
            return self.state
        elif self.model_select == "test":
            # random state (手臂長度隨機)
            rospy.loginfo("model_select:%s", self.model_select)
            total_arm_length = self.op_radius
            self.std_L2, self.std_L3 = self.robot_urdf.opt_specify_random_generate_write_urdf(total_arm_length) # 啟用隨機的L2,L3長度urdf, 並指定總臂長
            self.robot.__init__() # 重製機器人
            self.payload = self.op_payload
            self.payload_position = np.array(self.op_payload_position)
            self.robot.payload(self.payload, self.payload_position)  # set payload
            torque = self.dynamics_torque_limit()
            torque_over = self.torque_score_result(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3, torque)
            self.state[0] = torque_over
            self.point_Workspace_cal_Monte_Carlo() # 在當前reset出來的機械手臂構型下, 生成點位
            self.random_select_point() # 先隨機抽樣30個點位
            self.prev_shaping = None
            reach_score, motion_score = self.motion_planning_performance_evaluate(self.std_L2, self.std_L3, self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
            self.state[1] = reach_score
            self.state[2] = motion_score
            self.state[3] = self.std_L2
            self.state[4] = self.std_L3
            self.counts = 0
            return self.state        

    # 視覺化呈現，它只會回應出呼叫那一刻的畫面給你，要它持續出現，需要寫個迴圈
    def render(self, mode='human'):
        return None
        
    def close(self):
        return None
    
    def torque_score_result(self,model_select,axis2_motor_type, axis3_motor_type, torque):
        torque_over = 0
        if axis2_motor_type <= torque[1]:
            torque_over = torque_over +1
        if axis3_motor_type <= torque[2]:
            torque_over = torque_over +1
        return torque_over
    
    def dynamics_torque_limit(self):
        """
        Calculate the maximum torque required by

        each axis when the arm of each axis is the longest and the acceleration is the highest
        """
        torque = np.array([np.zeros(shape=6)])
        # axis_angle = np.array([np.zeros(shape=6)])
        axis_angle = []
        append_torque_limit_list = []
        temp_torque_max = []
        temp_torque_min = []
        Torque_Max = []
        # 角度轉換
        du = pi / 180
        # 度
        # radian = 180 / pi
        # 弧度
        self.robot.payload(self.payload, self.payload_position)  # set payload
        torque = np.array([np.zeros(shape=6)])
        q_list = [0, 90, -90, 180, -180]
        T_cell = (
            len(q_list)
            * len(q_list)
            * len(q_list)
            * len(q_list)
            * len(q_list)
            * len(q_list)
        )

        for i in range(len(q_list)):
            q1 = q_list[i]
            percent = i / T_cell * 100
            for j in range(len(q_list)):
                q2 = q_list[j]
                for k in range(len(q_list)):
                    q3 = q_list[k]
                    for l in range(len(q_list)):
                        q4 = q_list[l]
                        for m in range(len(q_list)):
                            q5 = q_list[m]
                            for n in range(len(q_list)):
                                q6 = q_list[n]
                                axis_angle.append([q1, q2, q3, q4, q5, q6])
                                load = np.array(
                                    [
                                        self.robot.rne(
                                            [
                                                q1 * du,
                                                q2 * du,
                                                q3 * du,
                                                q4 * du,
                                                q5 * du,
                                                q6 * du,
                                            ],
                                            self.vel,
                                            self.acc
                                        )
                                    ]
                                )
                                torque = np.append(torque, load, axis=0)

        for i in range(6):
            axis = i
            toque_max_index = np.argmax(torque[:, axis])
            toque_min_index = np.argmin(torque[:, axis])

            temp_torque_max = torque[toque_max_index].tolist()
            temp_torque_max.extend(axis_angle[toque_max_index])
            temp_torque_min = torque[toque_min_index].tolist()
            temp_torque_min.extend(axis_angle[toque_min_index])
            append_torque_limit_list.append(temp_torque_max)
            append_torque_limit_list.append(temp_torque_min)
            Torque_Max.append(abs(torque[toque_max_index][i]))
            self.torque_dynamics_limit = Torque_Max
        return self.torque_dynamics_limit
    def point_Workspace_cal_Monte_Carlo(self):
        """
        Through the "Work space" page in the interface to calculate of the robot
        """
        i = 0
        # 角度轉換
        du = pi / 180
        # 度
        radian = 180 / pi
        # 弧度
        # TODO: 真實情況下機構極限
        self.q1_s = -180
        self.q1_end = 180
        self.q2_s = -50
        self.q2_end = 230
        self.q3_s = -150
        self.q3_end = 150
        self.q4_s = -180
        self.q4_end = 180
        self.q5_s = -180
        self.q5_end = 180
        self.q6_s = -180
        self.q6_end = 180
        N = 10
        theta1 = np.around(self.q1_s + (self.q1_end - self.q1_s) * np.random.rand(N, 1), 0)
        theta2 = np.around(self.q2_s + (self.q2_end - self.q2_s) * np.random.rand(N, 1), 0)
        theta3 = np.around(self.q3_s + (self.q3_end - self.q3_s) * np.random.rand(N, 1), 0)
        theta4 = np.around(self.q4_s + (self.q4_end - self.q4_s) * np.random.rand(N, 1), 0)
        theta5 = np.around(self.q5_s + (self.q5_end - self.q5_s) * np.random.rand(N, 1), 0)
        theta6 = np.around(self.q6_s + (self.q6_end - self.q6_s) * np.random.rand(N, 1), 0)

        for i in range(N):
            q1 = theta1[i, 0]
            # TODO: 真實情況下機構極限
            q2 = theta2[i, 0]-90
            q3 = theta3[i, 0]
            q4 = theta4[i, 0]
            q5 = theta5[i, 0]
            q6 = theta6[i, 0]
            self.T = self.robot.fkine(
                [q1 * du, q2 * du, q3 * du, q4 * du, q5 * du, q6 * du]
            )
            t = np.round(self.T.t, 3)
            r = np.round(self.T.rpy('deg'), 3)
            self.T_x.append(t[0])
            self.T_y.append(t[1])
            self.T_z.append(t[2])
            self.T_roll.append(int(r[0]))
            self.T_pitch.append(int(r[1]))
            self.T_yaw.append(int(r[2]))
            i = i + 1
        
    def random_select_point(self):
        excel_file = Workbook()
        sheet = excel_file.active

        for i in range(10):
            sheet.cell(row=i + 1, column=1).value = self.T_x[i]
            sheet.cell(row=i + 1, column=2).value = self.T_y[i]
            sheet.cell(row=i + 1, column=3).value = self.T_z[i]
            sheet.cell(row=i + 1, column=4).value = self.T_roll[i]
            sheet.cell(row=i + 1, column=5).value = self.T_pitch[i]
            sheet.cell(row=i + 1, column=6).value = self.T_yaw[i]

        file_name = self.xlsx_outpath + "/task_point_motion" +".xlsx"
        excel_file.save(file_name)
    def reachability_performance_evaluate(self,model_select):
        if model_select == "train":
            # import xlsx
            df = load_workbook("./xlsx/task_point.xlsx")
        elif model_select == "test":
            df = load_workbook(self.point_test_excel)
        sheets = df.worksheets
        sheet1 = sheets[0]
        rows = sheet1.rows
        T_tmp = []
        i = 0
        count = 0
        for row in rows:
            row_val = [col.value for col in row]
            T_tmp.append(SE3(row_val[0], row_val[1], row_val[2]) * SE3.RPY([np.deg2rad(row_val[3]), np.deg2rad(row_val[4]), np.deg2rad(row_val[5])]))
            ik_q = self.robot.ikine_LMS(T=T_tmp[i])

            if ik_q.success == True:
                count += 1
            i = i + 1
        final_score = count / i
        if count == 0:
            return(0)
        else:
            return(final_score)
        
    def motion_planning_performance_evaluate(self, std_L2, std_L3, model_select,axis2_motor_type, axis3_motor_type):
        if model_select == "train":
            # import xlsx
            df = load_workbook("./xlsx/task_point_motion.xlsx")
        elif model_select == "test":
            # import xlsx
            df = load_workbook(self.point_test_excel)
            
        sheets = df.worksheets
        sheet1 = sheets[0]
        rows = sheet1.rows
        cols = sheet1.columns
        T_tmp = []
        T_point = []
        ratio_over = 0
        torque_over = 0
        num_torque = np.array([np.zeros(shape=6)])
        total_time = self.mission_time
        # 采样间隔
        sample_interval = 0.2
        Joint_tmp = []
        i = 0
        false_done = False
        count = 0
        plan_success_count = 0
        self.motion_plan.stl_trimesh_scaling(std_L2, std_L3)
        self.motion_plan.motion_planning_init(False)
        for row in rows:
            point_obstacle_val = [col.value for col in row]
            T_point.append([point_obstacle_val[0], point_obstacle_val[1], point_obstacle_val[2]])
        self.motion_plan.random_obstacle(T_point, 0.05) # distance5cm
        rows = sheet1.rows # init
        for row in rows:
            row_val = [col.value for col in row]
            T_tmp.append(SE3(row_val[0], row_val[1], row_val[2]) * SE3.RPY([np.deg2rad(row_val[3]), np.deg2rad(row_val[4]), np.deg2rad(row_val[5])]))
            ik_q = self.robot.ikine_LMS(T=T_tmp[i])
            
            # TODO: 撰寫motion planning 
            if ik_q.success == True:
                # TODO: 将第三、五、六轴的值加上负号
                ik_q.q[2] = -ik_q.q[2]
                ik_q.q[4] = -ik_q.q[4]
                ik_q.q[5] = -ik_q.q[5]
                Joint_tmp.append(ik_q.q)
                if count >= 1:
                    plan_success, path = self.motion_plan.motion_planning(Joint_tmp[count-1], Joint_tmp[count], collision = False, wait_duration = False)
                    
                    if plan_success == True:
                        plan_success_count = plan_success_count + 1 
                count += 1
            else:
                pass
            i = i + 1
        self.motion_plan.motion_planning_disconnect()
        if count == 0:
            print("ik fail")
            return(0,0)
            
        else:
            final_score = count / i
            if plan_success_count == 0:
                print("motion fail")
                return(final_score, 0)
            else:
                plan_success_score = plan_success_count / i
                return(final_score, plan_success_score) # 回傳 可達性 與 可運動規劃

class RobotOptEnv_3dof(gym.Env):
    metadata = {
        'render.modes': ['human', 'rgb_array'],
        'video.frames_per_second': 2
    }
    def __init__(self):
        self.robot = modular_robot_3dof() # TODO: fixed 3dof
        self.robot_urdf = stl_conv_urdf("single_arm_v12","test")
        self.motion_plan = motion_model()
        # callback:Enter the parameters of the algorithm to be optimized on the interface
        self.sub_optimal_design = rospy.Subscriber(
            "/optimal_design", optimal_design, self.optimal_design_callback
        )
        # 使用者設定參數
        self.payload = 5.0
        self.payload_position = np.array([0, 0, 0.04])
        self.vel = np.array([2.356194, 2.356194, 2.356194])# TODO: fixed 3dof
        self.acc = np.array([2.356194, 2.356194, 2.356194])# TODO: fixed 3dof
        self.total_weight = 20 # Kg
        self.total_cost = 1800 # 元
        # 預設二,三軸軸長
        self.std_L2 = 35.0 # 預設標準值 第二軸 35 cm
        self.std_L3 = 35.0 # 預設標準值 第三軸 35 cm
        # 觀察參數 motor
        self.motor = motor_data()
        self.res = self.motor.dynamixel_member
        self.high_torque = float('inf') # 預設標準值 馬達極限 120.0 N max
        self.low_torque = float('-inf') # 預設標準值 馬達極限 -12.0 N max
        self.motor_cost_init = np.array([0,400,400,0,0,0], dtype=np.float64) # 預設最大馬達費用
        self.motor_weight_init = np.array([0,0.855,0.855,0,0,0], dtype=np.float64) # 預設最大馬達重量
        self.motor_cost = np.array([0,400,400,0,0,0], dtype=np.float64) # 馬達費用 # TODO: fixed 3dof
        self.motor_weight = np.array([0,0.855,0.855], dtype=np.float64) # 馬達重量# TODO: fixed 3dof
        self.motor_rated = np.array([44.7,44.7,44.7], dtype=np.float64)# TODO: fixed 3dof
        # 使用者設定參數 & 觀察參數
        self.reach_distance = 0.6 # 使用者設定可達半徑最小值
        self.high_reach_eva = 1 # 預設觀測標準值
        self.low_reach_eva = 0 # 預設觀測標準值
        self.high_motion_eva = 1 # 預設觀測標準值
        self.low_motion_eva = 0 # 預設觀測標準值
        # self.high_manipulability = 1  # 預設觀測標準值
        # self.low_manipulability = 0  # 預設觀測標準值
        self.high_std_L2 = 70 # 預設觀測標準值
        self.low_std_L2 = -25 # 預設觀測標準值
        self.high_std_L3 = 70 # 預設觀測標準值
        self.low_std_L3 = -25 # 預設觀測標準值
        self.high_torque_over = 10 # 預設觀測標準值
        self.low_torque_over = 0 # 預設觀測標準值
        self.torque_done = np.array([false, false, false])# TODO: fixed 3dof
        self.torque_over = False
        self.ratio_over = False
        self.prev_shaping = None
        self.motor_type_axis_2 = 5.1
        self.motor_type_axis_3 = 5.1
        self.mission_time = 0
        self.low_torque_cost = -200
        self.high_torque_cost = 200
        # FIXME:
        self.action_select = 'fixed' 
        self.point_test_excel = './xlsx/task_point_6dof_tested_ori_random.xlsx'
        self.MAX_LENGTH = 40
        self.MIN_LENGTH = 5
        # TODO: 增加馬達模組選型action
        self.action_space = spaces.Discrete(10) # TODO: fixed 12種action

        # TODO: observation space for torque over 6DoF, reach, motion, axis 2, axis 3
        self.observation_space = spaces.Box(np.array([self.low_torque_over, self.low_reach_eva, self.low_motion_eva, self.low_std_L2, self.low_std_L3 ]), 
                                            np.array([self.high_torque_over, self.high_reach_eva, self.high_motion_eva,self.high_std_L2, self.high_std_L3]), 
                                            dtype=np.float64)
        # TODO: reward 歸一化
        self.state = np.array([0,0,0,0,0], dtype=np.float64)
        self.pre_state = np.array([0,0,0,0,0], dtype=np.float64)
        #隨機抽樣點位初始化
        self.T_x = []
        self.T_y = []
        self.T_z = []
        self.T_roll = []
        self.T_pitch = []
        self.T_yaw = []
        #隨機抽樣點位角度初始化
        self.joint_1 = []
        self.joint_2 = []
        self.joint_3 = []
        self.joint_4 = []
        self.joint_5 = []
        self.joint_6 = []
        self.xlsx_outpath = "./xlsx/"

        self.model_select = "train" # 選擇使用train or test model 

        self.op_dof = None
        self.op_payload = None
        self.op_payload_position = None
        self.op_vel = None
        self.op_acc = None
        self.op_radius = None
        self.op_weight = None
        self.op_cost = None

    def optimal_design_callback(self, data):
        # print(data.data)
        # TODO: 目標構型
        self.op_dof = data.dof
        self.op_payload = data.payload
        self.op_payload_position = data.payload_position
        self.op_vel = data.vel
        self.op_acc = data.acc
        self.op_radius = data.radius
        self.op_weight = data.arm_weight
        self.op_cost = data.cost
        
        print("op_dof:", self.op_dof)
        print("op_payload:", self.op_payload)
        print("op_payload_position:", self.op_payload_position)
        print("op_vel:", self.op_vel)
        print("op_acc:", self.op_acc)
        print("op_radius:", self.op_radius)
        print("op_weight:", self.op_weight)
        print("op_cost:", self.op_cost)
        
        self.payload = self.op_payload
        self.payload_position = np.array(self.op_payload_position)
        self.vel = np.array(self.op_vel[0:3])# TODO: fixed 3dof
        self.acc = np.array(self.op_acc[0:3])# TODO: fixed 3dof
        self.total_weight = self.op_weight # Kg
        self.total_cost = self.op_cost # 元
        self.reach_distance = self.op_radius # 使用者設定可達半徑最小值
        
    # TODO: fixed
    def step(self, action):
        assert self.action_space.contains(action), "%r (%s) invalid"%(action, type(action))
        if self.action_select == 'variable':
            if action == 0: 
                self.std_L2 -= 1.0
            elif action == 1:
                self.std_L2 += 1.0
            elif action == 2:
                self.std_L3 -= 1.0
            elif action == 3:
                self.std_L3 += 1.0
            elif action == 4:
                self.motor_type_axis_2 = 5.1
            elif action == 5:
                self.motor_type_axis_2 = 25.3
            elif action == 6:
                self.motor_type_axis_2 = 44.7
            elif action == 7:
                self.motor_type_axis_3 = 5.1 
            elif action == 8:
                self.motor_type_axis_3 = 25.3
            elif action == 9:
                self.motor_type_axis_3 = 44.7
          
        elif self.action_select == 'fixed':
            if action == 0: # 軸2  # 短 # 型號1
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                # 配置軸2 motor 型號1
                #[5.1, 25.3, 44.7]
                self.motor_type_axis_2 = 5.1 # 型號
                
            elif action == 1: # 軸2  # 長 # 型號1
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_2 = 5.1 # 型號
                
            elif action == 2: # 軸2  # 短 # 型號2
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                self.motor_type_axis_2 = 25.3 # 型號

            elif action == 3: # 軸2  # 長 # 型號2
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_2 = 25.3 # 型號
                
            elif action == 4: # 軸2  # 短 # 型號3
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                self.motor_type_axis_2 = 44.7 # 型號

            elif action == 5: # 軸2  # 長 # 型號3
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_2 = 44.7 # 型號
                
            elif action == 6: # 軸3  # 短 # 型號1
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                self.motor_type_axis_3 = 5.1 # 型號
                
            elif action == 7: # 軸3  # 長 # 型號1
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_3 = 5.1 # 型號
                
            elif action == 8: # 軸3  # 短 # 型號2
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                self.motor_type_axis_3 = 25.3 # 型號
                
            elif action == 9: # 軸3  # 長 # 型號2
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_3 = 25.3 # 型號

            elif action == 10: # 軸3  # 短 # 型號3
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                self.motor_type_axis_3 = 44.7 # 型號

            elif action == 11: # 軸3  # 長 # 型號3
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_3 = 44.7 # 型號

        # 輸入action後 二,三軸軸長
        self.robot_urdf.specified_generate_write_urdf(self.std_L2, self.std_L3)
        self.robot.__init__() # 重製機器人
        # FIXME: 修改未成功匯入payload給予機器人的問題
        self.robot.payload(self.payload, self.payload_position)  # set payload
        # TODO: 撰寫motion planning 
        torque = self.dynamics_torque_limit()
        rospy.loginfo("torque: %s", torque)
        rospy.loginfo("motor_type_axis_2: %s", self.motor_type_axis_2)
        rospy.loginfo("motor_type_axis_3: %s", self.motor_type_axis_3)
        torque_over = self.torque_score_result(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3, torque)
        self.state[0] = torque_over
        reach_score, motion_score = self.motion_planning_performance_evaluate(self.std_L2, self.std_L3, self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
        self.state[1] = reach_score
        self.state[2] = motion_score
        self.prev_shaping = None
        self.state[3] = self.std_L2
        self.state[4] = self.std_L3
        self.counts += 1
        reward = 0

        terminated = False
        # 避免軸長小於0
        if self.state[3] <= self.MIN_LENGTH or self.state[4] <= self.MIN_LENGTH  or self.state[3] >= self.MAX_LENGTH or self.state[4] >= self.MAX_LENGTH:
            terminated = True
            reward += -200
            
        percent = 100 - self.state[1] * 100
        reward += -percent
        
        motion_percent = 100 - self.state[2] * 100
        reward += -motion_percent

        torque_score = self.state[0]
        reward -= torque_score * 10 # TODO: fixed reward
        
        if percent == 0:
            torque_score = self.state[0]
            if torque_score == 0:
                reward += 100
        if motion_percent == 0:
            torque_score = self.state[0]
            if torque_score == 0:
                reward += 200
                terminated = True
                self.counts = 0

        if self.counts == 50: # max_steps
            terminated = True
            self.counts = 0

        current_design = [self.std_L2, self.std_L3, self.torque_over, motion_score]
        self.torque_over = False #reset
        rospy.loginfo("counts: %s", self.counts)
        rospy.loginfo("step_reward: %s", reward)
        rospy.loginfo("step_state: %s", self.state)
        rospy.loginfo("================================")
        # print("================================")
        
        return self.state, reward, terminated, current_design

    # reset环境状态 
    def reset(self):
        if self.model_select == "train":
            rospy.loginfo("model_select:%s", self.model_select)
            if self.action_select == 'variable':
                self.std_L2, self.std_L3 = self.robot_urdf.opt_random_generate_write_urdf() # 啟用隨機的L2,L3長度urdf
            elif self.action_select == 'fixed':
                random_total_arm_length = np.random.uniform(low=10, high=60) # FIX: 臂長長度
                self.std_L2, self.std_L3 = self.robot_urdf.opt_specify_random_generate_write_urdf(random_total_arm_length) # 啟用隨機的L2,L3長度urdf, 並指定總臂長
            self.robot.__init__() # 重製機器人
            self.motor_type_axis_2 = 44.7
            self.motor_type_axis_3 = 44.7
            rand_payload = np.random.uniform(low=1, high=4)
            self.payload = rand_payload
            self.robot.payload(self.payload, self.payload_position)  # set payload
            torque = self.dynamics_torque_limit()
            torque_over = self.torque_score_result(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3, torque)
            self.state[0] = torque_over
            self.point_Workspace_cal_Monte_Carlo() # 在當前reset出來的機械手臂構型下, 生成點位
            self.random_select_point() # 先隨機抽樣30個點位
            self.prev_shaping = None
            reach_score, motion_score = self.motion_planning_performance_evaluate(self.std_L2, self.std_L3, self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
            self.state[1] = reach_score
            self.state[2] = motion_score
            self.state[3] = self.std_L2
            self.state[4] = self.std_L3
            self.counts = 0
            return self.state
        elif self.model_select == "test":
            # random state (手臂長度隨機)
            rospy.loginfo("model_select:%s", self.model_select)
            total_arm_length = self.op_radius
            self.std_L2, self.std_L3 = self.robot_urdf.opt_specify_random_generate_write_urdf(total_arm_length) # 啟用隨機的L2,L3長度urdf, 並指定總臂長
            self.robot.__init__() # 重製機器人
            self.payload = self.op_payload
            self.payload_position = np.array(self.op_payload_position)
            self.robot.payload(self.payload, self.payload_position)  # set payload
            torque = self.dynamics_torque_limit()
            torque_over = self.torque_score_result(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3, torque)
            self.state[0] = torque_over
            self.point_Workspace_cal_Monte_Carlo() # 在當前reset出來的機械手臂構型下, 生成點位
            self.random_select_point() # 先隨機抽樣30個點位
            self.prev_shaping = None
            reach_score, motion_score = self.motion_planning_performance_evaluate(self.std_L2, self.std_L3, self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
            self.state[1] = reach_score
            self.state[2] = motion_score
            self.state[3] = self.std_L2
            self.state[4] = self.std_L3
            self.counts = 0
            return self.state        
        
    
    # 視覺化呈現，它只會回應出呼叫那一刻的畫面給你，要它持續出現，需要寫個迴圈
    def render(self, mode='human'):
        return None
        
    def close(self):
        return None
        
    def dynamics_torque_limit(self):
        """
        Calculate the maximum torque required by

        each axis when the arm of each axis is the longest and the acceleration is the highest
        """
        torque = np.array([np.zeros(shape=3)])# TODO: fixed 3dof
        # axis_angle = np.array([np.zeros(shape=6)])
        axis_angle = []
        append_torque_limit_list = []
        temp_torque_max = []
        temp_torque_min = []
        Torque_Max = []
        # 角度轉換
        du = pi / 180
        # 度
        # radian = 180 / pi
        # 弧度
        self.robot.payload(self.payload, self.payload_position)  # set payload
        torque = np.array([np.zeros(shape=3)])# TODO: fixed 3dof
        q_list = [0, 90, -90, 180, -180]
        # TODO: fixed 3dof
        T_cell = (
            len(q_list)
            * len(q_list)
            * len(q_list)
        )
        # TODO: fixed 3dof
        for i in range(len(q_list)):
            q1 = q_list[i]
            percent = i / T_cell * 100
            for j in range(len(q_list)):
                q2 = q_list[j]
                for k in range(len(q_list)):
                    q3 = q_list[k]
                    axis_angle.append([q1, q2, q3])
                    load = np.array(
                        [
                            self.robot.rne(
                                [
                                    q1 * du,
                                    q2 * du,
                                    q3 * du,
                                    # q4 * du,
                                    # q5 * du,
                                    # q6 * du,
                                ],
                                self.vel,
                                self.acc
                            )
                        ]
                    )
                    torque = np.append(torque, load, axis=0)
        # TODO: fixed 3dof
        for i in range(3):
            axis = i
            toque_max_index = np.argmax(torque[:, axis])
            toque_min_index = np.argmin(torque[:, axis])

            temp_torque_max = torque[toque_max_index].tolist()
            temp_torque_max.extend(axis_angle[toque_max_index])
            temp_torque_min = torque[toque_min_index].tolist()
            temp_torque_min.extend(axis_angle[toque_min_index])
            append_torque_limit_list.append(temp_torque_max)
            append_torque_limit_list.append(temp_torque_min)
            Torque_Max.append(abs(torque[toque_max_index][i]))
            self.torque_dynamics_limit = Torque_Max
        return self.torque_dynamics_limit
    def point_Workspace_cal_Monte_Carlo(self):
        """
        Through the "Work space" page in the interface to calculate of the robot
        """
        i = 0
        # 角度轉換
        du = pi / 180
        # 度
        radian = 180 / pi
        # 弧度

        self.q1_s = -180
        self.q1_end = 180
        self.q2_s = -50
        self.q2_end = 230
        self.q3_s = -150
        self.q3_end = 150
        N = 10
        theta1 = np.around(self.q1_s + (self.q1_end - self.q1_s) * np.random.rand(N, 1), 0)
        theta2 = np.around(self.q2_s + (self.q2_end - self.q2_s) * np.random.rand(N, 1), 0)
        theta3 = np.around(self.q3_s + (self.q3_end - self.q3_s) * np.random.rand(N, 1), 0)
       
        for i in range(N):
            q1 = theta1[i, 0]
            q2 = theta2[i, 0]
            q3 = theta3[i, 0]
            self.T = self.robot.fkine(
                [q1 * du, q2 * du, q3 * du]# TODO: fixed 3dof
            )
            t = np.round(self.T.t, 3)
            self.T_x.append(t[0])
            self.T_y.append(t[1])
            self.T_z.append(t[2])
            i = i + 1
    def random_select_point(self):
        excel_file = Workbook()
        sheet = excel_file.active

        for i in range(10):
            sheet.cell(row=i + 1, column=1).value = self.T_x[i]
            sheet.cell(row=i + 1, column=2).value = self.T_y[i]
            sheet.cell(row=i + 1, column=3).value = self.T_z[i]

        file_name = self.xlsx_outpath + "/task_point_3dof" +".xlsx"# TODO: fixed 3dof
        excel_file.save(file_name)
    def motion_planning_performance_evaluate(self, std_L2, std_L3, model_select,axis2_motor_type, axis3_motor_type):
        if model_select == "train":
            # import xlsx
            df = load_workbook("./xlsx/task_point_3dof.xlsx")# TODO: fixed 3dof
        elif model_select == "test":
            # import xlsx
            df = load_workbook(self.point_test_excel)
            
        sheets = df.worksheets
        sheet1 = sheets[0]
        rows = sheet1.rows
        cols = sheet1.columns
        T_tmp = []
        T_point = []
        Joint_tmp = []
        i = 0
        count = 0
        plan_success_count = 0
        self.motion_plan.stl_trimesh_scaling(std_L2, std_L3)
        self.motion_plan.motion_planning_init_3dof(False)
        for row in rows:
            point_obstacle_val = [col.value for col in row]
            T_point.append([point_obstacle_val[0], point_obstacle_val[1], point_obstacle_val[2]])
        self.motion_plan.random_obstacle(T_point, 0.05) # distance5cm
        rows = sheet1.rows # init
        for row in rows:
            row_val = [col.value for col in row]
            # TODO: fixed
            T_tmp.append(SE3(row_val[0], row_val[1], row_val[2]))# TODO: fixed 3dof
            ik_q = self.robot.ikine_LMS(T=T_tmp[i], mask = [1,1,1,0,0,0])# TODO: fixed 3dof
            # TODO: 撰寫motion planning 
            if ik_q.success == True:
                # TODO: 将第三轴的值加上负号
                ik_q.q[2] = -ik_q.q[2]
                # ik_q.q[4] = -ik_q.q[4]
                # ik_q.q[5] = -ik_q.q[5]
                Joint_tmp.append(ik_q.q)
                if count >= 1:
                    plan_success, path = self.motion_plan.motion_planning_3dof(Joint_tmp[count-1], Joint_tmp[count], collision = False, wait_duration = False)
                    
                    if plan_success == True:
                        plan_success_count = plan_success_count + 1 
                count += 1
            else:
                pass
            i = i + 1
        self.motion_plan.motion_planning_disconnect()
        if count == 0:
            print("ik fail")
            return(0,0)
            
        else:
            final_score = count / i
            if plan_success_count == 0:
                print("motion fail")
                return(final_score, 0)
            else:
                plan_success_score = plan_success_count / i
                return(final_score, plan_success_score) # 回傳 可達性 與 可運動規劃
    def torque_score_result(self,model_select,axis2_motor_type, axis3_motor_type, torque):
        torque_over = 0
        if axis2_motor_type <= torque[1]:
            torque_over = torque_over +1
        if axis3_motor_type <= torque[2]:
            torque_over = torque_over +1
        return torque_over
class RobotOptEnv_5dof(gym.Env):
    metadata = {
        'render.modes': ['human', 'rgb_array'],
        'video.frames_per_second': 2
    }
    def __init__(self):
        self.robot = modular_robot_5dof() # TODO: fixed 5dof
        self.robot_urdf = stl_conv_urdf("single_arm_v12","test")
        self.motion_plan = motion_model()
        # self.robot_urdf.init_dynamixel_diff_inertia()
        # callback:Enter the parameters of the algorithm to be optimized on the interface
        self.sub_optimal_design = rospy.Subscriber(
            "/optimal_design", optimal_design, self.optimal_design_callback
        )
        # 使用者設定參數
        self.payload = 5.0
        self.payload_position = np.array([0, 0, 0.04])
        self.vel = np.array([2.356194, 2.356194, 2.356194, 2.356194, 2.356194, 2.356194]) # TODO: fixed 5dof
        self.acc = np.array([2.356194, 2.356194, 2.356194, 2.356194, 2.356194, 2.356194]) # TODO: fixed 5dof
        self.total_weight = 20 # Kg
        self.total_cost = 1800 # 元
        # 預設二,三軸軸長
        self.std_L2 = 35.0 # 預設標準值 第二軸 35 cm
        self.std_L3 = 35.0 # 預設標準值 第三軸 35 cm
        # 觀察參數 motor
        self.motor = motor_data()
        self.res = self.motor.dynamixel_member
        self.high_torque = float('inf') # 預設標準值 馬達極限 120.0 N max
        self.low_torque = float('-inf') # 預設標準值 馬達極限 -12.0 N max
        self.motor_cost_init = np.array([0,400,400,0,0,0], dtype=np.float64) # 預設最大馬達費用
        self.motor_weight_init = np.array([0,0.855,0.855,0,0,0], dtype=np.float64) # 預設最大馬達重量
        self.motor_cost = np.array([0,400,400,0,0], dtype=np.float64) # 馬達費用 # TODO: fixed 5dof
        self.motor_weight = np.array([0,0.855,0.855,0,0], dtype=np.float64) # 馬達重量 # TODO: fixed 5dof
        self.motor_rated = np.array([44.7,44.7,44.7,44.7,44.7], dtype=np.float64) # TODO: fixed 5dof
        # 使用者設定參數 & 觀察參數
        self.reach_distance = 0.6 # 使用者設定可達半徑最小值
        self.high_reach_eva = 1 # 預設觀測標準值
        self.low_reach_eva = 0 # 預設觀測標準值
        self.high_motion_eva = 1 # 預設觀測標準值
        self.low_motion_eva = 0 # 預設觀測標準值
        # self.high_manipulability = 1  # 預設觀測標準值
        # self.low_manipulability = 0  # 預設觀測標準值
        self.high_std_L2 = 70 # 預設觀測標準值
        self.low_std_L2 = -25 # 預設觀測標準值
        self.high_std_L3 = 70 # 預設觀測標準值
        self.low_std_L3 = -25 # 預設觀測標準值
        self.high_torque_over = 10 # 預設觀測標準值
        self.low_torque_over = 0 # 預設觀測標準值
        self.torque_done = np.array([false, false, false, false, false]) # TODO: fixed 5dof
        self.torque_over = False
        self.ratio_over = False
        self.prev_shaping = None
        self.motor_type_axis_2 = 5.1
        self.motor_type_axis_3 = 5.1
        self.mission_time = 0
        self.low_torque_cost = -200
        self.high_torque_cost = 200
        # FIXME:
        self.action_select = 'fixed' 
        self.point_test_excel = './xlsx/task_point_6dof_tested_ori_random.xlsx'
        self.MAX_LENGTH = 40
        self.MIN_LENGTH = 5
        # TODO: 增加馬達模組選型action
        self.action_space = spaces.Discrete(10) # TODO: fixed 12種action

        # TODO: observation space for torque over 6DoF, reach, motion, axis 2, axis 3
        self.observation_space = spaces.Box(np.array([self.low_torque_over, self.low_reach_eva, self.low_motion_eva, self.low_std_L2, self.low_std_L3 ]), 
                                            np.array([self.high_torque_over, self.high_reach_eva, self.high_motion_eva,self.high_std_L2, self.high_std_L3]), 
                                            dtype=np.float64)
        # TODO: reward 歸一化
        self.state = np.array([0,0,0,0,0], dtype=np.float64)
        self.pre_state = np.array([0,0,0,0,0], dtype=np.float64)

        #隨機抽樣點位初始化
        self.T_x = []
        self.T_y = []
        self.T_z = []
        self.T_roll = []
        self.T_pitch = []
        self.T_yaw = []
        self.xlsx_outpath = "./xlsx/"

        self.model_select = "train" # 選擇使用train or test model 

        self.op_dof = None
        self.op_payload = None
        self.op_payload_position = None
        self.op_vel = None
        self.op_acc = None
        self.op_radius = None
        self.op_weight = None
        self.op_cost = None

    def optimal_design_callback(self, data):
        # print(data.data)
        # TODO: 目標構型
        self.op_dof = data.dof
        self.op_payload = data.payload
        self.op_payload_position = data.payload_position
        self.op_vel = data.vel
        self.op_acc = data.acc
        self.op_radius = data.radius
        self.op_weight = data.arm_weight
        self.op_cost = data.cost
        
        print("op_dof:", self.op_dof)
        print("op_payload:", self.op_payload)
        print("op_payload_position:", self.op_payload_position)
        print("op_vel:", self.op_vel)
        print("op_acc:", self.op_acc)
        print("op_radius:", self.op_radius)
        print("op_weight:", self.op_weight)
        print("op_cost:", self.op_cost)
        
        self.payload = self.op_payload
        self.payload_position = np.array(self.op_payload_position)
        self.vel = np.array(self.op_vel[0:6]) # TODO: fixed 5dof
        self.acc = np.array(self.op_acc[0:6]) # TODO: fixed 5dof
        self.total_weight = self.op_weight # Kg
        self.total_cost = self.op_cost # 元
        self.reach_distance = self.op_radius # 使用者設定可達半徑最小值
        
    # TODO: fixed
    def step(self, action):
        assert self.action_space.contains(action), "%r (%s) invalid"%(action, type(action))
        if self.action_select == 'variable':
            if action == 0: 
                self.std_L2 -= 1.0
            elif action == 1:
                self.std_L2 += 1.0
            elif action == 2:
                self.std_L3 -= 1.0
            elif action == 3:
                self.std_L3 += 1.0
            elif action == 4:
                self.motor_type_axis_2 = 5.1
            elif action == 5:
                self.motor_type_axis_2 = 25.3
            elif action == 6:
                self.motor_type_axis_2 = 44.7
            elif action == 7:
                self.motor_type_axis_3 = 5.1 
            elif action == 8:
                self.motor_type_axis_3 = 25.3
            elif action == 9:
                self.motor_type_axis_3 = 44.7
          
        elif self.action_select == 'fixed':
            if action == 0: # 軸2  # 短 # 型號1
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                # 配置軸2 motor 型號1
                #[5.1, 25.3, 44.7]
                self.motor_type_axis_2 = 5.1 # 型號
                
            elif action == 1: # 軸2  # 長 # 型號1
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_2 = 5.1 # 型號
                
            elif action == 2: # 軸2  # 短 # 型號2
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                self.motor_type_axis_2 = 25.3 # 型號

            elif action == 3: # 軸2  # 長 # 型號2
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_2 = 25.3 # 型號
                
            elif action == 4: # 軸2  # 短 # 型號3
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                self.motor_type_axis_2 = 44.7 # 型號

            elif action == 5: # 軸2  # 長 # 型號3
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_2 = 44.7 # 型號
                
            elif action == 6: # 軸3  # 短 # 型號1
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                self.motor_type_axis_3 = 5.1 # 型號
                
            elif action == 7: # 軸3  # 長 # 型號1
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_3 = 5.1 # 型號
                
            elif action == 8: # 軸3  # 短 # 型號2
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                self.motor_type_axis_3 = 25.3 # 型號
                
            elif action == 9: # 軸3  # 長 # 型號2
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_3 = 25.3 # 型號

            elif action == 10: # 軸3  # 短 # 型號3
                self.std_L2 -= 1.0
                self.std_L3 += 1.0
                self.motor_type_axis_3 = 44.7 # 型號

            elif action == 11: # 軸3  # 長 # 型號3
                self.std_L2 += 1.0
                self.std_L3 -= 1.0
                self.motor_type_axis_3 = 44.7 # 型號

        # 輸入action後 二,三軸軸長
        self.robot_urdf.specified_generate_write_urdf(self.std_L2, self.std_L3)
        self.robot.__init__() # 重製機器人
        # FIXME: 修改未成功匯入payload給予機器人的問題
        self.robot.payload(self.payload, self.payload_position)  # set payload
        # TODO: 撰寫motion planning 
        torque = self.dynamics_torque_limit()
        rospy.loginfo("torque: %s", torque)
        rospy.loginfo("motor_type_axis_2: %s", self.motor_type_axis_2)
        rospy.loginfo("motor_type_axis_3: %s", self.motor_type_axis_3)
        torque_over = self.torque_score_result(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3, torque)
        self.state[0] = torque_over
        reach_score, motion_score = self.motion_planning_performance_evaluate(self.std_L2, self.std_L3, self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
        self.state[1] = reach_score
        self.state[2] = motion_score
        self.prev_shaping = None
        self.state[3] = self.std_L2
        self.state[4] = self.std_L3
        self.counts += 1
        reward = 0

        terminated = False
        # 避免軸長小於0
        if self.state[3] <= self.MIN_LENGTH or self.state[4] <= self.MIN_LENGTH  or self.state[3] >= self.MAX_LENGTH or self.state[4] >= self.MAX_LENGTH:
            terminated = True
            reward += -200
            
        percent = 100 - self.state[1] * 100
        reward += -percent
        
        motion_percent = 100 - self.state[2] * 100
        reward += -motion_percent

        torque_score = self.state[0]
        reward -= torque_score * 10 # TODO: fixed reward
        
        if percent == 0:
            torque_score = self.state[0]
            if torque_score == 0:
                reward += 100
        if motion_percent == 0:
            torque_score = self.state[0]
            if torque_score == 0:
                reward += 200
                terminated = True
                self.counts = 0

        if self.counts == 50: # max_steps
            terminated = True
            self.counts = 0

        current_design = [self.std_L2, self.std_L3, self.torque_over, motion_score]
        self.torque_over = False #reset
        rospy.loginfo("counts: %s", self.counts)
        rospy.loginfo("step_reward: %s", reward)
        rospy.loginfo("step_state: %s", self.state)
        rospy.loginfo("================================")
        # print("================================")
        
        return self.state, reward, terminated, current_design

    # reset环境状态 
    def reset(self):
        if self.model_select == "train":
            rospy.loginfo("model_select:%s", self.model_select)
            if self.action_select == 'variable':
                self.std_L2, self.std_L3 = self.robot_urdf.opt_random_generate_write_urdf() # 啟用隨機的L2,L3長度urdf
            elif self.action_select == 'fixed':
                random_total_arm_length = np.random.uniform(low=10, high=60) # FIX: 臂長長度
                self.std_L2, self.std_L3 = self.robot_urdf.opt_specify_random_generate_write_urdf(random_total_arm_length) # 啟用隨機的L2,L3長度urdf, 並指定總臂長
            self.robot.__init__() # 重製機器人
            self.motor_type_axis_2 = 44.7
            self.motor_type_axis_3 = 44.7
            rand_payload = np.random.uniform(low=1, high=4)
            self.payload = rand_payload
            self.robot.payload(self.payload, self.payload_position)  # set payload
            torque = self.dynamics_torque_limit()
            torque_over = self.torque_score_result(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3, torque)
            self.state[0] = torque_over
            self.point_Workspace_cal_Monte_Carlo() # 在當前reset出來的機械手臂構型下, 生成點位
            self.random_select_point() # 先隨機抽樣30個點位
            self.prev_shaping = None
            reach_score, motion_score = self.motion_planning_performance_evaluate(self.std_L2, self.std_L3, self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
            self.state[1] = reach_score
            self.state[2] = motion_score
            self.state[3] = self.std_L2
            self.state[4] = self.std_L3
            self.counts = 0
            return self.state
        elif self.model_select == "test":
            # random state (手臂長度隨機)
            rospy.loginfo("model_select:%s", self.model_select)
            total_arm_length = self.op_radius
            self.std_L2, self.std_L3 = self.robot_urdf.opt_specify_random_generate_write_urdf(total_arm_length) # 啟用隨機的L2,L3長度urdf, 並指定總臂長
            self.robot.__init__() # 重製機器人
            self.payload = self.op_payload
            self.payload_position = np.array(self.op_payload_position)
            self.robot.payload(self.payload, self.payload_position)  # set payload
            torque = self.dynamics_torque_limit()
            torque_over = self.torque_score_result(self.model_select, self.motor_type_axis_2, self.motor_type_axis_3, torque)
            self.state[0] = torque_over
            self.point_Workspace_cal_Monte_Carlo() # 在當前reset出來的機械手臂構型下, 生成點位
            self.random_select_point() # 先隨機抽樣30個點位
            self.prev_shaping = None
            reach_score, motion_score = self.motion_planning_performance_evaluate(self.std_L2, self.std_L3, self.model_select, self.motor_type_axis_2, self.motor_type_axis_3)
            self.state[1] = reach_score
            self.state[2] = motion_score
            self.state[3] = self.std_L2
            self.state[4] = self.std_L3
            self.counts = 0
            return self.state        
    

    # 視覺化呈現，它只會回應出呼叫那一刻的畫面給你，要它持續出現，需要寫個迴圈
    def render(self, mode='human'):
        return None
        
    def close(self):
        return None
        
    def dynamics_torque_limit(self):
        """
        Calculate the maximum torque required by

        each axis when the arm of each axis is the longest and the acceleration is the highest
        """
        torque = np.array([np.zeros(shape=6)]) # TODO: fixed 5dof
        # axis_angle = np.array([np.zeros(shape=6)])
        axis_angle = []
        append_torque_limit_list = []
        temp_torque_max = []
        temp_torque_min = []
        Torque_Max = []
        # 角度轉換
        du = pi / 180
        # 度
        # radian = 180 / pi
        # 弧度
        self.robot.payload(self.payload, self.payload_position)  # set payload
        torque = np.array([np.zeros(shape=6)]) # TODO: fixed 5dof
        q_list = [0, 90, -90, 180, -180]
        # TODO: fixed 5dof
        T_cell = (
            len(q_list)
            * len(q_list)
            * len(q_list)
            * len(q_list)
            * len(q_list)
        )
        # TODO: fixed 5dof
        for i in range(len(q_list)):
            q1 = q_list[i]
            percent = i / T_cell * 100
            for j in range(len(q_list)):
                q2 = q_list[j]
                for k in range(len(q_list)):
                    q3 = q_list[k]
                    for l in range(len(q_list)):
                        q4 = q_list[l]
                        for m in range(len(q_list)):
                            q5 = q_list[m]
                            axis_angle.append([q1, q2, q3, q4, q5])
                            load = np.array(
                                [
                                    self.robot.rne(
                                        [
                                            q1 * du,
                                            q2 * du,
                                            q3 * du,
                                            q4 * du,
                                            q5 * du,
                                            0,
                                        ],
                                        self.vel,
                                        self.acc
                                    )
                                ]
                            )
                            torque = np.append(torque, load, axis=0)
        # TODO: fixed 5dof
        for i in range(5):
            axis = i
            toque_max_index = np.argmax(torque[:, axis])
            toque_min_index = np.argmin(torque[:, axis])

            temp_torque_max = torque[toque_max_index].tolist()
            temp_torque_max.extend(axis_angle[toque_max_index])
            temp_torque_min = torque[toque_min_index].tolist()
            temp_torque_min.extend(axis_angle[toque_min_index])
            append_torque_limit_list.append(temp_torque_max)
            append_torque_limit_list.append(temp_torque_min)
            Torque_Max.append(abs(torque[toque_max_index][i]))
            self.torque_dynamics_limit = Torque_Max
        return self.torque_dynamics_limit
    def point_Workspace_cal_Monte_Carlo(self):
        """
        Through the "Work space" page in the interface to calculate of the robot
        """
        i = 0
        # 角度轉換
        du = pi / 180
        # 度
        radian = 180 / pi
        # 弧度

        self.q1_s = -180
        self.q1_end = 180
        self.q2_s = -50
        self.q2_end = 230
        self.q3_s = -150
        self.q3_end = 150
        self.q4_s = -180
        self.q4_end = 180
        self.q5_s = -180
        self.q5_end = 180
        N = 10
        theta1 = np.around(self.q1_s + (self.q1_end - self.q1_s) * np.random.rand(N, 1), 0)
        theta2 = np.around(self.q2_s + (self.q2_end - self.q2_s) * np.random.rand(N, 1), 0)
        theta3 = np.around(self.q3_s + (self.q3_end - self.q3_s) * np.random.rand(N, 1), 0)
        theta4 = np.around(self.q4_s + (self.q4_end - self.q4_s) * np.random.rand(N, 1), 0)
        theta5 = np.around(self.q5_s + (self.q5_end - self.q5_s) * np.random.rand(N, 1), 0)
        
        for i in range(N):
            q1 = theta1[i, 0]
            q2 = theta2[i, 0]
            q3 = theta3[i, 0]
            q4 = theta4[i, 0]
            q5 = theta5[i, 0]
            # TODO: fixed 5dof
            # q6 = theta6[i, 0]
            self.T = self.robot.fkine(
                [q1 * du, q2 * du, q3 * du, q4 * du, q5 * du, 0]# TODO: fixed 5dof
            )
            t = np.round(self.T.t, 3)
            r = np.round(self.T.rpy('deg'), 3)
            self.T_x.append(t[0])
            self.T_y.append(t[1])
            self.T_z.append(t[2])
            self.T_roll.append(int(r[0]))
            self.T_pitch.append(int(r[1]))
            self.T_yaw.append(int(r[2]))
            i = i + 1
        
    def random_select_point(self):
        excel_file = Workbook()
        sheet = excel_file.active

        for i in range(10):
            # x = np.random.randint(1,999)
            sheet.cell(row=i + 1, column=1).value = self.T_x[i]
            sheet.cell(row=i + 1, column=2).value = self.T_y[i]
            sheet.cell(row=i + 1, column=3).value = self.T_z[i]
            sheet.cell(row=i + 1, column=4).value = self.T_roll[i]
            sheet.cell(row=i + 1, column=5).value = self.T_pitch[i]
            sheet.cell(row=i + 1, column=6).value = self.T_yaw[i]

        file_name = self.xlsx_outpath + "/task_point_5dof" +".xlsx" # TODO: fixed 5dof
        excel_file.save(file_name)
    def motion_planning_performance_evaluate(self, std_L2, std_L3, model_select,axis2_motor_type, axis3_motor_type):
        if model_select == "train":
            # import xlsx
            df = load_workbook("./xlsx/task_point_5dof.xlsx") # TODO: fixed 5dof
        elif model_select == "test":
            # import xlsx
            df = load_workbook("./xlsx/task_point_5dof_tested.xlsx")

            
        sheets = df.worksheets
        sheet1 = sheets[0]
        rows = sheet1.rows
        cols = sheet1.columns
        T_tmp = []
        T_point = []
        Joint_tmp = []
        i = 0
        count = 0
        plan_success_count = 0
        self.motion_plan.stl_trimesh_scaling(std_L2, std_L3)
        self.motion_plan.motion_planning_init(False)
        for row in rows:
            point_obstacle_val = [col.value for col in row]
            T_point.append([point_obstacle_val[0], point_obstacle_val[1], point_obstacle_val[2]])
        self.motion_plan.random_obstacle(T_point, 0.05) # distance5cm
        rows = sheet1.rows # init
        for row in rows:
            row_val = [col.value for col in row]
            # TODO: fixed
            T_tmp.append(SE3(row_val[0], row_val[1], row_val[2]) * SE3.RPY([np.deg2rad(row_val[3]), np.deg2rad(row_val[4]), np.deg2rad(row_val[5])]))
            ik_q = self.robot.ikine_LMS(T=T_tmp[i])
            # TODO: 撰寫motion planning 
            if ik_q.success == True:
                # TODO: 将第三、五轴的值加上负号
                ik_q.q[2] = -ik_q.q[2]
                ik_q.q[4] = -ik_q.q[4]
                # ik_q.q[5] = -ik_q.q[5]
                Joint_tmp.append(ik_q.q)
                if count >= 1:
                    plan_success, path = self.motion_plan.motion_planning(Joint_tmp[count-1], Joint_tmp[count], collision = False, wait_duration = False)
                    
                    if plan_success == True:
                        plan_success_count = plan_success_count + 1 
                count += 1
            else:
                pass
            i = i + 1
        self.motion_plan.motion_planning_disconnect()
        if count == 0:
            print("ik fail")
            return(0,0)
            
        else:
            final_score = count / i
            if plan_success_count == 0:
                print("motion fail")
                return(final_score, 0)
            else:
                plan_success_score = plan_success_count / i
                return(final_score, plan_success_score) # 回傳 可達性 與 可運動規劃
    def torque_score_result(self,model_select,axis2_motor_type, axis3_motor_type, torque):
        torque_over = 0
        if axis2_motor_type <= torque[1]:
            torque_over = torque_over +1
        if axis3_motor_type <= torque[2]:
            torque_over = torque_over +1
        return torque_over
if __name__ == '__main__':
    env = RobotOptEnv()
    
    action = env.action_space.sample()
    print(action)
    print(action[0])
    print(type(action[0]))
    print(action[1])
    print(action[2])
    
