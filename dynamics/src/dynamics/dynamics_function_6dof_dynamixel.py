#!/usr/bin/env python3
# coding: utf-8
import importlib
import sys

import rospy

importlib.reload(sys)
import argparse
import csv
import math
import time
from collections import namedtuple
from math import pi
from os import path

import geometry_msgs.msg
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import plotly.graph_objs as go
import plotly.offline as py
import roboticstoolbox as rtb
import sympy as sp
# import dyna_space
from interface_control.msg import (cal_cmd, cal_process, cal_result, dyna_data,
                                   dyna_space_data, optimal_design,
                                   optimal_random, specified_parameter_design)
from matplotlib import cm
from moveit_msgs.msg import DisplayTrajectory, RobotTrajectory
from mpl_toolkits.mplot3d import Axes3D
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, colors
from openpyxl.utils import get_column_letter
from plotly.offline import download_plotlyjs, iplot, plot
from scipy.interpolate import make_interp_spline  # draw smooth
from spatialmath import *
from trajectory_msgs.msg import JointTrajectory, JointTrajectoryPoint
np.set_printoptions(
    linewidth=100,
    formatter={"float": lambda x: f"{x:8.4g}" if abs(x) > 1e-10 else f"{0:8.4g}"},
)

import pandas as pd

from arm_workspace import arm_workspace_plane
from motor_module import motor_data
from modular_robot_6dof import modular_robot_6dof

class switch(object):
    def __init__(self, value):
        self.value = value
        self.fall = False

    def __iter__(self):
        """Return the match method once, then stop"""
        yield self.match
        raise StopIteration

    def match(self, *args):
        """Indicate whether or not to enter a case suite"""
        if self.fall or not args:
            return True
        elif self.value in args:  # changed for v1.5, see below
            self.fall = True
            return True
        else:
            return False


class Dynamics_dynamixel:
    def __init__(self):
        self.cmd = 0
        self.robot = modular_robot_6dof()
        
        self.robot.gravload(self.robot.qn)
        self.robot.inertia(self.robot.qn)
        self.torque = np.array([np.zeros(shape=6)])

        self.payload = 0
        self.payload_position = np.array([0, 0, 0])
        self.joint_angle = np.array([0, 0, 0, 0, 0, 0, 0])  # degree
        self.vel = np.array([0, 0, 0, 0, 0, 0, 0])  # rad / sec
        self.acc = np.array([0, 0, 0, 0, 0, 0, 0])  # rad / sec2

        self.payload_space = 0
        self.payload_position_space = np.array([0, 0, 0])
        self.axis_num = 1

        self.op_payload = 0
        self.op_payload_position = np.array([0, 0, 0])
        self.op_vel = np.array([0, 0, 0, 0, 0, 0])  # rad / sec
        self.op_acc = np.array([0, 0, 0, 0, 0, 0])  # rad / sec2
        self.op_radius = 0

        self.torque_dynamics_limit = np.array([0, 0, 0, 0, 0, 0])
        self.torque_static_limit = np.array([0, 0, 0, 0, 0, 0])

        self.tau_j = []

        self.qn = np.array([0, 0, 0, 0, 0, 0])  # degree

        self.joint_limit = [
            -100,
            100,
            -100,
            100,
            -100,
            100,
            -100,
            100,
            -100,
            100,
            -100,
            100,
        ]
        ## 數值法 求取工作空間
        # 關節角限位
        self.q1_s = -160
        self.q1_end = 160
        self.q2_s = -160
        self.q2_end = 160
        self.q3_s = -160
        self.q3_end = 160
        self.q4_s = -160
        self.q4_end = 160
        self.q5_s = -160
        self.q5_end = 160
        self.q6_s = -160
        self.q6_end = 160
        # 計算參數
        self.step = 20  # 計算步距 % 解析度   # original = 20
        # t=0:1:(q5_end-q5_s)/step # 產生時間向量
        step1 = (self.q1_end - self.q1_s) / self.step
        step2 = (self.q2_end - self.q2_s) / self.step
        step3 = (self.q3_end - self.q3_s) / self.step
        step4 = (self.q4_end - self.q4_s) / self.step
        step5 = (self.q5_end - self.q5_s) / self.step
        step6 = (self.q6_end - self.q6_s) / self.step
        self.step_num = int(step1 * step2 * step3 * step4 * step5)
        self.T_cell = step1 * step2 * step3 * step4 * step5
        self.T = np.zeros((3, 1))
        self.T_x = np.zeros((1, self.step_num))
        self.T_y = np.zeros((1, self.step_num))
        self.T_z = np.zeros((1, self.step_num))

        N = 100
        (Q2, Q3) = np.meshgrid(np.linspace(-pi, pi, N), np.linspace(-pi, pi, N))
        M11 = np.zeros((N, N))
        M12 = np.zeros((N, N))
        for i in range(N):
            for j in range(N):
                M = self.robot.inertia(np.r_[0, Q2[i, j], Q3[i, j], 0, 0, 0])
                M11[i, j] = M[0, 0]
                M12[i, j] = M[0, 1]

        self.xlsx_outpath = "./xlsx/"
        self.pic_outpath = "./picture/"
        # trajectory path generate postion velocity acceleration torque information parameter
        self.time_array = []
        self.tau_array = []
        self.point_nums = []
        self.traj_position = []
        self.traj_velocities = []
        self.traj_accelerations = []

        # ros topic 
        self.pub_dyna_space_progress = rospy.Publisher(
            "/dyna_space_progress", cal_process, queue_size=10
        )
        self.cal_process = cal_process()
        self.sub_taskcmd = rospy.Subscriber("/cal_command", cal_cmd, self.cmd_callback)
        self.sub_dyna = rospy.Subscriber(
            "/dynamics_data", dyna_data, self.dyna_callback
        )
        self.sub_dyna_space = rospy.Subscriber(
            "/dynamics_space_data", dyna_space_data, self.dyna_space_callback
        )
        self.sub_planned_path = rospy.Subscriber(
            "/move_group/display_planned_path",
            DisplayTrajectory,
            self.planned_path_callback,
        )
        self.specified_parameter_design = rospy.Subscriber(
            "/specified_parameter_design",
            specified_parameter_design,
            self.specified_parameter_design_callback,
        )

    def robot_rebuild(self):
        self.robot.__init__()
        rospy.loginfo("robot rebuild")

    def cmd_callback(self, data):
        self.cmd = data.cmd
        rospy.loginfo("I heard command is %s", data.cmd)

    def dyna_callback(self, data):
        self.joint_angle = data.joint_angle[0:6]
        self.payload = data.payload
        self.payload_position = data.payload_position
        self.vel = data.vel[0:6]
        self.acc = data.acc[0:6]
        rospy.loginfo("I heard command is %s", self.joint_angle)
        rospy.loginfo("I heard command is %s", self.payload)
        rospy.loginfo("I heard command is %s", self.payload_position)
        rospy.loginfo("I heard command is %s", self.vel)
        rospy.loginfo("I heard command is %s", self.acc)

    def dyna_space_callback(self, data):
        self.payload_space = data.payload
        self.payload_position_space = data.payload_position
        self.axis_num = data.analysis_axis
        rospy.loginfo("I heard command is %s", data.payload)
        rospy.loginfo("I heard command is %s", data.payload_position)
        rospy.loginfo("I heard command is %s", data.analysis_axis)

    def planned_path_callback(self, data):
        self.model_id = data.model_id
        self.robot_trajectory = data.trajectory
        self.robotstate = data.trajectory_start

        points_num = len(self.robot_trajectory[0].joint_trajectory.points)
        positions = []
        velocities = []
        accelerations = []
        time_from_start = []
        time_from_start_secs = []
        time_from_start_nsecs = []
        time_array = []
        for i in range(points_num):
            positions.append(
                self.robot_trajectory[0].joint_trajectory.points[i].positions
            )
            velocities.append(
                self.robot_trajectory[0].joint_trajectory.points[i].velocities
            )
            accelerations.append(
                self.robot_trajectory[0].joint_trajectory.points[i].accelerations
            )
            time_from_start.append(
                self.robot_trajectory[0].joint_trajectory.points[i].time_from_start
            )
            time_from_start_secs.append(
                self.robot_trajectory[0].joint_trajectory.points[i].time_from_start.secs
            )
            time_from_start_nsecs.append(
                self.robot_trajectory[0]
                .joint_trajectory.points[i]
                .time_from_start.nsecs
            )

        self.trajectory_dynamics_calc(
            points_num, positions, velocities, accelerations, time_from_start
        )
        time_array_secs = np.array(time_from_start_secs)
        time_array_nsecs = np.array(time_from_start_nsecs)
        time_array = time_array_secs + time_array_nsecs * (10) ** (-9)
        self.time_array = time_array
        self.tau_array = np.array(self.tau_j_array)
        self.traj_position = np.array(positions)
        self.traj_velocities = np.array(velocities)
        self.traj_accelerations = np.array(accelerations)
        rospy.loginfo("You can plot the trajectory information")

    def trajectory_torque_excel_write(self):
        excel_file = Workbook()
        sheet = excel_file.active
        sheet["A1"] = "time_array"
        sheet["B1"] = "torque 1"
        sheet["C1"] = "torque 2"
        sheet["D1"] = "torque 3"
        sheet["E1"] = "torque 4"
        sheet["F1"] = "torque 5"
        sheet["G1"] = "torque 6"
        sheet["H1"] = "traj_position 1"
        sheet["I1"] = "traj_position 2"
        sheet["J1"] = "traj_position 3"
        sheet["K1"] = "traj_position 4"
        sheet["L1"] = "traj_position 5"
        sheet["M1"] = "traj_position 6"
        sheet["N1"] = "traj_velocities 1"
        sheet["O1"] = "traj_velocities 2"
        sheet["P1"] = "traj_velocities 3"
        sheet["Q1"] = "traj_velocities 4"
        sheet["R1"] = "traj_velocities 5"
        sheet["S1"] = "traj_velocities 6"
        sheet["T1"] = "traj_accelerations 1"
        sheet["U1"] = "traj_accelerations 2"
        sheet["V1"] = "traj_accelerations 3"
        sheet["W1"] = "traj_accelerations 4"
        sheet["X1"] = "traj_accelerations 5"
        sheet["Y1"] = "traj_accelerations 6"
        
        for i in range(len(self.time_array)):
            sheet.cell(row=i + 2, column=1).value = self.time_array[i]
            sheet.cell(row=i + 2, column=2).value = self.tau_array[i][0]
            sheet.cell(row=i + 2, column=3).value = self.tau_array[i][1]
            sheet.cell(row=i + 2, column=4).value = self.tau_array[i][2]
            sheet.cell(row=i + 2, column=5).value = self.tau_array[i][3]
            sheet.cell(row=i + 2, column=6).value = self.tau_array[i][4]
            sheet.cell(row=i + 2, column=7).value = self.tau_array[i][5]
            sheet.cell(row=i + 2, column=8).value = self.traj_position[i][0]
            sheet.cell(row=i + 2, column=9).value = self.traj_position[i][1]
            sheet.cell(row=i + 2, column=10).value = self.traj_position[i][2]
            sheet.cell(row=i + 2, column=11).value = self.traj_position[i][3]
            sheet.cell(row=i + 2, column=12).value = self.traj_position[i][4]
            sheet.cell(row=i + 2, column=13).value = self.traj_position[i][5]
            sheet.cell(row=i + 2, column=14).value = self.traj_velocities[i][0]
            sheet.cell(row=i + 2, column=15).value = self.traj_velocities[i][1]
            sheet.cell(row=i + 2, column=16).value = self.traj_velocities[i][2]
            sheet.cell(row=i + 2, column=17).value = self.traj_velocities[i][3]
            sheet.cell(row=i + 2, column=18).value = self.traj_velocities[i][4]
            sheet.cell(row=i + 2, column=19).value = self.traj_velocities[i][5]
            sheet.cell(row=i + 2, column=20).value = self.traj_accelerations[i][0]
            sheet.cell(row=i + 2, column=21).value = self.traj_accelerations[i][1]
            sheet.cell(row=i + 2, column=22).value = self.traj_accelerations[i][2]
            sheet.cell(row=i + 2, column=23).value = self.traj_accelerations[i][3]
            sheet.cell(row=i + 2, column=24).value = self.traj_accelerations[i][4]
            sheet.cell(row=i + 2, column=25).value = self.traj_accelerations[i][5]

        file_name = self.xlsx_outpath + "/trajectory_torque" + "dynamixel"+".xlsx"
        excel_file.save(file_name)
        rospy.loginfo("trajectory torque excel write down.")
        print("================================")
# smooth plot the trajectory
    # def trajectory_torque_plot(self):
    #     x_smooth = np.linspace(self.time_array.min(), self.time_array.max(), 300)
    #     fig, ax = plt.subplots(4,6,figsize=(24,24))
    #     axis = 6
    #     for i in range(axis):
    #         y_smooth = make_interp_spline(self.time_array, self.tau_array[:, i])(x_smooth)
    #         ax[0,i].plot(x_smooth, y_smooth, "b--")
    #         string_axis = str(i + 1)
    #         ax[0,i].set_title("torque " + string_axis, {"fontsize": 10})  # 設定圖標題及其文字大小
    #         ax[0,i].set_xlabel("sec")
    #         ax[0,i].set_xlabel("N*m")

    #     for i in range(axis):
    #         y_smooth = make_interp_spline(self.time_array, self.traj_position[:, i])(x_smooth)
    #         ax[1,i].plot(x_smooth, y_smooth, "b--")
    #         string_axis = str(i + 1)
    #         ax[1,i].set_title("position " + string_axis, {"fontsize": 10})  # 設定圖標題及其文字大小
    #         ax[1,i].set_xlabel("sec")
    #         ax[1,i].set_xlabel("pos")

    #     for i in range(axis):
    #         y_smooth = make_interp_spline(self.time_array, self.traj_velocities[:, i])(x_smooth)
    #         ax[2,i].plot(x_smooth, y_smooth, "b--")
    #         string_axis = str(i + 1)
    #         ax[2,i].set_title("velocity " + string_axis, {"fontsize": 10})  # 設定圖標題及其文字大小
    #         ax[2,i].set_xlabel("sec")
    #         ax[2,i].set_xlabel("rad/sec")

    #     for i in range(axis):
    #         y_smooth = make_interp_spline(self.time_array, self.traj_accelerations[:, i])(x_smooth)
    #         ax[3,i].plot(x_smooth, y_smooth, "b--")
    #         string_axis = str(i + 1)
    #         ax[3,i].set_title("acceleration " + string_axis, {"fontsize": 10})
    #         ax[3,i].set_xlabel("sec")
    #         ax[3,i].set_xlabel("rad/sec^2")
    #     fig.tight_layout()
    #     plt.show()

    def trajectory_torque_plot(self):
        # x_smooth = np.linspace(self.time_array.min(), self.time_array.max(), 300)
        fig, ax = plt.subplots(4,6,figsize=(20,10))
        axis = 6
        for i in range(axis):
            # y_smooth = make_interp_spline(self.time_array, self.tau_array[:, i])(x_smooth)
            ax[0,i].plot(self.time_array, self.tau_array[:, i], "b--")
            string_axis = str(i + 1)
            ax[0,i].set_title("torque " + string_axis, {"fontsize": 10})  # 設定圖標題及其文字大小
            ax[0,i].set_xlabel("sec")
            ax[0,i].set_xlabel("N*m")

        for i in range(axis):
            # y_smooth = make_interp_spline(self.time_array, self.traj_position[:, i])(x_smooth)
            ax[1,i].plot(self.time_array, self.traj_position[:, i], "b--")
            string_axis = str(i + 1)
            ax[1,i].set_title("position " + string_axis, {"fontsize": 10})  # 設定圖標題及其文字大小
            ax[1,i].set_xlabel("sec")
            ax[1,i].set_xlabel("pos")

        for i in range(axis):
            # y_smooth = make_interp_spline(self.time_array, self.traj_velocities[:, i])(x_smooth)
            ax[2,i].plot(self.time_array, self.traj_velocities[:, i], "b--")
            string_axis = str(i + 1)
            ax[2,i].set_title("velocity " + string_axis, {"fontsize": 10})  # 設定圖標題及其文字大小
            ax[2,i].set_xlabel("sec")
            ax[2,i].set_xlabel("rad/sec")

        for i in range(axis):
            # y_smooth = make_interp_spline(self.time_array, self.traj_accelerations[:, i])(x_smooth)
            ax[3,i].plot(self.time_array, self.traj_accelerations[:, i], "b--")
            string_axis = str(i + 1)
            ax[3,i].set_title("acceleration " + string_axis, {"fontsize": 10})
            ax[3,i].set_xlabel("sec")
            ax[3,i].set_xlabel("rad/sec^2")
            
        fig.tight_layout()
        plt.show()

    def specified_parameter_design_callback(self, data):
        self.axis_2_length = data.axis_2_length
        self.axis_3_length = data.axis_3_length
        self.arm_weight = data.arm_weight
        self.payload = data.payload
        self.radius = data.radius  # 半径
        self.DoF = data.DoF
        self.joint_limit = data.joint_limit
        rospy.loginfo("I heard axis_2_length is %s", self.axis_2_length)
        rospy.loginfo("I heard axis_3_length is %s", self.axis_3_length)
        rospy.loginfo("I heard arm_weight is %s", self.arm_weight)
        rospy.loginfo("I heard payload is %s", self.payload)
        rospy.loginfo("I heard radius is %s", self.radius)
        rospy.loginfo("I heard DoF is %s", self.DoF)
        rospy.loginfo("I heard joint limit is %s", self.joint_limit)

    def payload_set(self):
        self.robot.payload(20, [0, 0, 0])  # set payload

    def dynamics_space_cal_Monte_Carlo(self, joint_limit):
        """
        Through the "dynamics space" page in the interface to calculate the dynamics of the robot
        """
        qd = np.r_[0, 1, 0, 0, 0, 0]
        # print("qd:",qd)
        self.robot.coriolis(self.robot.qn, qd) @ qd
        self.robot.rne(self.robot.qn, np.zeros((6,)), np.zeros((6,)))

        torque = np.array([np.zeros(shape=6)])
        # axis_angle = np.array([np.zeros(shape=6)])
        axis_angle = []
        append_torque_limit_list = []
        temp_torque_max = []
        temp_torque_min = []
        Torque_Max = []
        # 蒙地卡羅法正運動學計算工作空間
        start = time.time()
        rospy.loginfo("The time used to execute this is given below")
        i = 0
        # 角度轉換
        du = pi / 180
        # 度
        radian = 180 / pi
        # 弧度

        # fig = plt.figure()
        # self.ax = plt.subplot(111, projection='3d')
        # self.ax_2d = plt.subplot(111)

        self.q1_s = joint_limit[0]
        self.q1_end = joint_limit[1]
        self.q2_s = joint_limit[2]
        self.q2_end = joint_limit[3]
        self.q3_s = joint_limit[4]
        self.q3_end = joint_limit[5]
        self.q4_s = joint_limit[6]
        self.q4_end = joint_limit[7]
        self.q5_s = joint_limit[8]
        self.q5_end = joint_limit[9]
        self.q6_s = joint_limit[10]
        self.q6_end = joint_limit[11]

        self.robot.payload(
            self.payload_space, self.payload_position_space
        )  # set payload
        N = 20000
        theta1 = self.q1_end + (self.q1_end - self.q1_s) * np.random.rand(N, 1)
        theta2 = self.q2_end + (self.q2_end - self.q2_s) * np.random.rand(N, 1)
        theta3 = self.q3_end + (self.q3_end - self.q3_s) * np.random.rand(N, 1)
        theta4 = self.q4_end + (self.q4_end - self.q4_s) * np.random.rand(N, 1)
        theta5 = self.q5_end + (self.q5_end - self.q5_s) * np.random.rand(N, 1)
        theta6 = self.q6_end + (self.q6_end - self.q6_s) * np.random.rand(N, 1)

        for i in range(N):
            percent = (i + 1) / N * 100
            self.cal_process.dyna_space_progress = int(percent)
            self.pub_dyna_space_progress.publish(self.cal_process)
            q1 = theta1[i, 0]
            q2 = theta2[i, 0]
            q3 = theta3[i, 0]
            q4 = theta4[i, 0]
            q5 = theta5[i, 0]
            q6 = theta6[i, 0]
            axis_angle.append([q1, q2, q3, q4, q5, q6])
            self.T = self.robot.fkine(
                [q1 * du, q2 * du, q3 * du, q4 * du, q5 * du, q6 * du]
            )
            load = np.array(
                [
                    self.robot.gravload(
                        [q1 * du, q2 * du, q3 * du, q4 * du, q5 * du, q6 * du]
                    )
                ]
            )
            torque = np.append(torque, load, axis=0)
            self.T_x[0, i] = self.T.t[0]
            self.T_y[0, i] = self.T.t[1]
            self.T_z[0, i] = self.T.t[2]
            i = i + 1

        end = time.time()
        print("繪製工作空間運行時間：%f sec" % (end - start))

        for i in range(6):
            axis = i
            excel_file = Workbook()
            sheet = excel_file.active
            sheet["A1"] = "torque 1"
            sheet["B1"] = "torque 2"
            sheet["C1"] = "torque 3"
            sheet["D1"] = "torque 4"
            sheet["E1"] = "torque 5"
            sheet["F1"] = "torque 6"
            sheet["G1"] = "angle 1"
            sheet["H1"] = "angle 2"
            sheet["I1"] = "angle 3"
            sheet["J1"] = "angle 4"
            sheet["K1"] = "angle 5"
            sheet["L1"] = "angle 6"
            print("================================")
            print("軸%d torque正最大值時, 各軸torque, 末端位置, 各軸角度" % (axis + 1))
            print("axis_max_torque:", torque[np.argmax(torque[:, axis])])
            toque_max_index = np.argmax(torque[:, axis])
            print("toque_max_index:", toque_max_index)
            print("ik_sol:", axis_angle[np.argmax(torque[:, axis])])
            print("================================")
            print("軸%d torque負最大值時, 各軸torque, 末端位置, 各軸角度" % (axis + 1))
            print("axis_max_torque:", torque[np.argmin(torque[:, axis])])
            toque_min_index = np.argmin(torque[:, axis])
            print("toque_min_index:", toque_min_index)
            print("ik_sol:", axis_angle[np.argmin(torque[:, axis])])
            print("================================")
            temp_torque_max = torque[toque_max_index].tolist()
            temp_torque_max.extend(axis_angle[toque_max_index])
            temp_torque_min = torque[toque_min_index].tolist()
            temp_torque_min.extend(axis_angle[toque_min_index])
            append_torque_limit_list.append(temp_torque_max)
            append_torque_limit_list.append(temp_torque_min)

            Torque_Max.append(abs(torque[toque_max_index][i]))
            print("Torque_Max:", Torque_Max)
            self.torque_static_limit = Torque_Max

        for info in append_torque_limit_list:
            sheet.append(info)
        sheet.append(
            [
                "Torque 1 max",
                "Torque 2 max",
                "Torque 3 max",
                "Torque 4 max",
                "Torque 5 max",
                "Torque 6 max",
            ]
        )
        sheet.append(Torque_Max)

        file_name = self.xlsx_outpath + "/dynamics_static_torque_calc" + "dynamixel"+ ".xlsx"
        excel_file.save(file_name)

        rospy.loginfo("output dynamics static torque_calc axis excel file down.")
        print("================================")

    def Workspace_cal_Monte_Carlo(self):
        """
        Through the "Work space" page in the interface to calculate of the robot
        """
        i = 0
        # 角度轉換
        du = pi / 180
        # 度
        radian = 180 / pi
        # 弧度

        # fig = plt.figure()
        # self.ax = plt.subplot(111, projection='3d')
        # self.ax_2d = plt.subplot(111)

        self.q1_s = -160
        self.q1_end = 160
        self.q2_s = -160
        self.q2_end = 160
        self.q3_s = -160
        self.q3_end = 160
        self.q4_s = -160
        self.q4_end = 160
        self.q5_s = -160
        self.q5_end = 160
        self.q6_s = -160
        self.q6_end = 160
        N = 20000
        theta1 = self.q1_end + (self.q1_end - self.q1_s) * np.random.rand(N, 1)
        theta2 = self.q2_end + (self.q2_end - self.q2_s) * np.random.rand(N, 1)
        theta3 = self.q3_end + (self.q3_end - self.q3_s) * np.random.rand(N, 1)
        theta4 = self.q4_end + (self.q4_end - self.q4_s) * np.random.rand(N, 1)
        theta5 = self.q5_end + (self.q5_end - self.q5_s) * np.random.rand(N, 1)
        theta6 = self.q6_end + (self.q6_end - self.q6_s) * np.random.rand(N, 1)

        for i in range(N):
            percent = (i + 1) / N * 100
            q1 = theta1[i, 0]
            q2 = theta2[i, 0]
            q3 = theta3[i, 0]
            q4 = theta4[i, 0]
            q5 = theta5[i, 0]
            q6 = theta6[i, 0]
            self.T = self.robot.fkine(
                [q1 * du, q2 * du, q3 * du, q4 * du, q5 * du, q6 * du]
            )
            self.T_x[0, i] = self.T.t[0]
            self.T_y[0, i] = self.T.t[1]
            self.T_z[0, i] = self.T.t[2]
            i = i + 1

    def plot_space_scan(self):
        """
        Calculate the results through the "dynamics space" paging in the interface,
        and draw the distribution points
        """
        print("T_x:{} (meter)".format(self.T_x[0, :].max() - self.T_x[0, :].min()))
        print("T_y:{} (meter)".format(self.T_y[0, :].max() - self.T_y[0, :].min()))
        print("T_z:{} (meter)".format(self.T_z[0, :].max() - self.T_z[0, :].min()))
        print("================================")
        fig, axes = plt.subplots(nrows=3, ncols=1, figsize=(5, 15))

        axes[0].scatter(self.T_x[0, :], self.T_y[0, :], c="r", marker="o")

        axes[0].set_xlabel("x (meter)")
        axes[0].set_ylabel("y (meter)")

        # ax2 = plt.subplot(nrows=1, ncols=2, figsize=(5,5))
        axes[1].scatter(self.T_x[0, :], self.T_z[0, :], c="r", marker="o")

        axes[1].set_xlabel("x (meter)")
        axes[1].set_ylabel("z (meter)")

        # ax3 = plt.subplot(nrows=1, ncols=3, figsize=(5,5))
        axes[2].scatter(self.T_y[0, :], self.T_z[0, :], c="r", marker="o")

        axes[2].set_xlabel("y (meter)")
        axes[2].set_ylabel("z (meter)")

        ln1, ln2, ln3 = self.robot.return_configuration()

        # Ang_arr = np.array([- math.pi * 125 / 180, math.pi * 85 / 180, -math.pi * 145 / 180, math.pi * 95 / 180, -math.pi * 115 / 180, math.pi * 115 / 180])
        # Ang_arr1 = np.array([- math.pi * 170 / 180, math.pi * 170 / 180])
        Ang_arr = np.array(
            [
                math.pi * self.joint_limit[2] / 180,
                math.pi * self.joint_limit[3] / 180,
                math.pi * self.joint_limit[4] / 180,
                math.pi * self.joint_limit[5] / 180,
                math.pi * self.joint_limit[6] / 180,
                math.pi * self.joint_limit[7] / 180,
            ]
        )
        Ang_arr1 = np.array(
            [math.pi * self.joint_limit[0] / 180, math.pi * self.joint_limit[1] / 180]
        )
        L_arr = np.array([ln1, ln2, ln3])
        Wspace = arm_workspace_plane(
            ang_arr1=Ang_arr1, ang_arr=Ang_arr, link_lengths=L_arr
        )
        bb, aa = Wspace.xy_Wspace_mod(Wspace.L_arr, Wspace.Ang_arr_angle1, 360)
        CA, maxmin = Wspace.Wspace_mod(Wspace.L_arr, Wspace.Ang_arr, res=30)
        Wspace.plot_Wspace_mod(CA, maxmin, False)

    def dynamics_calc(self):
        """
        The "dynamics" page in the interface simply calculates the torque of each axis,

        Input: payload, current position, current velocity, current acceleration
        """
        qn = [0, 0, 0, 0, 0, 0]
        deg = pi / 180

        for i in range(6):
            qn[i] = self.joint_angle[i] * deg
        # print(self.robot.gravload(qn))
        self.robot.payload(self.payload, self.payload_position)  # set payload

        self.qn = qn
        self.tau_j = self.robot.rne(self.qn, self.vel, self.acc)
        print("tau_j:", self.tau_j)
        # axis = 2
        self.robot.plot(self.qn)

        plt.savefig(path.join(self.pic_outpath, "dataname_dynamics_calc.png"))
        plt.close()
        # excel output
        # 建立excel空白活頁簿
        excel_file = Workbook()
        # 建立一個工作表
        sheet = excel_file.active
        # 先填入第一列的欄位名稱
        sheet["A1"] = "axis 1"
        sheet["B1"] = "axis 2"
        sheet["C1"] = "axis 3"
        sheet["D1"] = "axis 4"
        sheet["E1"] = "axis 5"
        sheet["F1"] = "axis 6"
        sheet.append(
            [
                self.tau_j[0],
                self.tau_j[1],
                self.tau_j[2],
                self.tau_j[3],
                self.tau_j[4],
                self.tau_j[5],
            ]
        )
        file_name = self.xlsx_outpath + "/dynamics_calc"+ "dynamixel" + ".xlsx"
        # excel_file.save('dynamics_calc.xlsx')
        excel_file.save(file_name)

    def trajectory_dynamics_calc(self, num, pos, vel, acc, time):
        """
        Obtain the position, velocity, acceleration of the trajectory through moveit,
        and estimate the required torque
        """
        self.robot.payload(self.payload, self.payload_position)  # set payload
        self.tau_j_array = []
        for i in range(num):
            self.tau_j = self.robot.rne(pos[i], vel[i], acc[i])
            self.tau_j_array.append(self.tau_j)

    def arm_plot(self):
        """
        Through the "plot" button on the interface, draw the current posture of the arm
        """
        qn = [0, 0, 0, 0, 0, 0]
        deg = pi / 180

        for i in range(6):
            qn[i] = self.joint_angle[i] * deg

        self.qn = qn
        # fig = plt.figure()
        # self.robot.plot(self.qn,dt=0, block = False, loop=False)
        print("arm_plot:", self.qn)
        self.robot.plot(
            self.qn, backend="pyplot", block=False, vellipse=False, fellipse=False
        )
        # self.robot.plot(self.qt.q, backend=self.args.backend, block=False, movie="trajectory_generation.gif", vellipse=False, fellipse=False)
        plt.show()

    def plot_close(self):
        """
        Drawing on the interface is turned off
        """
        plt.close()

    def dynamics_torque_limit(self):
        """
        Calculate the maximum torque required by

        each axis when the arm of each axis is the longest and the acceleration is the highest
        """
        torque = np.array([np.zeros(shape=6)])
        # axis_angle = np.array([np.zeros(shape=6)])
        axis_angle = []
        append_torque_limit_list = []
        temp_torque_max = []
        temp_torque_min = []
        Torque_Max = []
        # 窮舉法正運動學計算工作空間
        start = time.time()
        rospy.loginfo("The time used to execute this is given below")
        # 角度轉換
        du = pi / 180
        # 度
        radian = 180 / pi
        # 弧度
        self.robot.payload(self.payload, self.payload_position)  # set payload
        torque = np.array([np.zeros(shape=6)])
        q_list = [0, 90, -90, 180, -180]
        T_cell = (
            len(q_list)
            * len(q_list)
            * len(q_list)
            * len(q_list)
            * len(q_list)
            * len(q_list)
        )

        T = np.zeros((3, 1))
        T_x = np.zeros(T_cell)
        T_y = np.zeros(T_cell)
        T_z = np.zeros(T_cell)

        for i in range(len(q_list)):
            q1 = q_list[i]
            percent = i / T_cell * 100
            for j in range(len(q_list)):
                q2 = q_list[j]
                for k in range(len(q_list)):
                    q3 = q_list[k]
                    for l in range(len(q_list)):
                        q4 = q_list[l]
                        for m in range(len(q_list)):
                            q5 = q_list[m]
                            for n in range(len(q_list)):
                                q6 = q_list[n]
                                axis_angle.append([q1, q2, q3, q4, q5, q6])
                                # T = self.robot.fkine([q1*du, q2*du, q3*du, q4*du, q5*du, q6*du])
                                load = np.array(
                                    [
                                        self.robot.rne(
                                            [
                                                q1 * du,
                                                q2 * du,
                                                q3 * du,
                                                q4 * du,
                                                q5 * du,
                                                q6 * du,
                                            ],
                                            self.vel,
                                            self.acc,
                                        )
                                    ]
                                )
                                torque = np.append(torque, load, axis=0)
                                # T_x[i] = T.t[0]
                                # T_y[i] = T.t[1]
                                # T_z[i] = T.t[2]
        end = time.time()
        # print("繪製工作空間運行時間：%f sec" % (end - start))

        for i in range(6):
            axis = i
            excel_file = Workbook()
            sheet = excel_file.active
            sheet["A1"] = "torque 1"
            sheet["B1"] = "torque 2"
            sheet["C1"] = "torque 3"
            sheet["D1"] = "torque 4"
            sheet["E1"] = "torque 5"
            sheet["F1"] = "torque 6"
            sheet["G1"] = "angle 1"
            sheet["H1"] = "angle 2"
            sheet["I1"] = "angle 3"
            sheet["J1"] = "angle 4"
            sheet["K1"] = "angle 5"
            sheet["L1"] = "angle 6"
            print("================================")
            print("軸%d torque正最大值時, 各軸torque, 末端位置, 各軸角度" % (axis + 1))
            print("axis_max_torque:", torque[np.argmax(torque[:, axis])])
            toque_max_index = np.argmax(torque[:, axis])
            print("toque_max_index:", toque_max_index)
            print("ik_sol:", axis_angle[np.argmax(torque[:, axis])])
            print("================================")
            print("軸%d torque負最大值時, 各軸torque, 末端位置, 各軸角度" % (axis + 1))
            print("axis_max_torque:", torque[np.argmin(torque[:, axis])])
            toque_min_index = np.argmin(torque[:, axis])
            print("toque_min_index:", toque_min_index)
            print("ik_sol:", axis_angle[np.argmin(torque[:, axis])])
            print("================================")
            temp_torque_max = torque[toque_max_index].tolist()
            temp_torque_max.extend(axis_angle[toque_max_index])
            temp_torque_min = torque[toque_min_index].tolist()
            temp_torque_min.extend(axis_angle[toque_min_index])
            append_torque_limit_list.append(temp_torque_max)
            append_torque_limit_list.append(temp_torque_min)

            Torque_Max.append(abs(torque[toque_max_index][i]))
            print("Torque_Max:", Torque_Max)
            self.torque_dynamics_limit = Torque_Max

        for info in append_torque_limit_list:
            sheet.append(info)
        sheet.append(
            [
                "Torque 1 max",
                "Torque 2 max",
                "Torque 3 max",
                "Torque 4 max",
                "Torque 5 max",
                "Torque 6 max",
            ]
        )
        sheet.append(Torque_Max)

        file_name = self.xlsx_outpath + "/dynamics_limit_torque_calc" + "dynamixel"+ ".xlsx"
        # 調整列寬 對齊
        ncols = 1
        nrows = 1

        for i in range(sheet.max_column):
            col_letter = get_column_letter(ncols)
            # print(type(col_letter))
            sheet.column_dimensions[col_letter].width = 20
            sheet[col_letter + str(1)].font = Font(
                name="Courier", size=14, color="EF0A1D"
            )
            sheet[col_letter + str(14)].font = Font(
                name="Courier", size=14, color="EF0A1D"
            )
            ncols = ncols + 1

            for i in range(sheet.max_row):
                alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    text_rotation=0,
                    wrap_text=True,
                )
                sheet[col_letter + str(i + 1)].alignment = alignment
        # 調整行高
        sheet.row_dimensions[1].height = 40
        # 調整行高
        sheet.row_dimensions[14].height = 40
        excel_file.save(file_name)
        rospy.loginfo("output dynamics_torque_limit_calc_axis excel file down.")
        print("================================")

    def CJM_select(self):
        # 初始化
        self.robot.__init__()
        rospy.loginfo("robot rebuild")
        motor = motor_data()
        res = motor.TECO_member
        print("torque dynamics static limit:", self.torque_static_limit)
        print("torque dynamics limit:", self.torque_dynamics_limit)

        static_sol_module = []
        dynamic_sol_module = []
        show_static = []
        show_dynamic = []
        static_data_input_head = [
            "static payload",
            "static payload position x",
            "static payload position y",
            "static payload position z",
        ]
        static_data_input = []
        dynamic_data_payload_input_head = [
            "dynamic payload",
            "dynamic payload position x",
            "dynamic payload position y",
            "dynamic payload position z",
        ]
        dynamic_data_vel_input_head = [
            "velocity of axis 1",
            "velocity of axis 2",
            "velocity of axis 3",
            "velocity of axis 4",
            "velocity of axis 5",
            "velocity of axis 6",
        ]
        dynamic_data_acc_input_head = [
            "acceralation of axis 1",
            "acceralation of axis 2",
            "acceralation of axis 3",
            "acceralation of axis 4",
            "acceralation of axis 5",
            "acceralation of axis 6",
        ]
        dynamic_data_input = []
        alignment = Alignment(
            horizontal="center", vertical="center", text_rotation=0, wrap_text=True
        )

        # excel output
        # 建立excel空白活頁簿
        excel_file = Workbook()
        # 建立一個工作表
        sheet = excel_file.active
        # 先填入第一列的欄位名稱
        sheet["A1"] = "model number of axis 1"
        sheet["B1"] = "model number of axis 2"
        sheet["C1"] = "model number of axis 3"
        sheet["D1"] = "model number of axis 4"
        sheet["E1"] = "model number of axis 5"
        sheet["F1"] = "model number of axis 6"

        for j in range(6):
            for i in range(len(res)):
                if res.rated_torque[i] > self.torque_static_limit[j]:
                    print("軸數:", j + 1)
                    print("關節型號:", i)
                    static_sol_module.append(i)
                    # show_static.append("static")
                    break
                # over size can load
                elif max(res.rated_torque) <= self.torque_static_limit[j]:
                    print("軸數:", j + 1)
                    print("關節型號: fail") #
                    static_sol_module.append("fail")
                    # show_static.append("static")
                    break

        show_static = ["static analysis result"]
        sheet.append(show_static)
        sheet.merge_cells("A2:F2")
        sheet["A2"].font = Font(name="Courier", size=12, color="8A3F2F")
        sheet["A2"].alignment = alignment
        sheet.append(static_sol_module)
        sheet.append([])

        for j in range(6):
            for i in range(len(res)):
                if res.rated_torque[i] > self.torque_dynamics_limit[j]:
                    print("軸數:", j + 1)
                    print("關節型號:", i)
                    dynamic_sol_module.append(i)
                    # show_dynamic.append("dynamic")
                    break
                # over size can load
                elif max(res.rated_torque) <= self.torque_dynamics_limit[j]:
                    print("軸數:", j + 1)
                    print("關節型號: fail") #
                    dynamic_sol_module.append("fail")
                    # show_static.append("static")
                    break
                
        show_dynamic = ["dynamic analysis result"]
        sheet.append(show_dynamic)
        sheet.merge_cells("A5:F5")
        sheet["A5"].font = Font(name="Courier", size=12, color="8A3F2F")
        sheet["A5"].alignment = alignment
        sheet.append(dynamic_sol_module)
        sheet.append([])

        show_static = ["static analysis input data"]
        sheet.append(show_static)
        sheet.merge_cells("A8:F8")
        sheet["A8"].font = Font(name="Courier", size=12, color="8A3F2F")
        sheet["A8"].alignment = alignment
        sheet.append(static_data_input_head)
        static_data_input.append(self.payload_space)
        static_data_input.append(self.payload_position_space[0])
        static_data_input.append(self.payload_position_space[1])
        static_data_input.append(self.payload_position_space[2])
        sheet.append(static_data_input)
        sheet.append([])

        show_dynamic = ["dynamic analysis input data"]
        sheet.append(show_dynamic)
        sheet.merge_cells("A12:F12")
        sheet["A12"].font = Font(name="Courier", size=12, color="8A3F2F")
        sheet["A12"].alignment = alignment
        sheet.append(dynamic_data_payload_input_head)
        dynamic_data_input.append(self.payload)
        dynamic_data_input.append(self.payload_position[0])
        dynamic_data_input.append(self.payload_position[1])
        dynamic_data_input.append(self.payload_position[2])
        sheet.append(dynamic_data_input)

        sheet.append(dynamic_data_vel_input_head)
        dynamic_data_input = []
        for i in range(6):
            dynamic_data_input.append(self.vel[i])
        sheet.append(dynamic_data_input)

        sheet.append(dynamic_data_acc_input_head)
        dynamic_data_input = []
        for i in range(6):
            dynamic_data_input.append(self.acc[i])
        sheet.append(dynamic_data_input)

        file_name = self.xlsx_outpath + "/CJM_select"+ "dynamixel"+".xlsx"
        # 調整列寬
        ncols = 1
        for i in range(sheet.max_column):
            col_letter = get_column_letter(ncols)
            # print(type(col_letter))
            sheet.column_dimensions[col_letter].width = 25
            sheet[col_letter + str(1)].font = Font(
                name="Courier", size=12, color="2F468A"
            )
            sheet[col_letter + str(1)].alignment = alignment
            sheet[col_letter + str(3)].font = Font(
                name="Courier", size=12, color="2F468A"
            )
            sheet[col_letter + str(3)].alignment = alignment
            sheet[col_letter + str(6)].font = Font(
                name="Courier", size=12, color="2F468A"
            )
            sheet[col_letter + str(6)].alignment = alignment
            # sheet[col_letter+str(14)].font = Font(name='Courier', size=14, color='EF0A1D')
            ncols = ncols + 1
        # 調整行高
        sheet.row_dimensions[1].height = 40
        # sheet['A'+str(i+1)].font = Font(name='Courier', size=14, color='EF0A1D')
        sheet["A1"].alignment = alignment
        sheet.append([])

        show_joint_module_model = ["The selected joint module model"]
        sheet.append(show_joint_module_model)
        sheet.merge_cells("A20:I20")
        sheet["A20"].font = Font(name="Courier", size=12, color="8A3F2F")
        sheet["A20"].alignment = alignment
        sheet.append(["model number"] + list(res.columns))
        for i in range(len(res.index)):
            sheet.append([i] + list(res.iloc[i]))

        excel_file.save(file_name)

    def excited_trajectory(
        self, q_list, q_list_2, q_list_3, q_list_4, q_list_5, q_list_6
    ):
        # 窮舉法正運動學計算工作空間
        start = time.time()
        print("The time used to execute this is given below")
        # 角度轉換
        du = pi / 180

    def robot_motor_random_build(self):
        self.robot.__init__()
        print("robot rebuild")
        motor = motor_data()
        # print(motor.TECO_member.head())
        # print(motor.TECO_member.groupby("rated_torque").mean())
        print(
            pd.concat(
                [
                    motor.TECO_member,
                    motor.Kollmorgen_member,
                    motor.UR_member,
                    motor.TM_member,
                ],
                axis=0,
            )
        )

    def task_set(self):
        for case in switch(self.cmd):
            if case(1):
                rospy.loginfo("Start Workspace Scan & success get subscriber data command")
                self.payload_set()
                # Dya.dynamics_space_cal()
                self.dynamics_space_cal_Monte_Carlo(self.joint_limit)
                # Dya.static_sol_output_axis()
                self.plot_space_scan()
                self.cmd = 0
                break

            if case(2):
                rospy.loginfo("Start Set payload & vel & acc analysis command")
                self.dynamics_calc()
                self.cmd = 0
                break
            # Select axis for dynamics space scan joint torque output
            if case(3):
                rospy.loginfo("Select axis for dynamics space scan joint torque output command")
                self.sol_output_axis()
                self.cmd = 0
                break

            if case(4):
                rospy.loginfo("Plot robotic arm command")
                self.arm_plot()
                self.cmd = 0
                break
            #
            if case(5):
                rospy.loginfo("Close plot command")
                self.plot_close()
                self.cmd = 0
                break

            if case(6):
                rospy.loginfo("Torque limit command")
                self.dynamics_torque_limit()
                self.cmd = 0
                break
            # arm data rebuild
            if case(7):
                rospy.loginfo("Robot rebuild command")
                self.robot_rebuild()
                self.cmd = 0
                break
            # arm data & motor data rebuild
            if case(8):
                rospy.loginfo("Robot motor random command")
                self.robot_motor_random_build()
                self.cmd = 0
                break
            # CJM select
            if case(9):
                rospy.loginfo("CJM select function command")
                self.CJM_select()
                self.cmd = 0
                break
            if case(10):
                rospy.loginfo("Plot trajectory information command")
                self.trajectory_torque_excel_write()
                self.trajectory_torque_plot()
                self.cmd = 0
                break
            if case():
                break

if __name__ == "__main__":
    rospy.init_node("Dynamics_dynamixel")

    Dya = Dynamics_dynamixel()
    while not rospy.is_shutdown():
        Dya.task_set()
