#!/usr/bin/env python3
# coding: utf-8
import rospy
import sys
from collections import namedtuple
# import dyna_space
from interface_control.msg import cal_cmd, dyna_data, dyna_space_data
import numpy as np
import roboticstoolbox as rtb
from spatialmath import *
from math import pi
from mpl_toolkits.mplot3d import Axes3D
import matplotlib.pyplot as plt
from matplotlib import cm
import argparse
import time
from moveit_msgs.msg import DisplayTrajectory, RobotTrajectory
from trajectory_msgs.msg import JointTrajectory, JointTrajectoryPoint
import geometry_msgs.msg
import csv
import openpyxl
from openpyxl import Workbook
from os import path
from scipy.interpolate import make_interp_spline # draw smooth 
np.set_printoptions(linewidth=100, formatter={'float': lambda x: f"{x:8.4g}" if abs(x) > 1e-10 else f"{0:8.4g}"})

from TECO_ARM1_robot import TECOARM1



class switch(object):
    def __init__(self, value):
        self.value = value
        self.fall = False

    def __iter__(self):
        """Return the match method once, then stop"""
        yield self.match
        raise StopIteration
    
    def match(self, *args):
        """Indicate whether or not to enter a case suite"""
        if self.fall or not args:
            return True
        elif self.value in args: # changed for v1.5, see below
            self.fall = True
            return True
        else:
            return False


class Dynamics_teco():
    def __init__(self):
        # self.pub_armstatus = rospy.Publisher("/reply_external_comm",peripheralCmd,queue_size=10)
        self.sub_taskcmd = rospy.Subscriber("/cal_command",cal_cmd, self.cmd_callback)
        self.sub_dyna = rospy.Subscriber("/dynamics_data",dyna_data, self.dyna_callback)
        self.sub_dyna_space = rospy.Subscriber("/dynamics_space_data",dyna_space_data, self.dyna_space_callback)
        self.sub_planned_path = rospy.Subscriber("/move_group/display_planned_path", DisplayTrajectory, self.planned_path_callback)
        # self.sub_planned_path = rospy.Subscriber("/move_group/display_planned_path",moveit_msgs.msg.JointTrajectory,self.planned_path_callback)
        self.cmd = 0
        self.teco = TECOARM1()
        # self.teco.plot(self.teco.qn, block=False)
        self.teco.gravload(self.teco.qn)
        self.teco.inertia(self.teco.qn)
        self.torque = np.array([np.zeros(shape=6)])

        self.payload = 0
        self.payload_position = np.array([0,0,0])
        self.joint_angle = np.array([0, 0, 0, 0, 0, 0, 0])  # degree
        self.vel = np.array([0, 0, 0, 0, 0, 0, 0])  # rad / sec
        self.acc = np.array([0, 0, 0, 0, 0, 0, 0])  # rad / sec2

        self.payload_space = 0
        self.payload_position_space = np.array([0,0,0])
        self.axis_num = 1

        self.tau_j = []

        self.qn = np.array([0,0,0,0,0,0]) # degree 
        
        ## 數值法 求取工作空間
        # 關節角限位
        self.q1_s=-160
        self.q1_end=160
        self.q2_s=-160
        self.q2_end=160
        self.q3_s=-160
        self.q3_end=160
        self.q4_s=-160
        self.q4_end=160
        self.q5_s=-160
        self.q5_end=160
        self.q6_s=-160
        self.q6_end=160
        # 計算參數
        self.step= 20 #計算步距 % 解析度   # original = 20
        # t=0:1:(q5_end-q5_s)/step # 產生時間向量 
        step1 = (self.q1_end - self.q1_s)/self.step 
        step2 = (self.q2_end - self.q2_s)/self.step 
        step3 = (self.q3_end - self.q3_s)/self.step 
        step4 = (self.q4_end - self.q4_s)/self.step 
        step5 = (self.q5_end - self.q5_s)/self.step
        step6 = (self.q6_end - self.q6_s)/self.step 
        self.step_num = int(step1*step2*step3*step4*step5)
        self.T_cell=step1*step2*step3*step4*step5
        self.T = np.zeros((3,1))
        self.T_x = np.zeros((1,self.step_num))
        self.T_y = np.zeros((1,self.step_num))
        self.T_z = np.zeros((1,self.step_num))
        
        N = 100
        (Q2, Q3) = np.meshgrid(np.linspace(-pi, pi, N), np.linspace(-pi, pi, N))
        M11 = np.zeros((N,N))
        M12 = np.zeros((N,N))
        for i in range(N):
            for j in range(N):
                M = self.teco.inertia(np.r_[0, Q2[i,j], Q3[i,j], 0, 0, 0])
                M11[i,j] = M[0,0]
                M12[i,j] = M[0,1]

        self.xlsx_outpath = "./xlsx/"
        self.pic_outpath = "./picture/"

    def cmd_callback(self,data):
        self.cmd = data.cmd
        rospy.loginfo("I heard command is %s", data.cmd)
    
    def dyna_callback(self, data):
        self.joint_angle = data.joint_angle[0:6]
        self.payload = data.payload
        self.payload_position = data.payload_position
        self.vel = data.vel[0:6]
        self.acc = data.acc[0:6]
        rospy.loginfo("I heard command is %s", self.joint_angle)
        rospy.loginfo("I heard command is %s", self.payload)
        rospy.loginfo("I heard command is %s", self.payload_position)
        rospy.loginfo("I heard command is %s", self.vel)
        rospy.loginfo("I heard command is %s", self.acc)

    def dyna_space_callback(self,data):
        self.payload_space = data.payload
        self.payload_position_space = data.payload_position
        self.axis_num = data.analysis_axis
        rospy.loginfo("I heard command is %s", data.payload)
        rospy.loginfo("I heard command is %s", data.payload_position)
        rospy.loginfo("I heard command is %s", data.analysis_axis)

    def planned_path_callback(self, data):
        self.model_id = data.model_id
        self.robot_trajectory = data.trajectory
        self.robotstate = data.trajectory_start

        points_num = len(self.robot_trajectory[0].joint_trajectory.points)
        positions = []
        velocities = []
        accelerations = []
        time_from_start = []
        time_from_start_secs = []
        time_from_start_nsecs = []
        time_array = []
        for i in range(points_num):
            positions.append(self.robot_trajectory[0].joint_trajectory.points[i].positions)
            velocities.append(self.robot_trajectory[0].joint_trajectory.points[i].velocities)
            accelerations.append(self.robot_trajectory[0].joint_trajectory.points[i].accelerations)
            time_from_start.append(self.robot_trajectory[0].joint_trajectory.points[i].time_from_start)
            time_from_start_secs.append(self.robot_trajectory[0].joint_trajectory.points[i].time_from_start.secs)
            time_from_start_nsecs.append(self.robot_trajectory[0].joint_trajectory.points[i].time_from_start.nsecs)

        self.trajectory_dynamics_calc(points_num, positions, velocities, accelerations, time_from_start)
        # TODO: plot trajectory torque 
        # rospy.loginfo("tau_j_array 1 is %s", self.tau_j_array[1,:].tolist())
        time_array_secs = np.array(time_from_start_secs)
        time_array_nsecs = np.array(time_from_start_nsecs)
        time_array = time_array_secs+time_array_nsecs*(10)**(-9)
        print(time_array)
        tau_array = np.array(self.tau_j_array)
        print(tau_array[:,1])

        x_smooth = np.linspace(time_array.min(),time_array.max(),300)

        plt.ion()
        fig = plt.figure()
        axis = 6
        for i in range(axis):
            y_smooth = make_interp_spline(time_array, tau_array[:,i])(x_smooth)
            plt.subplot(2,3,i+1) 
            plt.plot(x_smooth, y_smooth,"b--")
            string_axis = str(i+1)
            plt.title("torque "+string_axis, {'fontsize':10})  # 設定圖標題及其文字大小
            plt.xlabel("sec")
            plt.ylabel("N*m")

        plt.show()
        # plt.savefig(path.join(self.pic_outpath,"trajectory_torque.png"))
        plt.pause(0.0001)
        fig.canvas.draw()
 
    def payload_set(self):
        self.teco.payload(20, [0, 0, 0]) # set payload 

    def dynamics_cal(self):
        qd = np.r_[0, 1, 0, 0, 0, 0]
        # print("qd:",qd)
        self.teco.coriolis(self.teco.qn, qd) @ qd
        # TODO:  rne 逆動力學 add vel & acc analyses
        self.teco.rne(self.teco.qn, np.zeros((6,)), np.zeros((6,)))
        # 窮舉法正運動學計算工作空間
        start = time.time()
        print("The time used to execute this is given below")
        i = 0
        # 角度轉換
        du=pi/180;  #度
        radian=180/pi; #弧度

        fig = plt.figure()
        self.ax = plt.subplot(111, projection='3d')


        self.teco.payload(self.payload_space, self.payload_position_space) # set payload 

        for q1 in range(self.q1_s, self.q1_end, self.step):
            for q2 in range(self.q2_s, self.q2_end, self.step):
                percent = i/self.T_cell*100
                print("percent:{:.0f}%".format(percent), end="\r")
                for q3 in range(self.q3_s, self.q3_end, self.step):
                    for q4 in range(self.q4_s, self.q4_end, self.step):
                        for q5 in range(self.q5_s, self.q5_end, self.step):
                            self.T = self.teco.fkine([q1*du, q2*du, q3*du, q4*du, q5*du, 0*du])
                            load = np.array([self.teco.gravload([q1*du, q2*du, q3*du, q4*du, q5*du, 0*du])])
                            self.torque = np.append(self.torque,load,axis=0)
                            self.T_x[0,i] = self.T.t[0]
                            self.T_y[0,i] = self.T.t[1]
                            self.T_z[0,i] = self.T.t[2]
                            i=i+1
                            
        end = time.time()
        print("繪製工作空間運行時間：%f sec" % (end - start))

    def plot_space_scan(self):
        self.ax.scatter(self.T_x[0,:], self.T_y[0,:], self.T_z[0,:], c='r', marker='o')
        self.ax.set_xlabel("x")
        self.ax.set_ylabel("y") 
        self.ax.set_zlabel("z")
        plt.show()
        plt.pause(0)

    def sol_output(self):
        print(self.torque)
        print("torque type:",type(self.torque))
        print("torque size:",len(self.torque))

        print("軸1 torque正最大值:",np.max(self.torque[:,0]))
        print("軸2 torque正最大值:",np.max(self.torque[:,1]))
        print("軸3 torque正最大值:",np.max(self.torque[:,2]))
        print("軸4 torque正最大值:",np.max(self.torque[:,3]))
        print("軸5 torque正最大值:",np.max(self.torque[:,4]))
        print("軸6 torque正最大值:",np.max(self.torque[:,5]))

        print("軸1 torque負最大值:",np.min(self.torque[:,0]))
        print("軸2 torque負最大值:",np.min(self.torque[:,1]))
        print("軸3 torque負最大值:",np.min(self.torque[:,2]))
        print("軸4 torque負最大值:",np.min(self.torque[:,3]))
        print("軸5 torque負最大值:",np.min(self.torque[:,4]))
        print("軸6 torque負最大值:",np.min(self.torque[:,5]))

        print("軸2 torque正最大值時, 各軸torque, 末端位置, 各軸角度")
        torque_where = np.where(self.torque==np.max(self.torque[:,1]))
        for i in range(len(torque_where[0])):
            print(torque_where[0][i])
            max_torque = torque_where[0][i]
            print("torque:",self.torque[max_torque])
            print("末端位置",[self.T_x[0,i], self.T_y[0,i], self.T_z[0,i]])
            # TODO: 求解逆運動學 各軸角度self.teco.ikine_a()
            self.T.t[0] = self.T_x[0,i]
            self.T.t[1] = self.T_y[0,i]
            self.T.t[2] = self.T_z[0,i]
            sol = self.teco.ikine_LM(self.T)  # original sol = self.teco.ikine_a(self.T, "lun")
            print("sol:",sol)
            self.teco.plot(sol.q, dt=0.1 )
            # plt.show()

    def sol_output_axis(self):
        for i in range(6):
            axis = i 
            # excel output
            # 建立excel空白活頁簿
            excel_file = Workbook()
            # 建立一個工作表
            sheet = excel_file.active
            # 先填入第一列的欄位名稱
            sheet['A1'] = 'axis 1'
            sheet['B1'] = 'axis 2'
            sheet['C1'] = 'axis 3'
            sheet['D1'] = 'axis 4'
            sheet['E1'] = 'axis 5'
            sheet['F1'] = 'axis 6'
            sheet['G1'] = 'torque 1'
            sheet['H1'] = 'torque 2'
            sheet['I1'] = 'torque 3'
            sheet['J1'] = 'torque 4'
            sheet['K1'] = 'torque 5'
            sheet['L1'] = 'torque 6'


            self.ik_sol_positive= []
            print("軸%d torque正最大值:%f" %(axis+1, np.max(self.torque[:,axis])))
            print("軸%d torque負最大值:%f" %(axis+1, np.min(self.torque[:,axis])))
            print("軸%d torque正最大值時, 各軸torque, 末端位置, 各軸角度" %(axis+1))
            torque_where = np.where(self.torque==np.max(self.torque[:,axis]))
            for i in range(len(torque_where[0])):
                print(torque_where[0][i])
                max_torque = torque_where[0][i]
                print("torque:",self.torque[max_torque])
                print("末端位置",[self.T_x[0,i], self.T_y[0,i], self.T_z[0,i]])
                # TODO: 求解逆運動學 各軸角度self.teco.ikine_a()
                self.T.t[0] = self.T_x[0,i]
                self.T.t[1] = self.T_y[0,i]
                self.T.t[2] = self.T_z[0,i]
                sol = self.teco.ikine_LM(self.T)  # original sol = self.teco.ikine_a(self.T, "lun")
                print("sol:",sol)
                # TODO: 新增plot圖關閉功能, button close plot , 之後須增加需要自動關閉的情況
                # TODO: 匯入至CSV檔案
                # TODO:# output excel file
                append_sol_list_angle = sol.q.tolist()
                append_sol_list_torque = self.torque[max_torque].tolist()
                append_sol_list.extend(append_sol_list_torque)
                sheet.append(append_sol_list)

            file_name = self.xlsx_outpath+'/dynamics_space_calc_axis'+str(axis+1)+'_positive'+'.xlsx'
            excel_file.save(file_name)

            # excel output
            # 建立excel空白活頁簿
            excel_file = Workbook()
            # 建立一個工作表
            sheet = excel_file.active
            # 先填入第一列的欄位名稱
            sheet['A1'] = 'axis 1'
            sheet['B1'] = 'axis 2'
            sheet['C1'] = 'axis 3'
            sheet['D1'] = 'axis 4'
            sheet['E1'] = 'axis 5'
            sheet['F1'] = 'axis 6'
            sheet['G1'] = 'torque 1'
            sheet['H1'] = 'torque 2'
            sheet['I1'] = 'torque 3'
            sheet['J1'] = 'torque 4'
            sheet['K1'] = 'torque 5'
            sheet['L1'] = 'torque 6'

            self.ik_sol_negative= []
            print("軸%d torque負最大值時, 各軸torque, 末端位置, 各軸角度" %(axis+1))
            torque_where = np.where(self.torque==np.min(self.torque[:,axis]))
            for i in range(len(torque_where[0])):
                print(torque_where[0][i])
                max_torque = torque_where[0][i]
                print("torque:",self.torque[max_torque])
                print("末端位置",[self.T_x[0,i], self.T_y[0,i], self.T_z[0,i]])
                # TODO: 求解逆運動學 各軸角度self.teco.ikine_a()
                self.T.t[0] = self.T_x[0,i]
                self.T.t[1] = self.T_y[0,i]
                self.T.t[2] = self.T_z[0,i]
                sol = self.teco.ikine_LM(self.T)  # original sol = self.teco.ikine_a(self.T, "lun")
                print("sol:",sol)
                # TODO: 新增plot圖關閉功能, button close plot , 之後須增加需要自動關閉的情況
                # TODO: 匯入至CSV檔案
                # TODO:# output excel file
                append_sol_list_angle = sol.q.tolist()
                append_sol_list_torque = self.torque[max_torque].tolist()
                append_sol_list.extend(append_sol_list_torque)
                sheet.append(append_sol_list)

            file_name = self.xlsx_outpath+'/dynamics_space_calc_axis'+str(axis+1)+'_negative'+'.xlsx'
            excel_file.save(file_name)

    def dynamics_calc(self):
        qn = [0,0,0,0,0,0]
        deg = pi/180
        
        for i in range(6):
            qn[i] = self.joint_angle[i]*deg
        self.teco.payload(self.payload , self.payload_position) # set payload
        self.qn = qn
        self.tau_j = self.teco.rne(self.qn, self.vel, self.acc)
        print("tau_j:", self.tau_j)
        axis = 2
        self.teco.plot(self.qn)
        plt.savefig(path.join(self.pic_outpath,"dataname_dynamics_calc.png"))
        plt.close()
        # excel output
        # 建立excel空白活頁簿
        excel_file = Workbook()
        # 建立一個工作表
        sheet = excel_file.active
        # 先填入第一列的欄位名稱
        sheet['A1'] = 'axis 1'
        sheet['B1'] = 'axis 2'
        sheet['C1'] = 'axis 3'
        sheet['D1'] = 'axis 4'
        sheet['E1'] = 'axis 5'
        sheet['F1'] = 'axis 6'
        sheet.append([self.tau_j[0],self.tau_j[1],self.tau_j[2],self.tau_j[3],self.tau_j[4],self.tau_j[5]])
        file_name = self.xlsx_outpath+'/dynamics_calc'+'.xlsx'
        # excel_file.save('dynamics_calc.xlsx')
        excel_file.save(file_name)

    def trajectory_dynamics_calc(self, num, pos, vel, acc, time):
        self.teco.payload(self.payload , self.payload_position) # set payload
        self.tau_j_array = []
        for i in range(num):
            self.tau_j = self.teco.rne(pos[i], vel[i], acc[i])
            print("tau_j array:", self.tau_j)
            self.tau_j_array.append(self.tau_j)
        print("tau_j array size:", len(self.tau_j_array))

    def arm_plot(self):

        qn = [0,0,0,0,0,0]
        deg = pi/180
        
        for i in range(6):
            qn[i] = self.joint_angle[i]*deg

        self.qn = qn
        print("arm_plot:", self.qn)
        self.teco.plot(self.qn, backend='pyplot', block=False, vellipse=False, fellipse=False)
        plt.show()
        
    def plot_close(self):
        plt.close()

    def task_set(self):
        for case in switch(self.cmd):
            if case(1):
                print("Start Workspace Scan")
                print("success get subscriber data ")
                self.payload_set()
                self.dynamics_cal()
                self.sol_output_axis()
                self.plot_space_scan()
                self.cmd = 0
                break

            if case(2):
                print("Start Set payload & vel & acc analysis")
                self.dynamics_calc()
                self.cmd = 0
                break
            # Select axis for dynamics space scan joint torque output
            if case(3):
                print("Select axis for dynamics space scan joint torque output")
                self.sol_output_axis()
                self.cmd = 0
                break
            
            if case(4):
                self.arm_plot()
                self.cmd = 0
                break
            #
            if case(5):
                self.plot_close()
                self.cmd = 0
                break

            if case():
                break


if __name__=="__main__":
    rospy.init_node("dynamics_teco")
    
    Dya = Dynamics_teco()
    while not rospy.is_shutdown():
        Dya.task_set()