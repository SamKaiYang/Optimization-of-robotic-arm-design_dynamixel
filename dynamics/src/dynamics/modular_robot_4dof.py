#!/usr/bin/env python3
# coding: utf-8
import numpy as np
from roboticstoolbox import DHRobot, RevoluteDH
from spatialmath import SE3
import spatialmath as sm
from urdf_parser_py.urdf import URDF
import os
# import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
class modular_robot_4dof(DHRobot):
    """
    Class that models a TECO TECOARM1 manipulator

    :param symbolic: use symbolic constants
    :type symbolic: bool

    describes its kinematic and dynamic characteristics using standard DH
    conventions.

    .. runblock:: pycon

        >>> import roboticstoolbox as rtb
        >>> robot = rtb.models.DH.TECOARM1()
        >>> print(robot)

    Defined joint configurations are:

    - qz, zero joint angle configuration
    - qr, arm horizontal along x-axis

    .. note::
        - SI units are used.

    :References:

    .. codeauthor:: Yi He Yang
    """  # noqa

    def __init__(self, symbolic=False):
        #隨機抽樣點位初始化
        self.T_x = []
        self.T_y = []
        self.T_z = []
        self.T_roll = []
        self.T_pitch = []
        self.T_yaw = []
        self.xlsx_outpath = "./xlsx/"


        self.length_1 = 0.1
        self.length_2 = 0.1
        self.length_3 = 0.1

        # robot = URDF.from_xml_file(os.path.dirname(os.path.realpath(__file__))+"/urdf"+"/modular_robot_6dof.urdf")
        # robot = URDF.from_xml_file(os.path.dirname(os.path.realpath(__file__))+"/urdf"+"/random.urdf")
        robot = URDF.from_xml_file(os.path.dirname(os.path.realpath(__file__))+"/urdf"+"/single_arm_v12.urdf")
        # self.robot = robot
        if symbolic:
            import spatialmath.base.symbolic as sym
            zero = sym.zero()
            pi = sym.pi()
        else:
            from math import pi
            zero = 0.0

        deg = pi / 180
        inch = 0.0254

        self.length_1 = robot.joints[3].origin.xyz[1]
        self.length_2 = robot.joints[4].origin.xyz[1]
        self.length_3 = robot.joints[5].origin.xyz[2]

        # robot length values (metres)
        # a = [0, -0.314, -0.284, 0, 0, 0] #m # teco
        a = [0, robot.joints[3].origin.xyz[0], robot.joints[4].origin.xyz[0], 0, 0, 0]
        # a = [0, -j3.y, -j4.y, 0, 0, 0]
        # d = [0.1301, 0, 0, 0.1145, 0.090, 0.048] #m # teco 
        d = [robot.joints[1].origin.xyz[2]+robot.joints[2].origin.xyz[2],
            # 0.068,
            0,
            0,
            # robot.joints[2].origin.xyz[1]-robot.joints[4].origin.xyz[2], 
            robot.joints[5].origin.xyz[1] + robot.joints[5].origin.xyz[2],  
            robot.joints[5].origin.xyz[1] + robot.joints[5].origin.xyz[2],
            robot.joints[6].origin.xyz[1]
        ]
        # d = [j1.z, 0, 0, j2.y-j4.z , j5.z, j6.z] #m
        alpha = [pi/2, 0, zero, pi/2, -pi/2, zero]

        
        # mass data, no inertia available
        # mass = [3.7000, 8.3930, 2.33, 1.2190, 1.2190, 0.1897]
        mass = [robot.links[2].inertial.mass, 
                robot.links[3].inertial.mass,
                robot.links[4].inertial.mass,
                robot.links[5].inertial.mass,
                robot.links[6].inertial.mass,
                robot.links[7].inertial.mass
            ]
        
        G= [-80,-80,-80,-50,-50,-50]   # gear ratio
        G= [-1,-1,-1,-1,-1,-1]   # gear ratio
        # B = [[10,10], [10,10], [10,10], [10,10], [10,10], [10,10]]
        B = 10.0
        # center_of_mass = [
        #         [-1.55579201081481E-05, 0.00265005484815443, -0.00640979059142413],
        #         [4.90637956589368E-11, 0.205571973027702, -0.003359898058563426],
        #         [-9.75479428030767E-05, 0.271025707847572, 0.111573843205116],
        #         [-0.000181761828397664, 0.00219045749084071, -0.000800397394362884],
        #         [-0.000192919655058627, -0.00232492307126431, 0.00352418959262345],
        #         [-4.4856E-13 ,0 ,0.025]
        #     ]

        center_of_mass = [
                robot.links[2].inertial.origin.xyz,
                robot.links[3].inertial.origin.xyz,
                robot.links[4].inertial.origin.xyz,
                robot.links[5].inertial.origin.xyz,
                robot.links[6].inertial.origin.xyz,
                robot.links[7].inertial.origin.xyz
            ]
        # signed from a 3 3 matrix or a 6-element vector interpretted as Ixx Iyy Izz Ixy Iyz Ixz
        inertia = [
                [robot.links[2].inertial.inertia.ixx, robot.links[2].inertial.inertia.iyy, robot.links[2].inertial.inertia.izz, robot.links[2].inertial.inertia.ixy,robot.links[2].inertial.inertia.iyz, robot.links[2].inertial.inertia.ixz],
                [robot.links[3].inertial.inertia.ixx, robot.links[3].inertial.inertia.iyy, robot.links[3].inertial.inertia.izz, robot.links[3].inertial.inertia.ixy,robot.links[3].inertial.inertia.iyz, robot.links[3].inertial.inertia.ixz],
                [robot.links[4].inertial.inertia.ixx, robot.links[4].inertial.inertia.iyy, robot.links[4].inertial.inertia.izz, robot.links[4].inertial.inertia.ixy,robot.links[4].inertial.inertia.iyz, robot.links[4].inertial.inertia.ixz],
                [robot.links[5].inertial.inertia.ixx, robot.links[5].inertial.inertia.iyy, robot.links[5].inertial.inertia.izz, robot.links[5].inertial.inertia.ixy,robot.links[5].inertial.inertia.iyz, robot.links[5].inertial.inertia.ixz],
                [robot.links[6].inertial.inertia.ixx, robot.links[6].inertial.inertia.iyy, robot.links[6].inertial.inertia.izz, robot.links[6].inertial.inertia.ixy,robot.links[6].inertial.inertia.iyz, robot.links[6].inertial.inertia.ixz],
                [robot.links[7].inertial.inertia.ixx, robot.links[7].inertial.inertia.iyy, robot.links[7].inertial.inertia.izz, robot.links[7].inertial.inertia.ixy,robot.links[7].inertial.inertia.iyz, robot.links[7].inertial.inertia.ixz]
            ]
        # qlim = [
        #     [-pi,pi],
        #     [-pi,pi],
        #     [-pi,pi],
        #     [-pi,pi],
        #     [-pi,pi],
        #     [-pi,pi]
        # ]
        # qlim = [
        #     [-pi,pi],
        #     [-pi*5/36,pi*43/36], # [-25~215]
        #     [-pi*8/9,pi*8/9], # [-160~160]
        #     [-pi,pi],
        #     [-pi,pi],
        #     [-pi,pi]
        # ]
        links = []

        d[4] = 0
        d[5] = 0
        a[4] = 0
        a[5] = 0
        alpha[4] = 0
        alpha[5] = 0
        mass[4] = 0
        mass[5] = 0
        # center_of_mass[4] = 0
        # center_of_mass[5] = 0
        # inertia[4] = [0,0,0,0,0,0]
        # inertia[5] = [0,0,0,0,0,0]
        # G[4] = 0
        # G[5] = 0
        for j in range(6): # TODO: for 4 DoF
            link = RevoluteDH(
                d=d[j],
                a=a[j],
                alpha=alpha[j],
                m=mass[j],
                r=center_of_mass[j],
                I=inertia[j],
                G=G[j]
                # B=B[j]
                # qlim=qlim[j]
            )
            links.append(link)
    
        super().__init__(
            links,
            name="robot",
            manufacturer="Robotics",
            keywords=('dynamics', 'symbolic'),
            symbolic=symbolic
        )
        # print("FFFFUCCCUCUCUCUCCUCUCCKKKKK")
        # zero angles
        self.addconfiguration("qz", np.array([0, 0, 0, 0, 0, 0]))
        # horizontal along the x-axis
        self.addconfiguration("qr", np.r_[180, 0, 0, 0, 0, 0]*deg)

        # nominal table top picking pose
        self.addconfiguration("qn", np.array([0, 0, 0, 0, 0, 0]))
        
    def return_configuration(self):
        return self.length_1, self.length_2, self.length_3


    def point_Workspace_cal_Monte_Carlo(self, robot):
        """
        Through the "Work space" page in the interface to calculate of the robot
        """
        i = 0
        # 角度轉換
        du = pi / 180
        # 度
        radian = 180 / pi
        # 弧度

        self.q1_s = -160
        self.q1_end = 160
        self.q2_s = -160
        self.q2_end = 160
        self.q3_s = -160
        self.q3_end = 160
        self.q4_s = -160
        self.q4_end = 160
        self.q5_s = -160
        self.q5_end = 160
        self.q6_s = -160
        self.q6_end = 160
        N = 1000
        theta1 = self.q1_end + (self.q1_end - self.q1_s) * np.random.rand(N, 1)
        theta2 = self.q2_end + (self.q2_end - self.q2_s) * np.random.rand(N, 1)
        theta3 = self.q3_end + (self.q3_end - self.q3_s) * np.random.rand(N, 1)
        theta4 = self.q4_end + (self.q4_end - self.q4_s) * np.random.rand(N, 1)
        theta5 = self.q5_end + (self.q5_end - self.q5_s) * np.random.rand(N, 1)
        theta6 = self.q6_end + (self.q6_end - self.q6_s) * np.random.rand(N, 1)

        for i in range(N):
            q1 = theta1[i, 0]
            q2 = theta2[i, 0]
            q3 = theta3[i, 0]
            q4 = theta4[i, 0]
            q5 = theta5[i, 0]
            q6 = theta6[i, 0]
            self.T = robot.fkine(
                [q1 * du, q2 * du, q3 * du, q4 * du, 0, 0]
            )
            t = np.round(self.T.t, 3)
            r = np.round(self.T.rpy('deg'), 3)
            self.T_x.append(t[0])
            self.T_y.append(t[1])
            self.T_z.append(t[2])
            self.T_roll.append(int(r[0]))
            self.T_pitch.append(int(r[1]))
            self.T_yaw.append(int(r[2]))
            i = i + 1
        
    def random_select_point(self):
        excel_file = Workbook()
        sheet = excel_file.active

        for i in range(10):
            x = np.random.randint(1,999)
            sheet.cell(row=i + 1, column=1).value = self.T_x[x]
            sheet.cell(row=i + 1, column=2).value = self.T_y[x]
            sheet.cell(row=i + 1, column=3).value = self.T_z[x]
            sheet.cell(row=i + 1, column=4).value = self.T_roll[x]
            sheet.cell(row=i + 1, column=5).value = self.T_pitch[x]
            sheet.cell(row=i + 1, column=6).value = self.T_yaw[x]

        file_name = self.xlsx_outpath + "/task_point_4dof" +".xlsx"
        excel_file.save(file_name)
if __name__ == '__main__':    # pragma nocover

    robot = modular_robot_4dof(symbolic=False)
    print(robot)
    # print(robot.dynamics())
    symbolic = False
    if symbolic:
        import spatialmath.base.symbolic as sym
        zero = sym.zero()
        pi = sym.pi()
    else:
        from math import pi
        zero = 0.0
    # Set the desired end-effector pose
    deg = pi / 180
    q =  np.r_[0, 0, 0]*deg
    robot.point_Workspace_cal_Monte_Carlo(robot)
    robot.random_select_point()
    robot.teach()
    # print(robot.fkine_path(q) * sm.SE3(0, 0, 0.04))

    # T = robot.fkine(q)
    # t = np.round(T.t, 3)
    # r = np.round(T.rpy(), 3)
    '''# print(robot.q)
 
    # robot.ikine_6s
    # robot.ikine_global
    # robot.ikine_LMS
    # robot.ikine_mmc
    T = SE3(0.7, 0.2, 0.1) * SE3.RPY([0, 1, 0])
    print(T)
    print(robot.ikine_LM(T=T))
    print(robot.ikine_LMS(T=T))

    print(robot.ikine_LM(T=robot.fkine(q) * sm.SE3(0, 0, 0.04)))
    print(robot.ikine_LMS(T=robot.fkine(q) * sm.SE3(0, 0, 0.04)))
    
    print(robot.manipulability(q=q))
    # print(robot.manipulability(J=robot.fkine(q) * sm.SE3(0, 0, 0.04)))
    
'''


    df = load_workbook("./xlsx/task_point_4dof.xlsx")
    sheets = df.worksheets
    sheet1 = sheets[0]
    rows = sheet1.rows
    cols = sheet1.columns

    T_tmp = []
    manipulability_index = []
    i = 0
    for row in rows:
        row_val = [col.value for col in row]
        T_tmp.append(SE3(row_val[0], row_val[1], row_val[2]) * SE3.RPY([np.deg2rad(row_val[3]), np.deg2rad(row_val[4]), np.deg2rad(row_val[5])]))
        # print(T_tmp[i])
        ik_q = robot.ikine_LMS(T=T_tmp[i], ilimit = 500, tol = 1e-10, mask = [1,1,1,1,1,1])
        print(ik_q)
        print(robot.fkine(ik_q.q))
    #     if ik_q.success == True:
    #         manipulability_index.append(robot.manipulability(q=ik_q.q))
    #         # robot.plot_ellipse()
    #         # ik_np = np.array(ik_q.q)
    #         # print(ik_np)
    #         # robot.plot(q=ik_np, backend='pyplot', dt = 1)
        i = i + 1
    # print(np.mean(manipulability_index)) # manipulability 取平均
    