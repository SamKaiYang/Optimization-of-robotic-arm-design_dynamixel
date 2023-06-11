#!/usr/bin/env python3
# coding: utf-8
import numpy as np
from roboticstoolbox import DHRobot, RevoluteDH
from spatialmath import SE3
import spatialmath as sm
from urdf_parser_py.urdf import URDF
import os
# import pandas as pd
from openpyxl import load_workbook

class modular_robot_6dof(DHRobot):
    """
    Class that models a TECO TECOARM1 manipulator

    :param symbolic: use symbolic constants
    :type symbolic: bool

    describes its kinematic and dynamic characteristics using standard DH
    conventions.

    .. runblock:: pycon

        >>> import roboticstoolbox as rtb
        >>> robot = rtb.models.DH.TECOARM1()
        >>> print(robot)

    Defined joint configurations are:

    - qz, zero joint angle configuration
    - qr, arm horizontal along x-axis

    .. note::
        - SI units are used.

    :References:

    .. codeauthor:: Yi He Yang
    """  # noqa

    def __init__(self, symbolic=False):
        self.length_1 = 0.1
        self.length_2 = 0.1
        self.length_3 = 0.1

        # robot = URDF.from_xml_file(os.path.dirname(os.path.realpath(__file__))+"/urdf"+"/modular_robot_6dof.urdf")
        # robot = URDF.from_xml_file(os.path.dirname(os.path.realpath(__file__))+"/urdf"+"/random.urdf")
        robot = URDF.from_xml_file(os.path.dirname(os.path.realpath(__file__))+"/urdf"+"/single_arm_v22_19cm.urdf")
        if symbolic:
            import spatialmath.base.symbolic as sym
            zero = sym.zero()
            pi = sym.pi()
        else:
            from math import pi
            zero = 0.0

        deg = pi / 180
        inch = 0.0254

        self.length_1 = robot.joints[3].origin.xyz[1]
        self.length_2 = robot.joints[4].origin.xyz[1]
        self.length_3 = robot.joints[5].origin.xyz[2]

        # robot length values (metres)
        # a = [0, -0.314, -0.284, 0, 0, 0] #m # teco
        a = [0, robot.joints[3].origin.xyz[0], robot.joints[4].origin.xyz[0], 0, 0, 0]
        # a = [0, -j3.y, -j4.y, 0, 0, 0]
        # d = [0.1301, 0, 0, 0.1145, 0.090, 0.048] #m # teco 
        d = [robot.joints[1].origin.xyz[2]+robot.joints[2].origin.xyz[2],
            # 0.068,
            0,
            0,
            # robot.joints[2].origin.xyz[1]-robot.joints[4].origin.xyz[2], 
            robot.joints[5].origin.xyz[1] + robot.joints[5].origin.xyz[2],  
            robot.joints[5].origin.xyz[1] + robot.joints[5].origin.xyz[2],
            robot.joints[6].origin.xyz[1]
        ]
        # d = [j1.z, 0, 0, j2.y-j4.z , j5.z, j6.z] #m
        alpha = [pi/2, 0, zero, pi/2, -pi/2, zero]

        
        # mass data, no inertia available
        # mass = [3.7000, 8.3930, 2.33, 1.2190, 1.2190, 0.1897]
        mass = [robot.links[2].inertial.mass, 
                robot.links[3].inertial.mass,
                robot.links[4].inertial.mass,
                robot.links[5].inertial.mass,
                robot.links[6].inertial.mass,
                robot.links[7].inertial.mass
            ]
        
        G= [-80,-80,-80,-50,-50,-50]   # gear ratio
        G= [-1,-1,-1,-1,-1,-1]   # gear ratio
        # B = [[10,10], [10,10], [10,10], [10,10], [10,10], [10,10]]
        B = 10.0
        # center_of_mass = [
        #         [-1.55579201081481E-05, 0.00265005484815443, -0.00640979059142413],
        #         [4.90637956589368E-11, 0.205571973027702, -0.003359898058563426],
        #         [-9.75479428030767E-05, 0.271025707847572, 0.111573843205116],
        #         [-0.000181761828397664, 0.00219045749084071, -0.000800397394362884],
        #         [-0.000192919655058627, -0.00232492307126431, 0.00352418959262345],
        #         [-4.4856E-13 ,0 ,0.025]
        #     ]

        center_of_mass = [
                robot.links[2].inertial.origin.xyz,
                robot.links[3].inertial.origin.xyz,
                robot.links[4].inertial.origin.xyz,
                robot.links[5].inertial.origin.xyz,
                robot.links[6].inertial.origin.xyz,
                robot.links[7].inertial.origin.xyz
            ]
        # signed from a 3 3 matrix or a 6-element vector interpretted as Ixx Iyy Izz Ixy Iyz Ixz
        inertia = [
                [robot.links[2].inertial.inertia.ixx, robot.links[2].inertial.inertia.iyy, robot.links[2].inertial.inertia.izz, robot.links[2].inertial.inertia.ixy,robot.links[2].inertial.inertia.iyz, robot.links[2].inertial.inertia.ixz],
                [robot.links[3].inertial.inertia.ixx, robot.links[3].inertial.inertia.iyy, robot.links[3].inertial.inertia.izz, robot.links[3].inertial.inertia.ixy,robot.links[3].inertial.inertia.iyz, robot.links[3].inertial.inertia.ixz],
                [robot.links[4].inertial.inertia.ixx, robot.links[4].inertial.inertia.iyy, robot.links[4].inertial.inertia.izz, robot.links[4].inertial.inertia.ixy,robot.links[4].inertial.inertia.iyz, robot.links[4].inertial.inertia.ixz],
                [robot.links[5].inertial.inertia.ixx, robot.links[5].inertial.inertia.iyy, robot.links[5].inertial.inertia.izz, robot.links[5].inertial.inertia.ixy,robot.links[5].inertial.inertia.iyz, robot.links[5].inertial.inertia.ixz],
                [robot.links[6].inertial.inertia.ixx, robot.links[6].inertial.inertia.iyy, robot.links[6].inertial.inertia.izz, robot.links[6].inertial.inertia.ixy,robot.links[6].inertial.inertia.iyz, robot.links[6].inertial.inertia.ixz],
                [robot.links[7].inertial.inertia.ixx, robot.links[7].inertial.inertia.iyy, robot.links[7].inertial.inertia.izz, robot.links[7].inertial.inertia.ixy,robot.links[7].inertial.inertia.iyz, robot.links[7].inertial.inertia.ixz]
            ]
        # rad = deg/180*pi
        q_lim = [
            [-pi,pi],
            [0,pi],
            [-150/180*pi,150/180*pi],
            [-pi,pi],
            [-pi,pi],
            [-pi,pi]
        ]
        # q_lim = [
        #     [-pi,pi],
        #     [-150/180*pi,150/180*pi],
        #     [-150/180*pi,150/180*pi],
        #     [-150/180*pi,150/180*pi],
        #     [-150/180*pi,150/180*pi],
        #     [-pi,pi]
        # ]
        # qlim = [
        #     [-pi,pi],
        #     [-pi*5/36,pi*43/36], # [-25~215]
        #     [-pi*8/9,pi*8/9], # [-160~160]
        #     [-pi,pi],
        #     [-pi,pi],
        #     [-pi,pi]
        # ]
        links = []

        for j in range(6):
            link = RevoluteDH(
                d=d[j],
                a=a[j],
                alpha=alpha[j],
                m=mass[j],
                r=center_of_mass[j],
                I=inertia[j],
                G=G[j],
                # B=B[j]
                qlim=q_lim[j]
            )
            links.append(link)
    
        super().__init__(
            links,
            name="robot",
            manufacturer="Robotics",
            keywords=('dynamics', 'symbolic'),
            symbolic=symbolic
        )
        # print("FFFFUCCCUCUCUCUCCUCUCCKKKKK")
        # zero angles
        self.addconfiguration("qz", np.array([0, 0, 0, 0, 0, 0]))
        # horizontal along the x-axis
        self.addconfiguration("qr", np.r_[180, 0, 0, 0, 90, 0]*deg)

        # nominal table top picking pose
        self.addconfiguration("qn", np.array([0, 0, 0, 0, 0, 0]))
        
    def return_configuration(self):
        return self.length_1, self.length_2, self.length_3
if __name__ == '__main__':    # pragma nocover
    
    robot = modular_robot_6dof(symbolic=False)
    print(robot)
    # print(robot.dynamics())
    symbolic = False
    if symbolic:
        import spatialmath.base.symbolic as sym
        zero = sym.zero()
        pi = sym.pi()
    else:
        from math import pi
        zero = 0.0
    # Set the desired end-effector pose
    deg = pi / 180
    q =  np.r_[0, 0, 0, 0, 0, 0]*deg
    q_test =  np.r_[0, 90, -0, 0, 0, 0]*deg

    robot.teach()
    # df = load_workbook("./xlsx/task_point_6dof_tested_ori_random.xlsx")
    # sheets = df.worksheets
    # sheet1 = sheets[0]
    # rows = sheet1.rows
    # cols = sheet1.columns
    # T_tmp = []
    # T_traj = []
    # ik_q_traj = []
    # score = []
    # ratio_over = 0
    # torque_over = 0
    # num_torque = np.array([np.zeros(shape=6)])
    # total_time = 20
    # # 采样间隔
    # sample_interval = 0.2
    # manipulability_index = []
    # i = 0
    # false_done = False
    # count = 0
    # max_diff = []
    # max_diff_tol = 0
    # traj_time = []
    # diff = np.array([np.zeros(shape=6)])
    # for row in rows:
    #     row_val = [col.value for col in row]
    #     T_tmp.append(SE3(row_val[0], row_val[1], row_val[2]) * SE3.RPY([np.deg2rad(row_val[3]), np.deg2rad(row_val[4]), np.deg2rad(row_val[5])]))
    #     ik_q = robot.ikine_LMS(T=T_tmp[i])

    #     if ik_q.success == True:
    #         count += 1
    #         T_traj.append(T_tmp[i])
    #         ik_q_traj.append(ik_q.q)
    #         manipulability_index.append(robot.manipulability(q=ik_q.q))
    #         print("ik_q.q",ik_q.q)
    #         if count >= 2: # 兩個點位以上開始計算
    #             diff = diff + np.abs(np.subtract(ik_q_traj[-2], ik_q_traj[-1]))
    #             max_diff_tol = max_diff_tol + np.max(diff)
    #             max_diff.append(np.max(diff))
    #             print(max_diff)
    #     i = i + 1
    # for k in range(len(max_diff)):
    #     traj_time.append(max_diff[k] / max_diff_tol * total_time)
    #     print("traj_time",traj_time)
    # for m in range(len(max_diff)):
    #     time_vector = np.linspace(0, traj_time[m], int(traj_time[m]/sample_interval) + 1)
    #     traj = robot.jtraj(T_traj[m],T_traj[m+1],time_vector)
    #     print(traj.s)
    #     if np.amax(traj.sd) > 3.04:
    #         ratio_over = ratio_over + 1
    #     torque = robot.rne(traj.s,traj.sd,traj.sdd)
    #     row = abs(torque[:,1]) # 取出第2行
    #     result_2 = row[row > 44.7] # 取出大于阈值的数字
    #     row = abs(torque[:,2]) # 取出第3行
    #     result_3 = row[row > 44.7] # 取出大于阈值的数字
    #     if len(result_2)>0 or len(result_3) >0:
    #         torque_over = torque_over + 1
    #     num_torque = np.append(num_torque, torque)
    # total_energy = 0
    # for j in range(len(num_torque)):
    #     energy = abs(num_torque[j]) * sample_interval
    #     total_energy += energy
            
        # if count == 0:
        #     return(0, 0, 0,0,0)
        # else:
        #     final_score = count / i
            # if count == 1:
            #     return(0, 0, 0, final_score, manipulability_index[0]) # 回傳 manipulability[0]
            # else:
            #     return(ratio_over, torque_over, total_energy, final_score, np.mean(manipulability_index)) # 回傳 manipulability 取平均


    # robot.plot(q=q_test, backend='pyplot', dt = 10)
    
    # robot.teach()
    # manipulability teach view
    # robot.teach(limits= [-0.5, 0.5, -0.5, 0.5, -0, 1],vellipse=True)
    
    '''
    # manipulability test point
    q0 =  np.r_[0,90,0,23,90,0]*deg
    q0 =  np.r_[0,76,-60,66,90,0]*deg
    vellipse = robot.vellipse(q=q0)
    robot.plot_ellipse(ellipse = vellipse)
    print(robot.manipulability(q=q0))
    '''

    # print(robot.fkine_path(q) * sm.SE3(0, 0, 0.04))

    # T = robot.fkine(q)
    # t = np.round(T.t, 3)
    # r = np.round(T.rpy(), 3)
    '''# print(robot.q)
 
    # robot.ikine_6s
    # robot.ikine_global
    # robot.ikine_LMS
    # robot.ikine_mmc
    T = SE3(0.7, 0.2, 0.1) * SE3.RPY([0, 1, 0])
    print(T)
    print(robot.ikine_LM(T=T))
    print(robot.ikine_LMS(T=T))

    print(robot.ikine_LM(T=robot.fkine(q) * sm.SE3(0, 0, 0.04)))
    print(robot.ikine_LMS(T=robot.fkine(q) * sm.SE3(0, 0, 0.04)))
    
    print(robot.manipulability(q=q))
    # print(robot.manipulability(J=robot.fkine(q) * sm.SE3(0, 0, 0.04)))
    
'''

'''
    T_tmp = []
    T_tmp.append(SE3(0.25, 0.113, 0.199) * SE3.RPY([np.deg2rad(-173), np.deg2rad(-59), np.deg2rad(-147)]))
    T_tmp.append(SE3(-0.06, -0.09, 0.20) * SE3.RPY([np.deg2rad(-79), np.deg2rad(27), np.deg2rad(-99)]))

    print(robot.ikine_LM(T=T_tmp[0]))
    print(robot.ikine_LMS(T=T_tmp[1]))
'''

'''
    robot.teach(limits= [-1, 1, -1, 1, -1, 1],vellipse=True)

    
    # import xlsx
    df = load_workbook("./xlsx/task_point_6dof.xlsx")
    sheets = df.worksheets
    sheet1 = sheets[0]
    rows = sheet1.rows
    cols = sheet1.columns
    robot.payload( 1.5, [0,0,0.04])  # set payload
    
    T_tmp = []
    num_torque = np.array([np.zeros(shape=6)])
    manipulability_index = []
    i = 0
    for row in rows:
        row_val = [col.value for col in row]
        T_tmp.append(SE3(row_val[0], row_val[1], row_val[2]) * SE3.RPY([np.deg2rad(row_val[3]), np.deg2rad(row_val[4]), np.deg2rad(row_val[5])]))
        # print(T_tmp[i])
        if i >= 1:
            # 完成动作的总时间
            total_time = 20
            # 采样间隔
            sample_interval = 0.056
            # 生成时间向量
            traj_time = total_time/10
            time_vector = np.linspace(0, traj_time, int(traj_time/sample_interval) + 1)
            # t=[0:0.056:2]
            traj = robot.jtraj(T_tmp[i-1],T_tmp[i],time_vector)
            # print(traj.s)
            # print(traj.sd)
            # # print(traj.sdd)
            # print(traj.t)
            # robot.plot(traj.s)
            # np.append(torque, load, axis=0)
            # print(traj.sd)
            # print(traj.sd[0])
            if np.amax(traj.sd) > 3.04:
                print("Fail")
            torque = robot.rne(traj.s,traj.sd,traj.sdd)
            if np.amax(torque) > 44.7:
                print(torque)
            num_torque = np.append(num_torque, torque)
            
        ik_q = robot.ikine_LMS(T=T_tmp[i])
        # print(ik_q)
        ik_np = np.array(ik_q.q)
        # robot.plot(q=ik_np, backend='pyplot', dt = 1)
        # if ik_q.success == True:
        #     manipulability_index.append(robot.manipulability(q=ik_q.q))
            # robot.plot_ellipse()
            # ik_np = np.array(ik_q.q)
            # print(ik_np)
            # robot.plot(q=ik_np, backend='pyplot', dt = 1)
        i = i + 1
    # print(np.mean(manipulability_index)) # manipulability 取平均
    time_interval = 0.056 # in seconds
    time_total = 20 # in seconds
    num_samples = int(time_total / time_interval)

    # Example motor power data
    motor_power = num_torque

    total_energy = 0
    for i in range(num_samples):
        energy = motor_power[i] * time_interval
        total_energy += energy
    print(total_energy)

    import numpy as np

    # num_actions = 12
    # one_hot_array = np.eye(num_actions)
    # print(one_hot_array)
    # action = 5
    # one_hot_vector = one_hot_array[action, :]
    # action = 8
    # one_hot_vector = one_hot_array[action, :]
    # print(one_hot_vector)

    matrix = np.array([[1, 2, 3], [4, 5, 6], [7, 8, 9]])
    # threshold = 5
    # result = np.where(matrix > threshold, axis=0)
    # print(result)
    # row_max = np.amax(matrix, axis=1)
    # print(row_max)

    # if np.amax(matrix, axis=1) > 5:
    #     print("false")
    m = 3 # 行数，从0开始算
    threshold = 9 # 阈值

    row = matrix[:,m-1] # 取出第m-1行
    result = row[row > threshold] # 取出大于阈值的数字
    print(len(result))
'''