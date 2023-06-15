#!/usr/bin/env python3
# coding: utf-8
import rospy
import importlib
import sys
importlib.reload(sys)
from openpyxl import Workbook
import numpy as np
from roboticstoolbox import DHRobot, RevoluteDH
from spatialmath import SE3
import spatialmath as sm
from urdf_parser_py.urdf import URDF
import os
# import pandas as pd
from openpyxl import load_workbook
from stl_conv_6dof_urdf_dynamixel_real import stl_conv_urdf
from modular_robot_6dof_real import modular_robot_6dof
curr_path = os.path.dirname(os.path.abspath(__file__))  # 当前文件所在绝对路径
import spatialmath.base.symbolic as sym
zero = sym.zero()
pi = sym.pi()
deg = pi / 180
# TODO: 初版 只考慮 6 dof 機器人的關節長度變化, 觀察各軸馬達極限之輸出最大torque值
class RobotTraj():
    def __init__(self):
        self.robot = modular_robot_6dof()
        self.robot_urdf = stl_conv_urdf("single_arm_v22_19cm","test")
        self.payload = 1.1
        self.payload_position = [0, 0, 0.04]
        self.motor_type_axis_2 = 44.7
        self.motor_type_axis_3 = 25.3
        self.model_select = "test"
        self.std_L2 = 26.754
        self.std_L3 = 23.246
        self.point_test_excel = "./xlsx/task_point_6dof_tested_d.xlsx"
        self.mission_time = 10
    def only_reachable_manipulability(model_select):
        pass
    def performance_evaluate(model_select, motor_type_axis_2, motor_type_axis_3):
        pass
        
if __name__ == '__main__':    # pragma nocover
    
    Traj_robot = RobotTraj()
    # print(Traj_robot)
    Traj_robot.robot_urdf.specified_generate_write_urdf(Traj_robot.std_L2, Traj_robot.std_L3)
    Traj_robot.robot.__init__() # 重製機器人
    Traj_robot.robot.payload(Traj_robot.payload, Traj_robot.payload_position)  # set payload
    if Traj_robot.model_select == "test":
        df = load_workbook(Traj_robot.point_test_excel)
        sheets = df.worksheets
        sheet1 = sheets[0]
        rows = sheet1.rows
        T_tmp = []
        T_traj = []
        ik_q_traj = []
        ratio_over = 0
        torque_over = 0
        num_torque = np.array([np.zeros(shape=6)])
        total_time = Traj_robot.mission_time
        # 采样间隔
        sample_interval = 0.2
        manipulability_index = []
        i = 0
        count = 0
        max_diff = []
        max_diff_tol = 0
        traj_time = []
        traj = []
        diff = np.array([np.zeros(shape=6)])
        total_time_traj = 0
        for row in rows:
            row_val = [col.value for col in row]
            T_tmp.append(SE3(row_val[0], row_val[1], row_val[2]) * SE3.RPY([np.deg2rad(row_val[3]), np.deg2rad(row_val[4]), np.deg2rad(row_val[5])]))
            # ik_q = Traj_robot.robot.ikine_min(T=T_tmp[i],qlim=True)
            ik_q = Traj_robot.robot.ikine_LMS(T=T_tmp[i])
            # ikine_min(qlim=True)

            if ik_q.success == True:
                count += 1
                T_traj.append(T_tmp[i])
                ik_q_traj.append(ik_q.q)
                manipulability_index.append(Traj_robot.robot.manipulability(q=ik_q.q))
                # print("ik_q.q",ik_q.q)
                if count >= 2: # 兩個點位以上開始計算
                    diff = np.abs(np.subtract(ik_q_traj[-2], ik_q_traj[-1]))
                    max_diff_tol = max_diff_tol + np.max(diff)
                    max_diff.append(np.max(diff))
                    # print(max_diff)
            i = i + 1
        for k in range(len(max_diff)):
            traj_time.append(max_diff[k] / max_diff_tol * total_time)
        
        
        traj_num = 0
        for m in range(len(max_diff)):
            # TODO: save traj
            excel_file_traj = Workbook()
            sheet_traj = excel_file_traj.active
            
            time_vector = np.linspace(0, traj_time[m], int(traj_time[m]/sample_interval) + 1)
            # print("time_vector:",time_vector)
            traj_tmp = Traj_robot.robot.jtraj(T_traj[m],T_traj[m+1],time_vector)
            
            # if np.amax(traj.sd) > 3.04:
            #     ratio_over = ratio_over + 1
            torque = Traj_robot.robot.rne(traj_tmp.s,traj_tmp.sd,traj_tmp.sdd)
            row = abs(torque[:,1]) # 取出第2行
            result_2 = row[row > Traj_robot.motor_type_axis_2] # 取出大于阈值的数字
            row = abs(torque[:,2]) # 取出第3行
            result_3 = row[row > Traj_robot.motor_type_axis_3] # 取出大于阈值的数字
            if len(result_2)>0 or len(result_3) >0:
                torque_over = torque_over + 1
            num_torque = np.append(num_torque, torque)
            
            # traj.append(traj_tmp)
            # TODO: add save traj to sheet
            Traj_robot.robot.plot(traj_tmp.s)
            # 迭代矩陣的每一個元素，並填入工作表中
            for k in range(len(traj_tmp.s)):
                # sheet_traj.cell(row=k+1, column=1).value = traj_tmp.t[k] + total_time_traj
                sheet_traj.cell(row=k+1, column=1).value = traj_tmp.t[k]

                # print("traj_tmp.s length",len(traj_tmp.s))
                # print("traj_tmp.t length",len(traj_tmp.t))
                
                print(traj_tmp.t[k] + total_time_traj)
                for l in range(len(traj_tmp.s[k])):
                    # TODO: 将第三、五、六轴的值加上负号
                    if l == 0: # 第一軸改角度
                        sheet_traj.cell(row=k+1, column=l+2).value = -(traj_tmp.s[k][l])
                    elif l == 1: # 第二軸改角度
                        sheet_traj.cell(row=k+1, column=l+2).value = -(traj_tmp.s[k][l]-np.deg2rad(90))
                    elif l == 2: # 第三軸改角度
                        sheet_traj.cell(row=k+1, column=l+2).value = (traj_tmp.s[k][l])
                    elif l == 4: # 第五軸改角度
                        sheet_traj.cell(row=k+1, column=l+2).value = -(traj_tmp.s[k][l]-np.deg2rad(90))
                    elif l == 5: # 第六軸改角度
                        sheet_traj.cell(row=k+1, column=l+2).value = traj_tmp.s[k][l]
                    else:
                        sheet_traj.cell(row=k+1, column=l+2).value = traj_tmp.s[k][l]
            total_time_traj += traj_tmp.t[-1]
            # print("total_time_traj:",total_time_traj)
            traj_num = traj_num + 1  # 要更改的数字
            new_str = "_{}_".format(traj_num)
            if not os.path.exists(curr_path + "/tested_state_traj/"):
                os.makedirs(curr_path + "/tested_state_traj/")
            file_name_traj = curr_path + "/tested_state_traj/tested_robot_traj" + new_str +".xlsx"
            excel_file_traj.save(file_name_traj)
        
        
        total_energy = 0
        for j in range(len(num_torque)):
            energy = abs(num_torque[j]) * sample_interval
            total_energy += energy
        
        print("traj_time",traj_time)
        print(total_energy)