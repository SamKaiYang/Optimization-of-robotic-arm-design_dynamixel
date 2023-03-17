from openpyxl import load_workbook
import matplotlib.pyplot as plt

###　test 4
# 讀取Excel檔案
workbook = load_workbook(filename='tested_state_traj_10_20230315-140359.xlsx')
worksheet = workbook.active
# 設置圖表
# fig, ax = plt.subplots(nrows=2, ncols=1, figsize=(6, 12), sharex=True)


# for i in range(7):
#     # 提取第i行資料
#     col_data = list(worksheet.iter_cols(min_row=1, max_row=46, min_col=i+1, max_col=i+1, values_only=True))[0]
#     # 获取 x 坐标和 y 坐标
#     # x = len(data[:]) # 46 筆資料
#     x = range(1, len(col_data)+1)
#     y = col_data
#     # 绘制折线图
#     ax[i].plot(x, y)

#     # 添加标题和轴标签
#     ax[i].set_title('Column {}'.format(i))
#     ax[i].set_xlabel('x axis label')
#     ax[i].set_ylabel('y axis label')
i = 0
col_data = list(worksheet.iter_cols(min_row=1, max_row=46, min_col=i+1, max_col=i+1, values_only=True))[0]
# 获取 x 坐标和 y 坐标
# x = len(data[:]) # 46 筆資料
x = range(1, len(col_data)+1)
y = col_data
# 绘制折线图
plt.plot(x, y)

# 添加标题和轴标签
# plt.title('Column {}'.format(i))
plt.title('speed limit')
plt.xlabel('times')
plt.ylabel('counts')
# 顯示圖表
plt.show()

i = 1
col_data = list(worksheet.iter_cols(min_row=1, max_row=46, min_col=i+1, max_col=i+1, values_only=True))[0]
# 获取 x 坐标和 y 坐标
# x = len(data[:]) # 46 筆資料
x = range(1, len(col_data)+1)
y = col_data
# 绘制折线图
plt.plot(x, y)

# 添加标题和轴标签
# plt.title('Column {}'.format(i))
plt.title('torque limit')
plt.xlabel('times')
plt.ylabel('counts')
# 顯示圖表
plt.show()

i = 2
col_data = list(worksheet.iter_cols(min_row=1, max_row=46, min_col=i+1, max_col=i+1, values_only=True))[0]
# 获取 x 坐标和 y 坐标
# x = len(data[:]) # 46 筆資料
x = range(1, len(col_data)+1)
y = col_data
# 绘制折线图
plt.plot(x, y)

# 添加标题和轴标签
# plt.title('Column {}'.format(i))
plt.title('power consumption')
plt.xlabel('times')
plt.ylabel('W')
# 顯示圖表
plt.show()

i = 3
col_data = list(worksheet.iter_cols(min_row=1, max_row=46, min_col=i+1, max_col=i+1, values_only=True))[0]
# 获取 x 坐标和 y 坐标
# x = len(data[:]) # 46 筆資料
x = range(1, len(col_data)+1)
y = col_data
# 绘制折线图
plt.plot(x, y)

# 添加标题和轴标签
# plt.title('Column {}'.format(i))
plt.title('reachability')
plt.xlabel('times')
plt.ylabel('score')
# 顯示圖表
plt.show()
i = 4
col_data = list(worksheet.iter_cols(min_row=1, max_row=46, min_col=i+1, max_col=i+1, values_only=True))[0]
# 获取 x 坐标和 y 坐标
# x = len(data[:]) # 46 筆資料
x = range(1, len(col_data)+1)
y = col_data
# 绘制折线图
plt.plot(x, y)

# 添加标题和轴标签
# plt.title('Column {}'.format(i))
plt.title('manipulability')
plt.xlabel('times')
plt.ylabel('index')
# 顯示圖表
plt.show()
# i = 5
# col_data = list(worksheet.iter_cols(min_row=1, max_row=46, min_col=i+1, max_col=i+1, values_only=True))[0]
# # 获取 x 坐标和 y 坐标
# # x = len(data[:]) # 46 筆資料
# x = range(1, len(col_data)+1)
# y = col_data
# # 绘制折线图
# plt.plot(x, y)

# # 添加标题和轴标签
# # plt.title('Column {}'.format(i))
# plt.title('manipulability')
# plt.xlabel('times')
# plt.ylabel('index')
# # 顯示圖表
# plt.show()

fig, ax = plt.subplots(nrows=5, ncols=1, figsize=(6, 18), sharex=True)


#     # 绘制折线图
#     ax[i].plot(x, y)

#     # 添加标题和轴标签
#     ax[i].set_title('Column {}'.format(i))
#     ax[i].set_xlabel('x axis label')
#     ax[i].set_ylabel('y axis label')
i = 0
col_data = list(worksheet.iter_cols(min_row=1, max_row=46, min_col=i+1, max_col=i+1, values_only=True))[0]
# 获取 x 坐标和 y 坐标
# x = len(data[:]) # 46 筆資料
x = range(1, len(col_data)+1)
y = col_data
# 绘制折线图
ax[i].plot(x, y)

# 添加标题和轴标签
# plt.title('Column {}'.format(i))
ax[i].set_title('speed limit')
ax[i].set_xlabel('times')
ax[i].set_ylabel('counts')
ax[i].set_xticks([0,5,10,15,20,25,30,35,40,45,50])
# ax[i].set_aspect('auto')
# ## 於軸上顯示較小的刻度
# ax[i].minorticks_on()
# ax[i].set_xticklabels(
#         labels=labels,
#         fontsize=5,
#         rotation=0)
i = 1
col_data = list(worksheet.iter_cols(min_row=1, max_row=46, min_col=i+1, max_col=i+1, values_only=True))[0]
# 获取 x 坐标和 y 坐标
# x = len(data[:]) # 46 筆資料
x = range(1, len(col_data)+1)
y = col_data
# 绘制折线图
ax[i].plot(x, y)

# 添加标题和轴标签
# plt.title('Column {}'.format(i))
ax[i].set_title('torque limit')
ax[i].set_xlabel('times')
ax[i].set_ylabel('counts')


i = 2
col_data = list(worksheet.iter_cols(min_row=1, max_row=46, min_col=i+1, max_col=i+1, values_only=True))[0]
# 获取 x 坐标和 y 坐标
# x = len(data[:]) # 46 筆資料
x = range(1, len(col_data)+1)
y = col_data
# 绘制折线图
ax[i].plot(x, y)

# 添加标题和轴标签
# plt.title('Column {}'.format(i))
ax[i].set_title('power consumption')
ax[i].set_xlabel('times')
ax[i].set_ylabel('W')

i = 3
col_data = list(worksheet.iter_cols(min_row=1, max_row=46, min_col=i+1, max_col=i+1, values_only=True))[0]
# 获取 x 坐标和 y 坐标
# x = len(data[:]) # 46 筆資料
x = range(1, len(col_data)+1)
y = col_data
# 绘制折线图
ax[i].plot(x, y)

# 添加标题和轴标签
# plt.title('Column {}'.format(i))
ax[i].set_title('reachability')
ax[i].set_xlabel('times')
ax[i].set_ylabel('score')

i = 4
col_data = list(worksheet.iter_cols(min_row=1, max_row=46, min_col=i+1, max_col=i+1, values_only=True))[0]
# 获取 x 坐标和 y 坐标
# x = len(data[:]) # 46 筆資料
x = range(1, len(col_data)+1)
y = col_data
# 绘制折线图
ax[i].plot(x, y)

# 添加标题和轴标签
# plt.title('Column {}'.format(i))
ax[i].set_title('manipulability')
ax[i].set_xlabel('times')
ax[i].set_ylabel('index')
# 顯示圖表
plt.show()